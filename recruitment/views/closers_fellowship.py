"""
Views for Closers Fellowship applications submitted from the website form.
"""

from datetime import datetime
from io import BytesIO

from django.contrib import messages
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from django.views.decorators.http import require_http_methods
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from horilla.decorators import any_permission_required, login_required
from recruitment.models import ClosersFellowshipApplication
from recruitment.views.paginator_qry import paginator_qry


def _filter_params(request):
    return {
        "search": (request.GET.get("search") or "").strip(),
        "seat": (request.GET.get("seat") or "").strip(),
        "campaign": (request.GET.get("campaign") or "").strip(),
        "utm_source": (request.GET.get("utm_source") or "").strip(),
        "date_from": (request.GET.get("date_from") or "").strip(),
        "date_to": (request.GET.get("date_to") or "").strip(),
    }


def _apply_filters(queryset, params):
    search = params["search"]
    if search:
        queryset = queryset.filter(
            Q(full_name__icontains=search)
            | Q(email__icontains=search)
            | Q(phone__icontains=search)
        )
    if params["seat"]:
        queryset = queryset.filter(seat=params["seat"])
    if params["campaign"]:
        queryset = queryset.filter(utm_campaign=params["campaign"])
    if params["utm_source"]:
        queryset = queryset.filter(utm_source=params["utm_source"])
    if params["date_from"]:
        try:
            date_from = datetime.strptime(params["date_from"], "%Y-%m-%d").date()
            queryset = queryset.filter(created_at__date__gte=date_from)
        except ValueError:
            pass
    if params["date_to"]:
        try:
            date_to = datetime.strptime(params["date_to"], "%Y-%m-%d").date()
            queryset = queryset.filter(created_at__date__lte=date_to)
        except ValueError:
            pass
    return queryset


def _distinct_values(field_name):
    return (
        ClosersFellowshipApplication.objects.exclude(**{f"{field_name}__isnull": True})
        .exclude(**{field_name: ""})
        .values_list(field_name, flat=True)
        .distinct()
        .order_by(field_name)
    )


@login_required
@any_permission_required(
    perms=[
        "recruitment.view_closersfellowshipapplication",
        "recruitment.view_candidate",
    ]
)
def closers_fellowship_list(request):
    """List all Closers Fellowship applications."""
    params = _filter_params(request)
    applications = _apply_filters(
        ClosersFellowshipApplication.objects.all().order_by("-created_at"),
        params,
    )
    query_params = request.GET.copy()
    query_params.pop("page", None)
    has_filters = any(params.values())
    return render(
        request,
        "closers_fellowship/closers_fellowship_list.html",
        {
            "data": paginator_qry(applications, request.GET.get("page")),
            "filters": params,
            "has_filters": has_filters,
            "query_string": query_params.urlencode(),
            "seat_options": _distinct_values("seat"),
            "campaign_options": _distinct_values("utm_campaign"),
            "utm_source_options": _distinct_values("utm_source"),
        },
    )


@login_required
@any_permission_required(
    perms=[
        "recruitment.view_closersfellowshipapplication",
        "recruitment.view_candidate",
    ]
)
def closers_fellowship_detail(request, app_id):
    """View a single Closers Fellowship application."""
    application = get_object_or_404(ClosersFellowshipApplication, id=app_id)
    return render(
        request,
        "closers_fellowship/closers_fellowship_detail.html",
        {"application": application},
    )


@login_required
@any_permission_required(
    perms=[
        "recruitment.delete_closersfellowshipapplication",
        "recruitment.delete_candidate",
    ]
)
@require_http_methods(["POST"])
def closers_fellowship_delete(request, app_id):
    """Delete a Closers Fellowship application."""
    application = get_object_or_404(ClosersFellowshipApplication, id=app_id)
    name = application.full_name
    application.delete()
    messages.success(request, _("Application for %(name)s deleted.") % {"name": name})
    return redirect(reverse("closers-fellowship-list"))


def _format_applied_at(dt):
    if not dt:
        return ""
    local_dt = timezone.localtime(dt)
    return local_dt.strftime("%b %d, %Y %I:%M %p")


@login_required
@any_permission_required(
    perms=[
        "recruitment.view_closersfellowshipapplication",
        "recruitment.view_candidate",
    ]
)
def closers_fellowship_export(request):
    """Export Closers Fellowship applications to Excel (respects active filters)."""
    params = _filter_params(request)
    applications = _apply_filters(
        ClosersFellowshipApplication.objects.all().order_by("-created_at"),
        params,
    )

    headers = [
        str(_("Full Name")),
        str(_("Email")),
        str(_("Phone / WhatsApp")),
        str(_("Seat")),
        str(_("LinkedIn or Portfolio")),
        str(_("Q1")),
        str(_("Q2")),
        str(_("Q3")),
        str(_("Campaign")),
        str(_("Ad")),
        str(_("Adset")),
        str(_("UTM Source")),
        str(_("UTM Medium")),
        str(_("Applied At")),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Closers Fellowship"

    header_fill = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, app in enumerate(applications, 2):
        row_values = [
            app.full_name,
            app.email,
            app.phone or "",
            app.seat or "",
            app.linkedin_portfolio or "",
            app.answer_q1 or "",
            app.answer_q2 or "",
            app.answer_q3 or "",
            app.utm_campaign or "",
            app.utm_content or "",
            app.utm_term or "",
            app.utm_source or "",
            app.utm_medium or "",
            _format_applied_at(app.created_at),
        ]
        for col, value in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = len(str(headers[col - 1]))
        for row in range(2, ws.max_row + 1):
            value = ws.cell(row=row, column=col).value
            if value:
                max_len = max(max_len, min(len(str(value)), 60))
        ws.column_dimensions[letter].width = min(max_len + 2, 50)

    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"closers_fellowship_{timezone.localtime().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
