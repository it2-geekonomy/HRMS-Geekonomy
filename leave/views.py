"""
views.py
"""

import ast
import contextlib
import json
from collections import defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO
from urllib.parse import parse_qs, unquote

import pandas as pd
from django.apps import apps
from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import ProtectedError, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils.encoding import force_str
from django.utils.translation import gettext as __
from django.utils.translation import gettext_lazy as _
from django.views.decorators.http import require_http_methods
from xhtml2pdf import pisa

from base.filters import PenaltyFilter
from base.forms import PenaltyAccountForm
from base.methods import (
    choosesubordinates,
    closest_numbers,
    eval_validate,
    export_data,
    filtersubordinates,
    get_key_instances,
    get_pagination,
    is_reportingmanager,
    sortby,
)
from base.models import CompanyLeaves, Holidays, PenaltyAccounts
from employee.models import Employee
from horilla.decorators import (
    hx_request_required,
    logger,
    login_required,
    manager_can_enter,
    owner_can_enter,
    permission_required,
)
from horilla.group_by import group_by_queryset
from horilla.horilla_settings import DYNAMIC_URL_PATTERNS
from horilla.methods import get_horilla_model_class, remove_dynamic_url
from leave.decorators import *
from leave.filters import *
from leave.forms import *
from leave.methods import (
    attendance_days,
    calculate_requested_days,
    can_user_approve_leave_request,
    company_leave_dates_list,
    computed_balance_for_validation,
    filter_conditional_leave_request,
    get_init_days_and_reset_for_assign,
    holiday_dates_list,
    probation_leave_days_in_month,
)
from leave.models import *
from leave.models import leave_requested_dates
from leave.threading import LeaveMailSendThread, send_leave_approve_reject_email
from notifications.signals import notify


def generate_error_report(error_list, error_data, file_name):
    """
    Function used to generate error excle file for imported datas
    """
    for item in error_list:
        for key, value in error_data.items():
            if key in item:
                value.append(item[key])
            else:
                value.append(None)
    keys_to_remove = [
        key for key, value in error_data.items() if all(v is None for v in value)
    ]
    for key in keys_to_remove:
        del error_data[key]
    data_frame = pd.DataFrame(error_data, columns=error_data.keys())
    styled_data_frame = data_frame.style.applymap(
        lambda x: "text-align: center", subset=pd.IndexSlice[:, :]
    )
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = f'attachment; filename="{file_name}"'
    writer = pd.ExcelWriter(response, engine="xlsxwriter")
    styled_data_frame.to_excel(writer, index=False, sheet_name="Sheet1")
    worksheet = writer.sheets["Sheet1"]
    worksheet.set_column("A:Z", 30)
    writer.close()

    def get_error_sheet(request):
        remove_dynamic_url(path_info)
        return response

    from leave.urls import path, urlpatterns

    path_info = f"error-sheet-{uuid.uuid4()}"
    urlpatterns.append(path(path_info, get_error_sheet, name=path_info))
    DYNAMIC_URL_PATTERNS.append(path_info)
    path_info = f"leave/{path_info}"
    return path_info


@login_required
@permission_required("leave.add_leavetype")
def leave_type_creation(request):
    """
    function used to create leave type.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET : return leave type creation template
    POST : return leave view
    """
    form = LeaveTypeForm()
    if request.method == "POST":
        form = LeaveTypeForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, _("New leave type Created.."))
            return redirect(leave_type_view)
    return render(request, "leave/leave_type/leave_type_creation.html", {"form": form})


def paginator_qry(qryset, page_number):
    """
    function used to paginate query set
    """
    paginator = Paginator(qryset, get_pagination())
    qryset = paginator.get_page(page_number)
    return qryset


@login_required
@permission_required("leave.view_leavetype")
def leave_type_view(request):
    """
    function used to view leave type.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET : return leave type template
    """

    queryset = LeaveType.objects.all().exclude(is_compensatory_leave=True)
    if (
        LeaveGeneralSetting.objects.first()
        and LeaveGeneralSetting.objects.first().compensatory_leave
    ):
        queryset = LeaveType.objects.all()
    page_number = request.GET.get("page")
    page_obj = paginator_qry(queryset, page_number)
    previous_data = request.GET.urlencode()
    leave_type_filter = LeaveTypeFilter()
    requests_ids = json.dumps(list(queryset.values_list("id", flat=True)))
    if not queryset.exists():
        template_name = "leave/leave_type/leave_type_empty_view.html"
    else:
        template_name = "leave/leave_type/leave_type_view.html"
    return render(
        request,
        template_name,
        {
            "leave_types": page_obj,
            "form": leave_type_filter.form,
            "pd": previous_data,
            "requests_ids": requests_ids,
        },
    )


@login_required
@hx_request_required
@manager_can_enter("leave.view_leavetype")
def leave_type_individual_view(request, id):
    """
    function used to view one leave type.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET : return one leave type view template
    """
    leave_type = LeaveType.find(id)
    requests_ids_json = request.GET.get("instances_ids")
    compensatory = request.GET.get("compensatory")
    context = {"leave_type": leave_type, "compensatory": compensatory}
    if requests_ids_json:
        requests_ids = json.loads(requests_ids_json)
        previous_id, next_id = closest_numbers(requests_ids, id)
        context["previous"] = previous_id
        context["next"] = next_id
        context["requests_ids"] = requests_ids_json
    return render(request, "leave/leave_type/leave_type_individual_view.html", context)


@login_required
@hx_request_required
@permission_required("leave.view_leavetype")
def leave_type_filter(request):
    """
    function used to filter view leave type.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET : return leave types template
    """
    queryset = LeaveType.objects.all()
    page_number = request.GET.get("page")
    leave_type_filter = LeaveTypeFilter(request.GET, queryset).qs
    page_obj = paginator_qry(leave_type_filter, page_number)
    previous_data = request.GET.urlencode()
    requests_ids = json.dumps(list(leave_type_filter.values_list("id", flat=True)))
    data_dict = parse_qs(previous_data)
    get_key_instances(LeaveType, data_dict)
    return render(
        request,
        "leave/leave_type/leave_types.html",
        {
            "leave_types": page_obj,
            "pd": previous_data,
            "filter_dict": data_dict,
            "requests_ids": requests_ids,
        },
    )


@login_required
@permission_required("leave.change_leavetype")
def leave_type_update(request, id, **kwargs):
    """
    function used to update leave type.

    request (HttpRequest): The HTTP request object.
    id : leave type id

    Returns:
    GET : return leave type update template
    POST : return leave type view
    """
    try:
        leave_type = LeaveType.objects.get(id=id)
    except (LeaveType.DoesNotExist, OverflowError, ValueError):
        messages.error(request, _("Leave type not found"))
        return redirect(leave_type_view)
    form = UpdateLeaveTypeForm(instance=leave_type)
    compensatory = request.GET.get("compensatory")
    redirect_url = leave_type_view
    if compensatory:
        redirect_url = compensatory_leave_settings_view
    if request.method == "POST":
        form_data = UpdateLeaveTypeForm(
            request.POST, request.FILES, instance=leave_type
        )
        if form_data.is_valid():
            form_data.save()
            messages.success(request, _("Leave type is updated successfully.."))
            return redirect(redirect_url)
    return render(
        request,
        "leave/leave_type/leave_type_update.html",
        {"form": form, "compensatory": compensatory},
    )


@login_required
@permission_required("leave.delete_leavetype")
def leave_type_delete(request, obj_id):
    """
    function used to delete leave type.

    Parameters:
    request (HttpRequest): The HTTP request object.
    id : leave type id

    Returns:
    GET : return leave type view template
    """
    try:
        LeaveType.objects.get(id=obj_id).delete()
        messages.success(request, _("Leave type deleted successfully.."))
    except (LeaveType.DoesNotExist, OverflowError, ValueError):
        messages.error(request, _("Leave type not found."))
    except ProtectedError as e:
        models_verbose_name_sets = set()
        for obj in e.protected_objects:
            models_verbose_name_sets.add(__(obj._meta.verbose_name))
        models_verbose_name_str = (",").join(models_verbose_name_sets)
        messages.error(
            request,
            _(
                "This leave types are already in use for {}".format(
                    models_verbose_name_str
                )
            ),
        )
    if request.META.get("HTTP_HX_REQUEST") == "true":
        if request.META.get("HTTP_HX_TARGET") == "objectDetailsModalTarget":
            instances_ids = request.GET.get("instances_ids")
            instances_list = json.loads(instances_ids)
            if obj_id in instances_list:
                instances_list.remove(obj_id)
            previous_instance, next_instance = closest_numbers(
                json.loads(instances_ids), obj_id
            )
            return redirect(
                f"/leave/leave-type-individual-view/{next_instance}?instances_ids={instances_list}"
            )
        return redirect(f"/leave/type-filter?{request.GET.urlencode()}")
    return redirect(leave_type_view)


@login_required
@hx_request_required
@manager_can_enter("leave.add_leaverequest")
def get_employee_leave_types(request):
    employee_id = request.GET.get("employee_id")
    form = (
        LeaveRequestUpdationForm()
        if request.GET.get("form")
        and request.GET.get("form") == "LeaveRequestUpdationForm"
        else LeaveRequestCreationForm()
    )

    if employee_id:
        employee = get_object_or_404(Employee, id=employee_id)
        assigned_leave_types = LeaveType.objects.filter(
            id__in=employee.available_leave.values_list("leave_type_id", flat=True)
        )
        form.fields["leave_type_id"].queryset = assigned_leave_types
    else:
        form.fields["leave_type_id"].queryset = LeaveType.objects.none()

    leave_type_field_html = render_to_string(
        "leave/leave_request/leave_type_field.html",
        {
            "form": form,
            "field_name": "leave_type_id",
            "field": form.fields["leave_type_id"],
        },
    )
    return HttpResponse(leave_type_field_html)


def multiple_approvals_check(id):
    approvals = LeaveRequestConditionApproval.objects.filter(leave_request_id=id)
    requested_query = approvals.filter(is_approved=False).order_by("sequence")
    approved_query = approvals.filter(is_approved=True).order_by("sequence")
    managers = []
    for manager in approvals:
        managers.append(manager.manager_id)
    if approvals.exists():
        result = {
            "managers": managers,
            "approved": approved_query,
            "requested": requested_query,
            "approvals": approvals,
        }
    else:
        result = False
    return result


@login_required
@hx_request_required
@manager_can_enter("leave.add_leaverequest")
def leave_request_creation(request, type_id=None, emp_id=None):
    """
    function used to create leave request.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET : return leave request form template
    POST : return leave request view
    """
    referer_parts = [
        part for part in request.META.get("HTTP_REFERER").split("/") if part != ""
    ]
    if request.GET.urlencode().startswith("pd="):
        previous_data = unquote(request.GET.urlencode())[len("pd=") :]
    else:
        request_copy = request.GET.copy()
        if "confirm" in request_copy:
            request_copy.pop("confirm")
        previous_data = request_copy.urlencode()

    form = LeaveRequestCreationForm()
    if request:
        employee_qs = form.fields["employee_id"].queryset
        post_emp_id = request.POST.get("employee_id")
        employee = employee_qs.filter(id=post_emp_id).first() or (
            request.user.employee_get
            if request.user.employee_get in employee_qs
            else employee_qs.first()
        )

        if employee:
            leave_type_ids = employee.available_leave.values_list(
                "leave_type_id", flat=True
            )
            assigned_leave_types = LeaveType.objects.filter(id__in=leave_type_ids)

            form.fields["leave_type_id"].queryset = assigned_leave_types

    if type_id and emp_id:
        initial_data = {
            "leave_type_id": type_id,
            "employee_id": emp_id,
        }
        form = LeaveRequestCreationForm(initial=initial_data)

    form = choosesubordinates(request, form, "leave.add_leaverequest")
    if request.method == "POST":
        form = LeaveRequestCreationForm(request.POST, request.FILES)
        # Set the queryset again on the bound form for the validation
        form.fields["leave_type_id"].queryset = assigned_leave_types
        form = choosesubordinates(request, form, "leave.add_leaverequest")
        if form.is_valid():
            leave_request = form.save(commit=False)
            save = True

            if leave_request.leave_type_id.require_approval == "no":
                # Deduct only from this request's leave type (e.g. Casual Leave never deducts from EL).
                employee_id = leave_request.employee_id
                leave_type_id = leave_request.leave_type_id
                available_leave = AvailableLeave.objects.get(
                    leave_type_id=leave_type_id, employee_id=employee_id
                )
                leave_request.created_by = request.user.employee_get
                leave_request.save()
                if leave_request.requested_days > available_leave.available_days:
                    leave = (
                        leave_request.requested_days - available_leave.available_days
                    )
                    leave_request.approved_available_days = (
                        available_leave.available_days
                    )
                    available_leave.available_days = 0
                    available_leave.carryforward_days = (
                        available_leave.carryforward_days - leave
                    )
                    leave_request.approved_carryforward_days = leave
                else:
                    available_leave.available_days = (
                        available_leave.available_days - leave_request.requested_days
                    )
                    leave_request.approved_available_days = leave_request.requested_days
                leave_request.status = "approved"
            if save:
                leave_request.created_by = request.user.employee_get
                leave_request.save()
                try:
                    available_leave.save()
                except:
                    pass

                if multiple_approvals_check(leave_request.id):
                    conditional_requests = multiple_approvals_check(leave_request.id)
                    managers = []
                    for manager in conditional_requests["managers"]:
                        managers.append(manager.employee_user_id)
                    with contextlib.suppress(Exception):
                        notify.send(
                            request.user.employee_get,
                            recipient=managers[0],
                            verb="You have a new leave request to validate.",
                            verb_ar="لديك طلب إجازة جديد يجب التحقق منه.",
                            verb_de="Sie haben eine neue Urlaubsanfrage zur Validierung.",
                            verb_es="Tiene una nueva solicitud de permiso que debe validar.",
                            verb_fr="Vous avez une nouvelle demande de congé à valider.",
                            icon="people-circle",
                            redirect=f"/leave/request-view?id={leave_request.id}",
                        )

                mail_thread = LeaveMailSendThread(
                    request, leave_request, type="request"
                )
                mail_thread.start()
                messages.success(request, _("Leave request created successfully.."))
                with contextlib.suppress(Exception):
                    notify.send(
                        request.user.employee_get,
                        recipient=leave_request.employee_id.employee_work_info.reporting_manager_id.employee_user_id,
                        verb=f"New leave request created for {leave_request.employee_id}.",
                        verb_ar=f"تم إنشاء طلب إجازة جديد لـ {leave_request.employee_id}.",
                        verb_de=f"Neuer Urlaubsantrag erstellt für {leave_request.employee_id}.",
                        verb_es=f"Nueva solicitud de permiso creada para {leave_request.employee_id}.",
                        verb_fr=f"Nouvelle demande de congé créée pour {leave_request.employee_id}.",
                        icon="people-circle",
                        redirect=reverse("request-view") + f"?id={leave_request.id}",
                    )
                form = LeaveRequestCreationForm()
                if referer_parts[-2] == "employee-view":
                    return HttpResponse("<script>window.location.reload();</script>")

            leave_requests = LeaveRequest.objects.all()
            if len(leave_requests) == 1:
                return HttpResponse("<script>window.location.reload()</script>")
    referrer = request.META.get("HTTP_REFERER", "")
    referrer = "/" + "/".join(referrer.split("/")[3:])
    if referrer == "/":
        hx_url = reverse("leave-request-and-approve")
        hx_target = "#leaveApproveCardBody"
    else:
        hx_url = "/leave/request-filter?"
        hx_target = "#leaveRequest"
    return render(
        request,
        "leave/leave_request/leave_request_form.html",
        {
            "form": form,
            "pd": previous_data,
            "hx_url": hx_url,
            "hx_target": hx_target,
        },
    )


@login_required
@manager_can_enter("leave.view_leaverequest")
def leave_request_view(request):
    """
    function used to view leave request.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET : return leave request view template
    """
    # Default: show only "requested" leaves; when user applies filter (status in GET), use filter.
    get_data = request.GET.copy()
    if "status" not in get_data:
        get_data.setdefault("status", "requested")
    queryset = LeaveRequestFilter(get_data).qs.order_by("-id").distinct()
    multiple_approvals = filter_conditional_leave_request(request).distinct()
    normal_requests = filtersubordinates(request, queryset, "leave.view_leaverequest")

    if not request.user.is_superuser:
        multi_approve_requests = LeaveRequestConditionApproval.objects.filter(
            is_approved=False, is_rejected=False
        )

        multi_ids = [request.leave_request_id.id for request in multi_approve_requests]

        # Create a new list excluding leave requests with IDs in multi_ids
        normal_requests = [
            leave.id for leave in normal_requests if leave.id not in multi_ids
        ]

        # Convert the list of IDs back to a queryset
        normal_requests = LeaveRequest.objects.filter(id__in=normal_requests).distinct()

    queryset = normal_requests | multiple_approvals
    page_number = request.GET.get("page")
    page_obj = paginator_qry(queryset, page_number)
    leave_request_filter = LeaveRequestFilter(get_data)

    # Fetching leave requests
    leave_requests = queryset

    leave_requests_with_interview = []
    if apps.is_installed("recruitment"):
        for leave_request in leave_requests:

            # Fetch interviews for the employee within the requested leave period
            InterviewSchedule = get_horilla_model_class(
                app_label="recruitment", model="interviewschedule"
            )

            interviews = InterviewSchedule.objects.filter(
                employee_id=leave_request.employee_id,
                interview_date__range=[
                    leave_request.start_date,
                    leave_request.end_date,
                ],
            )
            if interviews:
                # If interview exists then adding the leave request to the list
                leave_requests_with_interview.append(leave_request)

    requests = queryset.filter(status="requested").count()
    requests_ids = json.dumps(list(page_obj.object_list.values_list("id", flat=True)))
    approved_requests = queryset.filter(status="approved").count()
    rejected_requests = queryset.filter(status="cancelled").count()
    previous_data = request.GET.urlencode()
    data_dict = parse_qs(previous_data)
    get_key_instances(LeaveRequest, data_dict)
    return render(
        request,
        "leave/leave_request/request_view.html",
        {
            "leave_requests": page_obj,
            "pd": previous_data,
            "form": leave_request_filter.form,
            "requests": requests,
            "approved_requests": approved_requests,
            "rejected_requests": rejected_requests,
            "gp_fields": LeaveRequestReGroup.fields,
            "requests_ids": requests_ids,
            "current_date": date.today(),
            "filter_dict": data_dict,
            "leave_requests_with_interview": leave_requests_with_interview,
        },
    )


@login_required
@manager_can_enter("leave.view_leaverequest")
def leave_requests_export(request):
    if request.META.get("HTTP_HX_REQUEST") == "true":
        excel_column = LeaveRequestExportForm()
        export_filter = LeaveRequestFilter()
        context = {
            "excel_column": excel_column,
            "export_filter": export_filter.form,
        }

        return render(
            request,
            "leave/leave_request/leave_requests_export_filter.html",
            context=context,
        )
    return export_data(
        request=request,
        model=LeaveRequest,
        filter_class=LeaveRequestFilter,
        form_class=LeaveRequestExportForm,
        file_name="Leave_requests",
        perm="leave.view_leaverequest",
    )


def generate_leave_request_pdf(template_path, context, html=False):
    """
    Generate a PDF file from an HTML template and context data.

    Args:
        template_path (str): The path to the HTML template.
        context (dict): The context data to render the template.
        html (bool): If True, return raw HTML instead of a PDF.

    Returns:
        HttpResponse: A response with the generated PDF file or raw HTML.
    """
    try:
        html_content = render_to_string(template_path, context)

        if html:
            return HttpResponse(html_content)

        result = BytesIO()
        pdf_status = pisa.CreatePDF(src=html_content, dest=result)

        if pdf_status.err:
            logger.error("Error creating PDF")
            return HttpResponse("Error generating PDF", status=500)

        response = HttpResponse(result.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = 'inline; filename="leave_request.pdf"'
        return response

    except Exception as e:
        logger.exception("Error generating PDF")
        return HttpResponse(f"Error generating PDF: {str(e)}", status=500)


@login_required
@manager_can_enter("leave.view_leaverequest")
def create_leave_report(request):
    """
    Generate a Leave Report as a PDF and return it in an HttpResponse.

    Args:
        request (HttpRequest): The request object.

    Returns:
        HttpResponse: A response containing the PDF content.
    """
    employee_data = {}
    company_id = request.session.get("selected_company")
    if company_id == "all" or not company_id:
        company = Company.objects.all()
    else:
        company = Company.objects.filter(id=company_id).first()

    leave_requests = LeaveRequest.objects.filter(status="approved").select_related(
        "employee_id", "leave_type_id"
    )
    used_days_map = defaultdict(float)
    leave_request_map = defaultdict(list)

    for lreq in leave_requests:
        key = (
            lreq.employee_id.id,
            lreq.leave_type_id.id if lreq.leave_type_id else None,
        )
        used_days_map[key] += lreq.requested_days
        leave_request_map[lreq.employee_id.id].append(lreq)

    employees = Employee.objects.all()

    for employee in employees:
        employee_id = employee.id
        emp_data = {
            "employee": employee,
            "total_leave_days": 0,
            "used_leave_days": 0,
            "remaining_leave_days": 0,
            "leave_requests": leave_request_map.get(employee_id, []),
            "leave_types_counted": set(),
            "new_hire": False,
        }

        if employee.employee_work_info:
            hire_date = employee.employee_work_info.date_joining
            if hire_date and (date.today() - hire_date) <= timedelta(days=365):
                emp_data["new_hire"] = True

        assigned_leave_types = LeaveType.objects.filter(
            id__in=employee.available_leave.values_list("leave_type_id", flat=True)
        )

        for leave_type in assigned_leave_types:
            leave_type_id = leave_type.id

            if leave_type_id in emp_data["leave_types_counted"]:
                continue

            emp_data["leave_types_counted"].add(leave_type_id)

            total_days = leave_type.total_days or 0
            emp_data["total_leave_days"] += total_days

            used_days = used_days_map.get((employee_id, leave_type_id), 0)
            emp_data["used_leave_days"] += used_days

        emp_data["remaining_leave_days"] = (
            emp_data["total_leave_days"] - emp_data["used_leave_days"]
        )

        sorted_reqs = sorted(
            emp_data["leave_requests"],
            key=lambda x: (x.end_date - x.start_date).days,
            reverse=True,
        )
        for i in range(3):
            if i < len(sorted_reqs):
                emp_data[f"period{i+1}_start"] = sorted_reqs[i].start_date
                emp_data[f"period{i+1}_end"] = sorted_reqs[i].end_date
            else:
                emp_data[f"period{i+1}_start"] = ""
                emp_data[f"period{i+1}_end"] = ""

        employee_data[employee_id] = emp_data

    final_employee_data = list(employee_data.values())
    final_employee_data.sort(key=lambda x: x["employee"].get_full_name())

    context = {
        "employee_data": final_employee_data,
        "company_data": company,
        "report_creation_date": date.today(),
        "request": request,
    }

    template_path = "leave/leave_request/leave_request_pdf.html"
    return generate_leave_request_pdf(template_path, context=context, html=False)


@login_required
@hx_request_required
# @manager_can_enter("leave.view_leaverequest")
def leave_request_filter(request):
    """
    function used to filter leave request.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET : return leave request view template
    """
    previous_data = request.GET.urlencode()
    # Default: show only "requested" leaves; when user applies filter (status in GET), use filter.
    get_data = request.GET.copy()
    if "status" not in get_data:
        get_data.setdefault("status", "requested")
    queryset = LeaveRequestFilter(get_data).qs.order_by("-id")

    # Fetching leave requests
    leave_requests = queryset

    leave_requests_with_interview = []
    if apps.is_installed("recruitment"):
        for leave_request in leave_requests:

            # Fetch interviews for the employee within the requested leave period
            InterviewSchedule = get_horilla_model_class(
                app_label="recruitment", model="interviewschedule"
            )

            interviews = InterviewSchedule.objects.filter(
                employee_id=leave_request.employee_id,
                interview_date__range=[
                    leave_request.start_date,
                    leave_request.end_date,
                ],
            )
            if interviews:
                # If interview exists then adding the leave request to the list
                leave_requests_with_interview.append(leave_request)

    field = request.GET.get("field")
    multiple_approvals = filter_conditional_leave_request(request)
    queryset = filtersubordinates(request, queryset, "leave.view_leaverequest")

    if not request.user.is_superuser:
        multi_approve_requests = LeaveRequestConditionApproval.objects.filter(
            is_approved=False, is_rejected=False
        )

        multi_ids = [request.leave_request_id.id for request in multi_approve_requests]

        # Create a new list excluding leave requests with IDs in multi_ids
        queryset = [leave.id for leave in queryset if leave.id not in multi_ids]

        # Convert the list of IDs back to a queryset
        queryset = LeaveRequest.objects.filter(id__in=queryset)

    queryset = queryset.distinct()
    multiple_approvals = multiple_approvals.distinct()

    queryset = queryset | multiple_approvals
    leave_request_filter = LeaveRequestFilter(request.GET, queryset).qs
    page_number = request.GET.get("page")
    template = ("leave/leave_request/leave_requests.html",)
    if request.GET.get("sortby"):
        leave_request_filter = sortby(request, leave_request_filter, "sortby")

    if field != "" and field is not None:
        leave_request_filter = group_by_queryset(
            leave_request_filter, field, request.GET.get("page"), "page"
        )
        list_values = [entry["list"] for entry in leave_request_filter]
        id_list = []
        for value in list_values:
            for instance in value.object_list:
                id_list.append(instance.id)

        requests_ids = json.dumps(list(id_list))
        template = "leave/leave_request/group_by.html"

    else:
        leave_request_filter = paginator_qry(
            leave_request_filter, request.GET.get("page")
        )
        requests_ids = json.dumps(
            [instance.id for instance in leave_request_filter.object_list]
        )

    data_dict = []

    if not request.GET.get("dashboard"):
        data_dict = parse_qs(previous_data)
        get_key_instances(LeaveRequest, data_dict)

    if "status" in data_dict:
        status_list = data_dict["status"]
        if len(status_list) > 1:
            data_dict["status"] = [status_list[-1]]
    return render(
        request,
        template,
        {
            "leave_requests": leave_request_filter,
            "pd": previous_data,
            "filter_dict": data_dict,
            "field": field,
            "dashboard": request.GET.get("dashboard"),
            "requests_ids": requests_ids,
            "current_date": date.today(),
            "leave_requests_with_interview": leave_requests_with_interview,
        },
    )


@login_required
@hx_request_required
@manager_can_enter("leave.change_leaverequest")
def leave_request_update(request, id):
    """
    function used to update leave request.
    """
    leave_request = LeaveRequest.objects.get(id=id)
    if request.method == "POST":
        form = LeaveRequestUpdationForm(
            request.POST, request.FILES, instance=leave_request
        )
        form = choosesubordinates(request, form, "leave.add_leaverequest")
        if form.is_valid():
            leave_request = form.save(commit=False)
            save = True

            if save:
                leave_request.save()
                messages.success(request, _("Leave request is updated successfully.."))
                with contextlib.suppress(Exception):
                    notify.send(
                        request.user.employee_get,
                        recipient=leave_request.employee_id.employee_work_info.reporting_manager_id.employee_user_id,
                        verb=f"Leave request updated for {leave_request.employee_id}.",
                        verb_ar=f"تم تحديث طلب الإجازة لـ {leave_request.employee_id}.",
                        verb_de=f"Urlaubsantrag aktualisiert für {leave_request.employee_id}.",
                        verb_es=f"Solicitud de permiso actualizada para {leave_request.employee_id}.",
                        verb_fr=f"Demande de congé mise à jour pour {leave_request.employee_id}.",
                        icon="people-circle",
                        redirect=reverse("request-view") + f"?id={leave_request.id}",
                    )
                response = render(
                    request,
                    "leave/leave_request/request_update_form.html",
                    {"form": form, "id": id},
                )
                return HttpResponse(
                    response.content.decode("utf-8")
                    + "<script>location.reload();</script>"
                )
    else:
        form = LeaveRequestUpdationForm(instance=leave_request)
        form = choosesubordinates(request, form, "leave.add_leaverequest")

    return render(
        request,
        "leave/leave_request/request_update_form.html",
        {
            "form": form,
            "id": id,
        },
    )


@login_required
@hx_request_required
@manager_can_enter("leave.delete_leaverequest")
def leave_request_delete(request, id):
    """
    function used to delete leave request.

    Parameters:
    request (HttpRequest): The HTTP request object.
    id : leave request id

    Returns:
    GET : return leave request view template
    """
    previous_data = request.GET.urlencode()
    try:
        leave_request = LeaveRequest.objects.get(id=id)
        messages.success(request, _("Leave request deleted successfully.."))
        leave_request.delete()
    except (LeaveRequest.DoesNotExist, OverflowError, ValueError):
        messages.error(request, _("Leave request not found."))
    except ProtectedError:
        messages.error(request, _("Related entries exists"))
    hx_target = request.META.get("HTTP_HX_TARGET", None)
    if hx_target == "leaveRequest":
        leave_requests = LeaveRequest.objects.all()
        if leave_requests.exists():
            return redirect(f"/leave/request-filter?{previous_data}")
        else:
            return HttpResponse("<script>window.location.reload();</script>")
    return redirect(leave_request_view)


@login_required
@manager_can_enter("leave.change_leaverequest")
def leave_request_approve(request, id, emp_id=None):
    """
    function used to approve a leave request.

    Parameters:
    request (HttpRequest): The HTTP request object.
    id : leave request id
    emp_id : employee id if the approval operation comes from "/employee/employee-view/{employee_id}/" template.


    Returns:
    GET : If `emp_id` is provided, it returns to the "/employee/employee-view/{employee_id}/" template after approval.
          Otherwise, it returns to the default leave request view template.
    """
    leave_request = LeaveRequest.objects.get(id=id)
    if not can_user_approve_leave_request(request.user, leave_request):
        messages.error(
            request,
            _(
                "You cannot approve your own leave request. Your reporting manager must approve it."
            ),
        )
        if emp_id is not None:
            return redirect(f"/employee/employee-view/{emp_id}/")
        return HttpResponseRedirect(request.META.get("HTTP_REFERER", "/"))
    # Deduct only from this request's leave type (Casual Leave never deducts from EL).
    employee_id = leave_request.employee_id
    leave_type_id = leave_request.leave_type_id
    available_leave = AvailableLeave.objects.get(
        leave_type_id=leave_type_id, employee_id=employee_id
    )
    # Probation Leave & Interns Leave: same logic (1/month, no carryforward, computed balance). Case-insensitive.
    lt_name = (leave_type_id.name or "").strip()
    lt_name_lower = lt_name.lower()
    is_probation_or_interns = (
        "probation" in lt_name_lower or "interns leave" in lt_name_lower or "intern" in lt_name_lower
    )
    as_of = date.today() if is_probation_or_interns else leave_request.start_date
    computed = computed_balance_for_validation(available_leave, as_of_date=as_of)
    if computed is not None:
        total_available_leave = computed
    else:
        total_available_leave = (
            available_leave.available_days + available_leave.carryforward_days
        )
    send_notification = False
    if leave_request.status != "approved":
        # Probation/Interns Leave: max 1 day per calendar month — block approval if already 1 in that month
        if is_probation_or_interns:
            days_in_month = probation_leave_days_in_month(
                leave_request.employee_id,
                leave_type_id,
                leave_request.start_date.year,
                leave_request.start_date.month,
                exclude_leave_request_id=leave_request.id,
            )
            if days_in_month + leave_request.requested_days > 1:
                messages.error(
                    request,
                    _(
                        "Probation Leave / Interns Leave: maximum 1 day per calendar month. "
                        "%(employee)s already has %(days)s day(s) in %(month)s for this leave type."
                    )
                    % {
                        "employee": employee_id,
                        "days": days_in_month,
                        "month": leave_request.start_date.strftime("%B %Y"),
                    },
                )
                if emp_id is not None:
                    employee_id = emp_id
                    return redirect(f"/employee/employee-view/{employee_id}/")
                return HttpResponseRedirect(request.META.get("HTTP_REFERER", "/"))
        if total_available_leave >= leave_request.requested_days:
            if computed is not None:
                # Monthly accrual leave types (e.g., Casual/Earned/Probation): deduct from computed balance
                # so approval logic matches request validation and dashboard balance.
                new_balance = max(0.0, computed - leave_request.requested_days)
                leave_request.approved_available_days = leave_request.requested_days
                leave_request.approved_carryforward_days = 0
                available_leave.available_days = new_balance
                available_leave.carryforward_days = 0
            elif leave_request.requested_days > available_leave.carryforward_days:
                leave = leave_request.requested_days - available_leave.carryforward_days
                leave_request.approved_carryforward_days = (
                    available_leave.carryforward_days
                )
                available_leave.carryforward_days = 0
                available_leave.available_days = available_leave.available_days - leave
                leave_request.approved_available_days = leave
            else:
                temp = available_leave.carryforward_days
                available_leave.carryforward_days = temp - leave_request.requested_days
                leave_request.approved_carryforward_days = leave_request.requested_days
            leave_request.status = "approved"
            if not leave_request.multiple_approvals():
                leave_request.save()
                available_leave.save()
                send_notification = True
            else:
                if request.user.is_superuser:
                    LeaveRequestConditionApproval.objects.filter(
                        leave_request_id=leave_request
                    ).update(is_approved=True)
                    leave_request.save()
                    available_leave.save()
                    send_notification = True
                else:
                    conditional_requests = leave_request.multiple_approvals()
                    approver = next(
                        (
                            manager
                            for manager in conditional_requests["managers"]
                            if manager == request.user.employee_get
                        ),
                        None,
                    )
                    condition_approval = LeaveRequestConditionApproval.objects.filter(
                        manager_id=approver, leave_request_id=leave_request
                    ).first()
                    condition_approval.is_approved = True
                    managers = []
                    for manager in conditional_requests["managers"]:
                        managers.append(manager.employee_user_id)
                    if len(managers) > condition_approval.sequence:
                        with contextlib.suppress(Exception):
                            notify.send(
                                request.user.employee_get,
                                recipient=managers[condition_approval.sequence],
                                verb="You have a new leave request to validate.",
                                verb_ar="لديك طلب إجازة جديد يجب التحقق منه.",
                                verb_de="Sie haben eine neue Urlaubsanfrage zur Validierung.",
                                verb_es="Tiene una nueva solicitud de permiso que debe validar.",
                                verb_fr="Vous avez une nouvelle demande de congé à valider.",
                                icon="people-circle",
                                redirect=f"/leave/request-view?id={leave_request.id}",
                            )

                    condition_approval.save()
                    if approver == conditional_requests["managers"][-1]:
                        leave_request.save()
                        available_leave.save()
                        send_notification = True
            messages.success(request, _("Leave request approved successfully.."))
            if send_notification:
                with contextlib.suppress(Exception):
                    notify.send(
                        request.user.employee_get,
                        recipient=leave_request.employee_id.employee_user_id,
                        verb="Your Leave request has been approved",
                        verb_ar="تمت الموافقة على طلب الإجازة الخاص بك",
                        verb_de="Ihr Urlaubsantrag wurde genehmigt",
                        verb_es="Se ha aprobado su solicitud de permiso",
                        verb_fr="Votre demande de congé a été approuvée",
                        icon="people-circle",
                        redirect=reverse("user-request-view")
                        + f"?id={leave_request.id}",
                    )
            # Always send email when we just approved (so employee gets notification)
            if leave_request.status == "approved":
                if not send_leave_approve_reject_email(request, leave_request, approved=True):
                    messages.warning(
                        request,
                        _("Leave approved but email not sent: employee has no email set."),
                    )
        else:
            messages.error(
                request,
                f"{employee_id} dont have enough leave days to approve the request..",
            )
    else:
        messages.error(request, _("Leave request already approved"))
    if emp_id is not None:
        employee_id = emp_id
        return redirect(f"/employee/employee-view/{employee_id}/")
    return HttpResponseRedirect(request.META.get("HTTP_REFERER", "/"))


@login_required
@manager_can_enter("leave.change_leaverequest")
def leave_request_bulk_approve(request):
    if request.method == "POST":
        request_ids = request.POST.getlist("ids")
        for request_id in request_ids:
            try:
                leave_request = (
                    LeaveRequest.objects.get(id=int(request_id)) if request_id else None
                )
                if leave_request and not can_user_approve_leave_request(
                    request.user, leave_request
                ):
                    continue
                if leave_request.status == "requested" and (
                    leave_request.start_date >= datetime.today().date()
                    or request.user.has_perm("leave.change_leaverequest")
                ):
                    leave_request_approve(request, leave_request.id)
                else:
                    if leave_request.status == "approved":
                        messages.info(
                            request,
                            _("{} {} request already approved").format(
                                leave_request.employee_id, leave_request.leave_type_id
                            ),
                        )
                    elif leave_request.start_date < datetime.today().date():
                        messages.warning(
                            request,
                            _("{} {} request date exceeded").format(
                                leave_request.employee_id, leave_request.leave_type_id
                            ),
                        )
                    else:
                        messages.warning(
                            request,
                            _("{} {} can't approve.").format(
                                leave_request.employee_id, leave_request.leave_type_id
                            ),
                        )
            except (ValueError, OverflowError, LeaveRequest.DoesNotExist):
                messages.error(request, _("Leave request not found"))
                pass
    return HttpResponse("<script>window.location.reload();</script>")


@login_required
@manager_can_enter("leave.change_leaverequest")
def leave_bulk_reject(request):
    request_ids = request.POST.getlist("request_ids")

    for request_id in request_ids:
        leave_request = (
            LeaveRequest.objects.get(id=int(request_id)) if request_id else None
        )
        leave_request_cancel(request, leave_request.id)

    return HttpResponse("<script>window.location.reload();</script>")


@login_required
@manager_can_enter("leave.change_leaverequest")
def leave_request_cancel(request, id, emp_id=None):
    """
    function used to Reject leave request.

    Parameters:
    request (HttpRequest): The HTTP request object.
    id : leave request id
    emp_id : employee id if the cancel operation comes from "/employee/employee-view/{employee_id}/" template.

    Returns:
    GET : If `emp_id` is provided, it returns to the "/employee/employee-view/{employee_id}/" template after cancel.
          Otherwise, it returns to the default leave request view template.

    """
    form = RejectForm()
    if request.method == "POST":
        form = RejectForm(request.POST)
        if form.is_valid():
            leave_request = LeaveRequest.objects.get(id=id)
            # Add back only to this request's leave type (Casual Leave never touched EL).
            employee_id = leave_request.employee_id
            leave_type_id = leave_request.leave_type_id
            available_leave = AvailableLeave.objects.filter(
                leave_type_id=leave_type_id, employee_id=employee_id
            ).first()
            if leave_request.status != "rejected":
                if available_leave:
                    available_leave.available_days += leave_request.approved_available_days
                    available_leave.carryforward_days += (
                        leave_request.approved_carryforward_days
                    )
                    available_leave.save()
                leave_request.approved_available_days = 0
                leave_request.approved_carryforward_days = 0
                leave_request.status = "rejected"
                leave_request.leave_clashes_count = 0

                if leave_request.multiple_approvals() and not request.user.is_superuser:
                    conditional_requests = leave_request.multiple_approvals()
                    approver = [
                        manager
                        for manager in conditional_requests["managers"]
                        if manager.employee_user_id == request.user
                    ]
                    condition_approval = LeaveRequestConditionApproval.objects.filter(
                        manager_id=approver[0], leave_request_id=leave_request
                    ).first()
                    condition_approval.is_approved = False
                    condition_approval.is_rejected = True
                    condition_approval.save()

                leave_request.reject_reason = form.cleaned_data["reason"]
                leave_request.save()
                comment = LeaverequestComment()
                comment.request_id = leave_request
                comment.employee_id = request.user.employee_get
                comment.comment = leave_request.reject_reason
                comment.save()

                messages.success(request, _("Leave request rejected successfully.."))
                with contextlib.suppress(Exception):
                    notify.send(
                        request.user.employee_get,
                        recipient=leave_request.employee_id.employee_user_id,
                        verb="Your leave request has been rejected.",
                        verb_ar="تم رفض طلب الإجازة الخاص بك",
                        verb_de="Ihr Urlaubsantrag wurde abgelehnt",
                        verb_es="Tu solicitud de permiso ha sido rechazada",
                        verb_fr="Votre demande de congé a été rejetée",
                        icon="people-circle",
                        redirect=reverse("user-request-view")
                        + f"?id={leave_request.id}",
                    )

                # Send reject email in background so response returns immediately (avoids 5s delay and double-click "already rejected")
                owner = leave_request.employee_id
                to_email = getattr(owner, "get_mail", None)
                to_email = to_email() if callable(to_email) else None
                if not to_email:
                    messages.warning(
                        request,
                        _("Leave rejected but email not sent: employee has no email set."),
                    )
                else:
                    LeaveMailSendThread(request, leave_request, type="reject").start()
            else:
                messages.error(request, _("Leave request already rejected."))

            if emp_id is not None:
                employee_id = emp_id
                return redirect(f"/employee/employee-view/{employee_id}/")
            return HttpResponse("<script>location.reload();</script>")
    return render(
        request, "leave/leave_request/cancel_form.html", {"form": form, "id": id}
    )


@login_required
@hx_request_required
def user_leave_cancel(request, id):
    """
    function used to cancel approved leave request by employee.

    Parameters:
    request (HttpRequest): The HTTP request object.
    id : leave request id

    Returns:
    GET :  it returns to the default my leave request view template.

    """
    leave_request = LeaveRequest.objects.get(id=id)
    employee_id = leave_request.employee_id
    if employee_id.employee_user_id.id == request.user.id:
        current_date = date.today()
        if (
            leave_request.status == "approved"
            and leave_request.end_date >= current_date
        ):
            form = RejectForm()
            if request.method == "POST":
                form = RejectForm(request.POST)
                if form.is_valid():
                    leave_request.reject_reason = form.cleaned_data["reason"]
                    leave_request.status = "cancelled"
                    leave_request.save()
                    messages.success(
                        request, _("Leave request cancelled successfully..")
                    )

                    mail_thread = LeaveMailSendThread(
                        request, leave_request, type="cancel"
                    )
                    mail_thread.start()
                    return HttpResponse("<script>location.reload();</script>")
            return render(
                request,
                "leave/leave_request/user_cancel_form.html",
                {"form": form, "id": id},
            )
        messages.error(request, _("You can't cancel this leave request."))
        return HttpResponse("<script>location.reload();</script>")
    messages.error(request, _("You don't have the permission."))
    return HttpResponse("<script>location.reload();</script>")


@login_required
@hx_request_required
@manager_can_enter("leave.view_leaverequest")
def one_request_view(request, id):
    """
    function used to view one leave request.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET : return one leave request view template
    """
    leave_request = LeaveRequest.objects.get(id=id)
    context = {
        "leave_request": leave_request,
        "current_date": date.today(),
        "dashboard": request.GET.get("dashboard"),
    }
    requests_ids_json = request.GET.get("instances_ids")

    if requests_ids_json:
        requests_ids = json.loads(requests_ids_json)
        previous_id, next_id = closest_numbers(requests_ids, id)
        context["previous"] = previous_id
        context["next"] = next_id
        context["requests_ids"] = requests_ids_json
    return render(request, "leave/leave_request/one_request_view.html", context)


@login_required
@hx_request_required
@manager_can_enter("leave.add_availableleave")
def leave_assign_one(request, obj_id):
    """
    Assigns leave types to employees.

    Parameters:
    request (HttpRequest): The HTTP request object.
    obj_id: ID of the leave type.

    Returns:
    GET: Renders the leave type assignment form template.
    POST: Processes and assigns the leave type to selected employees.
    """
    form = LeaveOneAssignForm()
    form = choosesubordinates(request, form, "leave.add_availableleave")

    # Fetch the leave type
    leave_type = LeaveType.objects.filter(id=obj_id).first()
    if not leave_type:
        messages.error(request, _("Leave type not found."))
        return render(
            request,
            "leave/leave_assign/leave_assign_one_form.html",
            {"form": form, "id": obj_id},
        )

    if request.method == "POST":
        if leave_type.is_compensatory_leave:
            messages.info(
                request, _("Compensatory leave type cannot be assigned manually.")
            )
            return render(
                request,
                "leave/leave_assign/leave_assign_one_form.html",
                {"form": form, "id": obj_id},
            )

        employee_ids = list(map(int, request.POST.getlist("employee_id")))

        existing_leaves_set = set(
            AvailableLeave.objects.filter(
                leave_type_id=leave_type, employee_id__in=employee_ids
            ).values_list("employee_id", flat=True)
        )

        expiry_date = (
            leave_type.carryforward_expire_date
            if leave_type.carryforward_expire_date
            else None
        )
        new_employees = list(set(employee_ids) - existing_leaves_set)

        assigned_count = 0
        if new_employees:
            assigned_date = date.today()
            emp_map = {e.id: e for e in Employee.objects.filter(id__in=new_employees)}
            available_leaves = []
            for employee_id in new_employees:
                employee = emp_map.get(employee_id)
                init_days, reset_override = get_init_days_and_reset_for_assign(
                    leave_type, assigned_date, employee=employee
                )
                leave = AvailableLeave(
                    leave_type_id=leave_type,
                    employee_id_id=employee_id,
                    available_days=init_days,
                    assigned_date=assigned_date,
                )
                if reset_override is not None:
                    leave.reset_date = reset_override
                elif leave.reset_date is None and leave_type.reset:
                    leave.reset_date = leave.set_reset_date(
                        assigned_date=assigned_date, available_leave=leave
                    )

                if leave_type.carryforward_type == "carryforward expire":
                    if not expiry_date:
                        expiry_date = leave.assigned_date
                    leave.expired_date = expiry_date

                leave.total_leave_days = max(
                    leave.available_days + leave.carryforward_days, 0
                )
                leave.carryforward_days = max(leave.carryforward_days, 0)
                available_leaves.append(leave)

            AvailableLeave.objects.bulk_create(available_leaves)
            assigned_count = len(available_leaves)

            messages.success(
                request,
                _("Successfully assigned leave type to {} employees.").format(
                    assigned_count
                ),
            )
            form = LeaveOneAssignForm()

            employees = Employee.objects.filter(id__in=new_employees).only(
                "id", "employee_user_id"
            )
            notifications = [
                notify.send(
                    request.user.employee_get,
                    recipient=employee.employee_user_id,
                    verb="New leave type is assigned to you",
                    verb_ar="تم تعيين نوع إجازة جديد لك",
                    verb_de="Ihnen wurde ein neuer Urlaubstyp zugewiesen",
                    verb_es="Se le ha asignado un nuevo tipo de permiso",
                    verb_fr="Un nouveau type de congé vous a été attribué",
                    icon="people-circle",
                    redirect=reverse("user-request-view"),
                )
                for employee in employees
            ]

        if len(employee_ids) != assigned_count:
            messages.info(
                request,
                _(
                    "Leave type is already assigned to some selected {} employees."
                ).format(len(employee_ids) - assigned_count),
            )

    return render(
        request,
        "leave/leave_assign/leave_assign_one_form.html",
        {"form": form, "id": obj_id},
    )


@login_required
@manager_can_enter("leave.view_availableleave")
def leave_assign_view(request):
    """
    Function to view assigned employee leaves.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET : return leave assigned view template
    """
    queryset = filtersubordinates(
        request, AvailableLeave.objects.all(), "leave.view_availableleave"
    )
    previous_data = request.GET.urlencode() or "field=leave_type_id"
    field = request.GET.get("field", "leave_type_id")
    page_number = request.GET.get("page")

    # Paginate and group queryset by field
    page_obj = group_by_queryset(queryset.order_by("-id"), field, page_number)
    available_leave_ids = json.dumps(
        [instance.id for entry in page_obj for instance in entry["list"].object_list]
    )

    # Setting a condition for the template
    request.GET = request.GET.copy()
    request.GET["field"] = True

    return render(
        request,
        "leave/leave_assign/assign_view.html",
        {
            "available_leaves": page_obj,
            "f": AssignedLeaveFilter(),
            "pd": previous_data,
            "filter_dict": parse_qs(previous_data),
            "gp_fields": LeaveAssignReGroup.fields,
            "assign_form": AssignLeaveForm(),
            "available_leave_ids": available_leave_ids,
        },
    )


@login_required
@hx_request_required
@manager_can_enter("leave.view_availableleave")
def available_leave_single_view(request, obj_id):
    get_data = request.GET.copy()
    get_data.pop("instances_ids", None)
    previous_data = get_data.urlencode()

    available_leave = AvailableLeave.objects.filter(id=obj_id).select_related("leave_type_id", "employee_id").first()
    instance_ids = json.loads(request.GET.get("instances_ids", "[]"))
    previous_instance, next_instance = (
        closest_numbers(instance_ids, obj_id) if instance_ids else (None, None)
    )

    computed_leave_display = None
    if available_leave:
        disp = _computed_leave_display_for_available(available_leave)
        if disp is not None:
            computed_leave_display = disp

    content = {
        "available_leave": available_leave,
        "previous_instance": previous_instance,
        "next_instance": next_instance,
        "instance_ids_json": json.dumps(instance_ids),
        "pd": previous_data,
        "computed_leave_display": computed_leave_display,
    }
    return render(
        request, "leave/leave_assign/single_assign_view.html", context=content
    )


@login_required
@hx_request_required
@manager_can_enter("leave.view_availableleave")
def leave_assign_filter(request):
    """
    function used to filter assign leave type to employees.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET : return leave type assigned view template
    """
    queryset = AvailableLeave.objects.all()
    assign_form = AssignLeaveForm()
    queryset = filtersubordinates(request, queryset, "leave.view_availableleave")
    assigned_leave_filter = AssignedLeaveFilter(request.GET, queryset).qs
    previous_data = request.GET.urlencode()
    field = request.GET.get("field")
    page_number = request.GET.get("page")
    template = ("leave/leave_assign/assigned_leave.html",)
    available_leaves = assigned_leave_filter.order_by("-id")
    if request.GET.get("sortby"):
        available_leaves = sortby(request, available_leaves, "sortby")
        available_leave_ids = json.dumps(
            [instance.id for instance in paginator_qry(available_leaves, None)]
        )
    if field != "" and field is not None:
        page_obj = group_by_queryset(available_leaves, field, page_number)
        list_values = [entry["list"] for entry in page_obj]
        id_list = []
        for value in list_values:
            for instance in value.object_list:
                id_list.append(instance.id)
        available_leave_ids = json.dumps(list(id_list))
        template = "leave/leave_assign/group_by.html"
    else:
        available_leave_ids = json.dumps(
            [instance.id for instance in paginator_qry(available_leaves, None)]
        )
        page_obj = paginator_qry(available_leaves, page_number)

    data_dict = parse_qs(previous_data)
    get_key_instances(AvailableLeave, data_dict)

    # Build computed display for monthly accrual types (EL, CL, Probation Leave) so All Assigned Leaves shows correct balances
    if field:
        leaves_flat = []
        for entry in page_obj:
            leaves_flat.extend(entry["list"].object_list)
    else:
        leaves_flat = list(page_obj.object_list)
    computed_leave_display = {}
    for av in leaves_flat:
        disp = _computed_leave_display_for_available(av)
        if disp is not None:
            computed_leave_display[av.id] = disp

    return render(
        request,
        template,
        {
            "available_leaves": page_obj,
            "pd": previous_data,
            "filter_dict": data_dict,
            "field": field,
            "assign_form": assign_form,
            "available_leave_ids": available_leave_ids,
            "computed_leave_display": computed_leave_display,
        },
    )


@login_required
@hx_request_required
@manager_can_enter("leave.add_availableleave")
def leave_assign(request):
    """
    Function to assign multiple leave types to employees.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET: Render the leave assign form template.
    POST: Handle the leave type assignment.
    """
    form = AssignLeaveForm()
    form = choosesubordinates(request, form, "leave.add_availableleave")
    page_reload = AvailableLeave.objects.count() == 0

    if request.method == "POST":
        leave_type_ids = request.POST.getlist("leave_type_id")
        employee_ids = request.POST.getlist("employee_id")

        if leave_type_ids and employee_ids:
            leave_types = LeaveType.objects.filter(id__in=leave_type_ids)
            employees = Employee.objects.filter(id__in=employee_ids)

            existing_assignments = set(
                AvailableLeave.objects.filter(
                    leave_type_id__in=leave_type_ids, employee_id__in=employee_ids
                ).values_list("leave_type_id", "employee_id")
            )

            new_assignments = []
            success_messages = set()
            info_messages = set()

            for employee in employees:
                for leave_type in leave_types:
                    assignment_key = (leave_type.id, employee.id)
                    if assignment_key not in existing_assignments:
                        assigned_date = date.today()
                        init_days, reset_override = get_init_days_and_reset_for_assign(
                            leave_type, assigned_date, employee=employee
                        )
                        new_assignment = AvailableLeave(
                            leave_type_id=leave_type,
                            employee_id=employee,
                            available_days=init_days,
                            assigned_date=assigned_date,
                        )
                        if reset_override is not None:
                            new_assignment.reset_date = reset_override
                        new_assignments.append(new_assignment)
                        new_assignment.pre_save_processing()
                        success_messages.add(employee.employee_user_id)
                    else:
                        info_messages.add(employee.employee_user_id)

            # Bulk create new assignments
            if new_assignments:
                with transaction.atomic():
                    AvailableLeave.objects.bulk_create(new_assignments)
                    for user_id in success_messages:
                        with contextlib.suppress(Exception):
                            notify.send(
                                request.user.employee_get,
                                recipient=user_id,
                                verb="New leave type is assigned to you",
                                verb_ar="تم تعيين نوع إجازة جديد لك",
                                verb_de="Dir wurde ein neuer Urlaubstyp zugewiesen",
                                verb_es="Se te ha asignado un nuevo tipo de permiso",
                                verb_fr="Un nouveau type de congé vous a été attribué",
                                icon="people-circle",
                                redirect=reverse("user-request-view"),
                            )
                    messages.success(request, _("Leave types assigned successfully."))

            if info_messages:
                messages.info(
                    request,
                    _("Some leave types were already assigned to {} employees.").format(
                        len(info_messages)
                    ),
                )

        if page_reload:
            return HttpResponse("<script>window.location.reload()</script>")

    return render(
        request, "leave/leave_assign/leave_assign_form.html", {"assign_form": form}
    )


@login_required
@hx_request_required
@manager_can_enter("leave.change_availableleave")
def available_leave_update(request, id):
    """
    function used to update available leave of an assigned leave type of an employee.

    Parameters:
    request (HttpRequest): The HTTP request object.
    id : available leave id

    Returns:
    GET : return available leave update form template
    POST : return leave type assigned  view
    """
    leave_assign = AvailableLeave.objects.get(id=id)
    form = AvailableLeaveUpdateForm(instance=leave_assign)
    previous_data = request.GET.urlencode() or "field=leave_type_id"
    if request.method == "POST":
        form = AvailableLeaveUpdateForm(request.POST, instance=leave_assign)
        if form.is_valid():
            available_leave = form.save()
            messages.success(request, _("Available leaves updated successfully..."))
            with contextlib.suppress(Exception):
                notify.send(
                    request.user.employee_get,
                    recipient=available_leave.employee_id.employee_user_id,
                    verb=f"Your {available_leave.leave_type_id} leave type updated.",
                    verb_ar=f"تم تحديث نوع الإجازة {available_leave.leave_type_id} الخاص بك.",
                    verb_de=f"Ihr Urlaubstyp {available_leave.leave_type_id} wurde aktualisiert.",
                    verb_es=f"Se ha actualizado su tipo de permiso {available_leave.leave_type_id}.",
                    verb_fr=f"Votre type de congé {available_leave.leave_type_id} a été mis à jour.",
                    icon="people-circle",
                    redirect=reverse("user-request-view"),
                )
    return render(
        request,
        "leave/leave_assign/available_update_form.html",
        {"form": form, "id": id, "pd": previous_data},
    )


@login_required
@hx_request_required
@manager_can_enter("leave.delete_availableleave")
def leave_assign_delete(request, obj_id):
    """
    Function to delete an assigned leave type of an employee.

    Parameters:
    - request (HttpRequest): The HTTP request object.
    - obj_id (int): Available leave ID.

    Returns:
    - Redirects to the assigned leave type view or refreshes the page.
    """
    pd = request.GET.urlencode()

    try:
        AvailableLeave.objects.get(id=obj_id).delete()
        messages.success(request, _("Assigned leave successfully deleted."))
    except AvailableLeave.DoesNotExist:
        messages.error(request, _("Assigned leave not found."))
    except ProtectedError:
        messages.error(request, _("Related entries exist."))

    if instances_ids := request.GET.get("instances_ids"):
        instances_list = json.loads(instances_ids)
        previous_instance, next_instance = closest_numbers(instances_list, obj_id)
        if obj_id in instances_list:
            instances_list.remove(obj_id)
        return redirect(
            f"/leave/available-leave-single-view/{next_instance}/?{pd}&instances_ids={json.dumps(instances_list)}"
        )

    if not AvailableLeave.objects.exists():
        return HttpResponse("<script>window.location.reload()</script>")
    return redirect(f"/leave/assign-filter?{pd}")


@require_http_methods(["POST"])
@permission_required("leave.delete_availableleave")
def leave_assign_bulk_delete(request):
    """
    This method is used to delete bulk of assigned leaves
    """
    ids = request.POST["ids"]
    ids = json.loads(ids)
    count = 0
    for assigned_leave_id in ids:
        try:
            assigned_leave = AvailableLeave.objects.get(id=assigned_leave_id)
            assigned_leave.delete()
            count += 1
        except Exception as e:
            messages.error(request, _("Assigned leave not found."))
    messages.success(
        request, _("{} assigned leaves deleted successfully ").format(count)
    )
    return JsonResponse({"message": "Success"})


def assign_leave_type_excel(_request):
    """
    Generate an empty Excel template for asisgn leave type to employee with predefined columns.

    Returns:
        HttpResponse: An HTTP response containing an empty Excel template with predefined columns.
    """
    try:
        columns = [
            "Employee Badge ID",
            "Leave Type",
        ]
        data_frame = pd.DataFrame(columns=columns)
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = (
            'attachment; filename="assign_leave_type_excel.xlsx"'
        )
        data_frame.to_excel(response, index=False)
        return response
    except Exception as exception:
        return HttpResponse(exception)


@login_required
@manager_can_enter("leave.add_availableleave")
def assign_leave_type_import(request):
    """
    This function accepts a POST request containing an Excel file with assigned leave type to employee data.
    It processes the data, checks for errors, and either assigns leave types to employees
    or generates an error report in the form of an Excel file.
    """
    error_data = {
        "Employee Badge ID": [],
        "Leave Type": [],
        "Badge ID Error": [],
        "Leave Type Error": [],
        "Assigned Error": [],
        "Available Days": [],
        "Carry Forward Days": [],
        "Other Errors": [],
    }

    if request.method == "POST":
        file = request.FILES["assign_leave_type_import"]
        data_frame = pd.read_excel(file)
        assign_leave_dicts = data_frame.to_dict("records")

        # Pre-fetch all employees and leave types
        employees = {
            emp.badge_id.lower(): emp for emp in Employee.objects.all() if emp.badge_id
        }
        leave_types = {lt.name.lower(): lt for lt in LeaveType.objects.all()}
        available_leaves = {
            (al.leave_type_id.id, al.employee_id.id): al
            for al in AvailableLeave.objects.all()
        }

        assign_leave_list = []
        error_list = []

        for assign_leave in assign_leave_dicts:
            badge_id = assign_leave.get("Employee Badge ID", "").strip().lower()
            assign_leave_type = assign_leave.get("Leave Type", "").strip().lower()
            available_days = assign_leave.get("Available Days", "0")
            cfd = assign_leave.get("Carry Forward Days", "0")
            employee = employees.get(badge_id)
            leave_type = leave_types.get(assign_leave_type)

            errors = []
            if employee is None:
                errors.append(_("This badge id does not exist."))
            if leave_type is None:
                errors.append(_("This leave type does not exist."))
            if errors:
                assign_leave[
                    "Badge ID Error" if "badge id" in errors[0] else "Leave Type Error"
                ] = " ".join(force_str(error) for error in errors)
                error_list.append(assign_leave)
                continue

            # Check if leave type has already been assigned to the employee
            if (leave_type.id, employee.id) in available_leaves:
                assign_leave["Assigned Error"] = _(
                    "Leave type has already been assigned to the employee."
                )
                error_list.append(assign_leave)
                continue

            # If no errors, create the AvailableLeave instance
            if available_days == 0:
                available_days = leave_type.total_days

            available_leave = AvailableLeave(
                leave_type_id=leave_type,
                employee_id=employee,
                available_days=available_days,
            )
            if cfd:
                available_leave.carryforward_days = cfd
                available_leave.expired_date = leave_type.carryforward_expire_date
                try:
                    available_leave.reset_date = leave_type.leave_type_next_reset_date()
                except:
                    pass
                available_leave.assigned_date = datetime.today()
                available_leave.total_leave_days = (
                    available_leave.carryforward_days + available_leave.available_days
                )
            assign_leave_list.append(available_leave)

        # Bulk create available leaves
        if assign_leave_list:
            AvailableLeave.objects.bulk_create(assign_leave_list)

        # Generate error report if there are errors
        path_info = None
        if error_list:
            path_info = generate_error_report(
                error_list, error_data, "AssignLeaveError.xlsx"
            )

        context = {
            "created_count": len(assign_leave_dicts) - len(error_list),
            "error_count": len(error_list),
            "model": _("Assigned Leaves"),
            "path_info": path_info,
        }
        html = render_to_string("import_popup.html", context)
        return HttpResponse(html)


@login_required
def assigned_leaves_export(request):
    hx_request = request.META.get("HTTP_HX_REQUEST")
    if hx_request:
        export_filter = AssignedLeaveFilter()
        export_column = AvailableLeaveColumnExportForm()
        content = {
            "export_filter": export_filter,
            "export_column": export_column,
        }
        return render(
            request,
            "leave/leave_assign/assigned_leaves_export_form.html",
            context=content,
        )
    return export_data(
        request=request,
        model=AvailableLeave,
        filter_class=AssignedLeaveFilter,
        form_class=AvailableLeaveColumnExportForm,
        file_name="Assign_Leave",
        perm="leave.view_availableleave",
    )


@login_required
@hx_request_required
def get_job_positions(request):
    department_id = request.GET.get("department")
    form = RestrictLeaveForm()
    form.fields["job_position"].queryset = JobPosition.objects.filter(
        department_id=department_id
    )

    return render(request, "leave/job_position_field.html", {"form": form})


@login_required
@hx_request_required
@permission_required("leave.add_restrictleave")
def restrict_creation(request):
    """
    function used to create restricted days.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET : return restricted days creation form template
    POST : return restricted days view template
    """

    query_string = request.GET.urlencode()
    if query_string.startswith("pd="):
        previous_data = unquote(query_string[len("pd=") :])
    else:
        previous_data = unquote(query_string)
    form = RestrictLeaveForm()

    if request.method == "POST":
        form = RestrictLeaveForm(request.POST)
        if form.is_valid():
            form.save()
            form = RestrictLeaveForm()
            messages.success(request, _("Restricted day created successfully.."))
            if RestrictLeave.objects.filter().count() == 1:
                return HttpResponse("<script>window.location.reload();</script>")
    return render(
        request,
        "leave/restrict/restrict_form.html",
        {"form": form, "pd": previous_data},
    )


@login_required
def restrict_view(request):
    """
    function used to view restricted days.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET : return restricted days view  template
    """
    queryset = RestrictLeave.objects.all()[::-1]
    previous_data = request.GET.urlencode()
    page_number = request.GET.get("page")
    page_obj = paginator_qry(queryset, page_number)
    restrictday_filter = RestrictLeaveFilter()
    return render(
        request,
        "leave/restrict/view_restrict.html",
        {
            "restrictday": page_obj,
            "form": restrictday_filter.form,
            "pd": previous_data,
        },
    )


@login_required
@hx_request_required
def restrict_filter(request):
    """
    function used to filter restricted days.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET : return restricted days view template
    """
    queryset = RestrictLeave.objects.all()
    previous_data = request.GET.urlencode()
    restrictday_filter = RestrictLeaveFilter(request.GET, queryset).qs
    if request.GET.get("sortby"):
        restrictday_filter = sortby(request, restrictday_filter, "sortby")
    page_number = request.GET.get("page")
    page_obj = paginator_qry(restrictday_filter[::-1], page_number)
    data_dict = parse_qs(previous_data)
    get_key_instances(RestrictLeave, data_dict)
    return render(
        request,
        "leave/restrict/restrict.html",
        {"restrictday": page_obj, "pd": previous_data, "filter_dict": data_dict},
    )


@login_required
@hx_request_required
@permission_required("leave.change_restrictleave")
def restrict_update(request, id):
    """
    function used to update restricted days.

    Parameters:
    request (HttpRequest): The HTTP request object.
    id : restricted days id

    Returns:
    GET : return restricted days update form template
    POST : return restricted days view template
    """
    query_string = request.GET.urlencode()
    if query_string.startswith("pd="):
        previous_data = unquote(query_string[len("pd=") :])
    else:
        previous_data = unquote(query_string)
    restrictday = RestrictLeave.objects.get(id=id)
    form = RestrictLeaveForm(instance=restrictday)
    if request.method == "POST":
        form = RestrictLeaveForm(request.POST, instance=restrictday)
        if form.is_valid():
            form.save()
            messages.success(request, _("Restricted day updated successfully.."))
    return render(
        request,
        "leave/restrict/restrict_update_form.html",
        {"form": form, "id": id, "pd": previous_data},
    )


@login_required
@hx_request_required
@permission_required("leave.delete_restrictleave")
def restrict_delete(request, id):
    """
    function used to delete restricted days.

    Parameters:
    request (HttpRequest): The HTTP request object.
    id : restricted days id

    Returns:
    GET : return restricted days view template
    """
    query_string = request.GET.urlencode()
    try:
        RestrictLeave.objects.get(id=id).delete()
        messages.success(request, _("Restricted day deleted successfully.."))
    except RestrictLeave.DoesNotExist:
        messages.error(request, _("Restricted day not found."))
    except ProtectedError:
        messages.error(request, _("Related entries exists"))
    if not RestrictLeave.objects.filter():
        return HttpResponse("<script>window.location.reload();</script>")
    return redirect(f"/leave/restrict-filter?{query_string}")


@login_required
@hx_request_required
@permission_required("leave.delete_restrictleave")
def restrict_days_bulk_delete(request):
    """
    function used to delete multiple restricted days.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET : return restricted days view template
    """
    pd = request.GET.urlencode()
    if request.method == "POST":
        restrict_day_ids = request.POST.getlist("ids")
        try:
            restrict_days = RestrictLeave.objects.filter(
                id__in=restrict_day_ids
            ).delete()
            count = len(restrict_day_ids)
            messages.success(
                request,
                _("{} Leave restricted days deleted successfully").format(count),
            )
        except (OverflowError, ValueError):
            messages.error(request, _("Restricted Days not found"))
        except:
            messages.error(request, _("Something went wrong"))
    return redirect(f"/leave/restrict-filter?{pd}")


@login_required
@permission_required("leave.add_restrictleave")
def restrict_day_select(request):
    page_number = request.GET.get("page")
    if page_number == "all":
        restrict_days = RestrictLeave.objects.all()
    restrict_day_ids = [str(day.id) for day in restrict_days]
    total_count = len(restrict_day_ids)
    context = {"restrict_day_ids": restrict_day_ids, "total_count": total_count}
    return JsonResponse(context)


@login_required
@permission_required("leave.add_restrictleave")
def restrict_day_select_filter(request):
    page_number = request.GET.get("page")
    filtered = request.GET.get("filter")
    filters = json.loads(filtered) if filtered else {}

    if page_number == "all":
        restrictday_filter = RestrictLeaveFilter(
            filters, queryset=RestrictLeave.objects.all()
        )
        restrictday_filter = restrictday_filter.qs
        restrictday_ids = [str(restrictday.id) for restrictday in restrictday_filter]
        total_count = restrictday_filter.count()
        context = {"restrict_day_ids": restrictday_ids, "total_count": total_count}
        return JsonResponse(context)


@login_required
@hx_request_required
def user_leave_request(request, id):
    """
    function used to create user leave request.

    Parameters:
    request (HttpRequest): The HTTP request object.
    id : leave type id

    Returns:
    GET : return user leave request creation form template
    POST : user my leave view template
    """
    previous_data = unquote(request.GET.urlencode())[len("pd=") :]
    employee = request.user.employee_get
    leave_type = LeaveType.objects.get(id=id)
    form = UserLeaveRequestForm(
        initial={"employee_id": employee, "leave_type_id": leave_type},
        employee=employee,
    )
    if request.method == "POST":
        form = UserLeaveRequestForm(request.POST, request.FILES, employee=employee)
        start_date = datetime.strptime(request.POST.get("start_date"), "%Y-%m-%d")
        end_date = datetime.strptime(request.POST.get("end_date"), "%Y-%m-%d")
        start_date_breakdown = request.POST.get("start_date_breakdown")
        end_date_breakdown = request.POST.get("end_date_breakdown")
        available_leave = AvailableLeave.objects.get(
            employee_id=employee, leave_type_id=leave_type
        )
        available_total_leave = (
            available_leave.available_days + available_leave.carryforward_days
        )
        # Calculate requested days with automatic holiday/company leave exclusion
        requested_days = calculate_requested_days(
            start_date, end_date, start_date_breakdown, end_date_breakdown, exclude_holidays=True, employee_id=employee
        )

        if form.is_valid():
            leave_request = form.save(commit=False)
            save = True
            leave_request.leave_type_id = leave_type
            leave_request.employee_id = employee

            if leave_request.leave_type_id.require_approval == "no":
                # Deduct only from this request's leave type (Casual Leave never deducts from EL).
                employee_id = leave_request.employee_id
                leave_type_id = leave_request.leave_type_id
                available_leave = AvailableLeave.objects.get(
                    leave_type_id=leave_type_id, employee_id=employee_id
                )
                if leave_request.requested_days > available_leave.available_days:
                    leave = (
                        leave_request.requested_days - available_leave.available_days
                    )
                    leave_request.approved_available_days = (
                        available_leave.available_days
                    )
                    available_leave.available_days = 0
                    available_leave.carryforward_days = (
                        available_leave.carryforward_days - leave
                    )
                    leave_request.approved_carryforward_days = leave
                else:
                    available_leave.available_days = (
                        available_leave.available_days - leave_request.requested_days
                    )
                    leave_request.approved_available_days = leave_request.requested_days
                leave_request.status = "approved"
                available_leave.save()
            if save:
                leave_request.created_by = employee
                leave_request.save()

                if multiple_approvals_check(leave_request.id):
                    conditional_requests = multiple_approvals_check(leave_request.id)
                    managers = []
                    for manager in conditional_requests["managers"]:
                        managers.append(manager.employee_user_id)
                    with contextlib.suppress(Exception):
                        notify.send(
                            request.user.employee_get,
                            recipient=managers[0],
                            verb="You have a new leave request to validate.",
                            verb_ar="لديك طلب إجازة جديد يجب التحقق منه.",
                            verb_de="Sie haben eine neue Urlaubsanfrage zur Validierung.",
                            verb_es="Tiene una nueva solicitud de permiso que debe validar.",
                            verb_fr="Vous avez une nouvelle demande de congé à valider.",
                            icon="people-circle",
                            redirect=f"/leave/request-view?id={leave_request.id}",
                        )
                mail_thread = LeaveMailSendThread(
                    request, leave_request, type="request"
                )
                mail_thread.start()
                messages.success(request, _("Leave request created successfully.."))
                with contextlib.suppress(Exception):
                    notify.send(
                        request.user.employee_get,
                        recipient=leave_request.employee_id.employee_work_info.reporting_manager_id.employee_user_id,
                        verb="You have a new leave request to validate.",
                        verb_ar="لديك طلب إجازة جديد يجب التحقق منه.",
                        verb_de="Sie haben eine neue Urlaubsanfrage zur Validierung.",
                        verb_es="Tiene una nueva solicitud de permiso que debe validar.",
                        verb_fr="Vous avez une nouvelle demande de congé à valider.",
                        icon="people-circle",
                        redirect=reverse("request-view") + f"?id={leave_request.id}",
                    )
                if len(
                    LeaveRequest.objects.filter(employee_id=employee)
                ) == 1 or request.META.get("HTTP_REFERER").endswith(
                    "employee-profile/"
                ):
                    return HttpResponse("<script>window.location.reload();</script>")

        return render(
            request,
            "leave/user_leave/user_request_form.html",
            {
                "form": form,
                "id": id,
                "leave_type": leave_type,
                "pd": previous_data,
            },
        )
    form.fields["leave_type_id"].queryset = LeaveType.objects.filter(id=id)
    return render(
        request,
        "leave/user_leave/user_request_form.html",
        {
            "form": form,
            "id": id,
            "leave_type": leave_type,
            "pd": previous_data,
        },
    )


@login_required
@hx_request_required
def user_request_update(request, id):
    """
    function used to update user leave request.

    Parameters:
    request (HttpRequest): The HTTP request object.
    id : leave request id

    Returns:
    GET : return user leave request update form template
    POST : return user leave request view template
    """
    previous_data = request.GET.urlencode()
    leave_request = LeaveRequest.objects.get(id=id)
    try:
        if (
            request.user.employee_get == leave_request.employee_id
            and leave_request.status != "approved"
        ):
            form = UserLeaveRequestForm(
                employee=leave_request.employee_id, instance=leave_request
            )
            if request.method == "POST":
                # Original requested days (before update) - not yet deducted from balance when status is requested
                original_requested_days = leave_request.requested_days
                form = UserLeaveRequestForm(
                    request.POST,
                    request.FILES,
                    instance=leave_request,
                    employee=leave_request.employee_id,
                )
                if form.is_valid():
                    leave_request = form.save(commit=False)

                    start_date = leave_request.start_date
                    end_date = leave_request.end_date
                    start_date_breakdown = leave_request.start_date_breakdown
                    end_date_breakdown = leave_request.end_date_breakdown
                    leave_type = leave_request.leave_type_id
                    employee = request.user.employee_get
                    available_leave = AvailableLeave.objects.get(
                        employee_id=employee, leave_type_id=leave_type
                    )
                    available_total_leave = (
                        available_leave.available_days
                        + available_leave.carryforward_days
                    )
                    # When updating a request that is not yet approved, its days were never deducted;
                    # add them back so we don't double-count when validating the updated request.
                    if leave_request.status != "approved":
                        available_total_leave = (
                            available_total_leave + original_requested_days
                        )
                    # Calculate requested days with automatic holiday/company leave exclusion
                    requested_days = calculate_requested_days(
                        start_date, end_date, start_date_breakdown, end_date_breakdown, exclude_holidays=True, employee_id=leave_request.employee_id
                    )
                    if requested_days <= available_total_leave:
                        leave_request.save()
                        messages.success(
                            request, _("Leave request updated successfully..")
                        )
                    else:
                        form.add_error(
                            None,
                            _("You dont have enough leave days to make the request.."),
                        )
            return render(
                request,
                "leave/user_leave/user_request_update.html",
                {
                    "form": form,
                    "id": id,
                    "pd": previous_data,
                },
            )
        else:
            messages.error(request, _("You can't update this leave request..."))
            return HttpResponse("<script>window.location.reload();</script>")
    except Exception as e:
        messages.error(request, _("User has no leave request.."))
    return render(
        request,
        "leave/user_leave/user_request_update.html",
        {
            "form": form,
            "id": id,
            "pd": previous_data,
        },
    )


@login_required
@hx_request_required
def user_request_delete(request, id):
    """
    function used to delete user leave request.

    Parameters:
    request (HttpRequest): The HTTP request object.
    id : leave request id

    Returns:
    GET : return user leave request view template
    """
    previous_data = request.GET.urlencode()
    try:
        leave_request = LeaveRequest.objects.get(id=id)
        if request.user.employee_get == leave_request.employee_id:
            messages.success(request, _("Leave request deleted successfully.."))
            leave_request.delete()
    except LeaveRequest.DoesNotExist:
        messages.error(request, _("User has no leave request.."))
    except ProtectedError:
        messages.error(request, _("Related entries exists"))
    if not LeaveRequest.objects.filter(employee_id=request.user.employee_get):
        return HttpResponse("<script>window.location.reload();</script>")
    else:
        return redirect(f"/leave/user-request-filter?{previous_data}")


@login_required
@hx_request_required
def user_leave_filter(request):
    """
    function used to filter user assigned leave types.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET : return user assigned leave types view template
    """
    try:
        employee = request.user.employee_get
        queryset = employee.available_leave.all()
        previous_data = request.GET.urlencode()
        page_number = request.GET.get("page")
        assigned_leave_filter = AssignedLeaveFilter(request.GET, queryset).qs
        data_dict = parse_qs(previous_data)
        get_key_instances(AvailableLeave, data_dict)
        page_obj = paginator_qry(assigned_leave_filter, page_number)
        return render(
            request,
            "leave/user_leave/user_leave.html",
            {"user_leaves": page_obj, "pd": previous_data, "filter_dict": data_dict},
        )
    except Exception as e:
        messages.error(request, _("User is not an employee.."))
        return redirect("/")


def _computed_leave_display_for_available(available_leave):
    """
    For Earned Leave (monthly 1.25) and Casual Leave (monthly 1, cap 12), return
    computed balance as of today so cards show correctly. Returns None to use model values.
    """
    if not available_leave or not getattr(available_leave, "leave_type_id", None):
        return None
    leave_type = available_leave.leave_type_id
    lt_name = (leave_type.name or "").strip()
    reset_based = getattr(leave_type, "reset_based", None) or ""
    total_days_val = round(float(leave_type.total_days or 0), 2)
    EL_EPOCH_DATE = date(2026, 1, 1)
    # Earned Leave: monthly 1.25
    is_el = (
        (reset_based == "monthly" and ("Earned Leave" in lt_name or total_days_val == 1.25))
        or ("Earned Leave" in lt_name and total_days_val == 1.25)
    )
    # Casual Leave: monthly 1/day, cap 12 (not Probation Casual Leave)
    is_cl = (
        reset_based == "monthly"
        and total_days_val == 1.0
        and "probation" not in lt_name.lower()
        and ("Casual Leave" in lt_name or "Casual" in lt_name)
    )
    # Probation Leave / Probation Casual Leave / Interns Leave: monthly 1/day, no carryforward
    lt_name_lower = lt_name.lower()
    is_pl = (
        total_days_val == 1.0
        and ("probation" in lt_name_lower or "interns leave" in lt_name_lower or "intern" in lt_name_lower)
    )
    if is_el and leave_type.total_days:
        raw_start = getattr(available_leave, "assigned_date", None)
        accrual_start = max(EL_EPOCH_DATE, raw_start.replace(day=1)) if raw_start else EL_EPOCH_DATE
        today = date.today()
        months = (today.year - accrual_start.year) * 12 + (today.month - accrual_start.month) + 1
        months = max(0, months)
        total_accrued = float(leave_type.total_days) * months
        cap = leave_type.carryforward_max if leave_type.carryforward_max is not None else 9999
        approved_taken = LeaveRequest.objects.filter(
            employee_id=available_leave.employee_id,
            leave_type_id=leave_type,
            status="approved",
        ).aggregate(total=Sum("requested_days"))["total"] or 0
        balance = min(max(0.0, total_accrued - approved_taken), cap)
        total = round(balance, 3)
        taken = round(float(approved_taken), 3)
        return {"available_days": total, "carryforward_days": 0, "total_leave_days": total, "leave_taken": taken}
    if is_cl and leave_type.total_days:
        from leave.methods import get_casual_leave_balance_stats

        stats = get_casual_leave_balance_stats(available_leave)
        return {
            "available_days": stats["available_days"],
            "carryforward_days": 0,
            "total_leave_days": stats["available_days"],
            "leave_taken": stats["leave_taken"],
        }
    # Probation Leave / Interns Leave: 1 per month, NO carry forward. Cap at 2 (this + last month).
    # Display "available_days" as usable in current month (1 - days already taken this month).
    if is_pl and leave_type.total_days:
        from leave.methods import (
            PROBATION_PERIOD_LEAVE_CAP,
            get_probation_period_leave_balance_stats,
        )

        stats = get_probation_period_leave_balance_stats(available_leave)
        total_accrued_cap = round(
            min(stats["months_accrued"], PROBATION_PERIOD_LEAVE_CAP), 3
        )
        return {
            "available_days": stats["available_days"],
            "carryforward_days": 0,
            "total_leave_days": total_accrued_cap,
            "leave_taken": stats["leave_taken"],
        }
    return None


@login_required
def user_request_view(request):
    """
    function used to view user leave request.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET : return user leave request view template
    """
    try:
        user = request.user.employee_get
        queryset = user.leaverequest_set.all()
        previous_data = request.GET.urlencode()
        page_number = request.GET.get("page")

        # Fetching leave requests
        leave_requests = queryset

        leave_requests_with_interview = []
        if apps.is_installed("recruitment"):
            for leave_request in leave_requests:

                # Fetch interviews for the employee within the requested leave period
                InterviewSchedule = get_horilla_model_class(
                    app_label="recruitment", model="interviewschedule"
                )

                interviews = InterviewSchedule.objects.filter(
                    employee_id=leave_request.employee_id,
                    interview_date__range=[
                        leave_request.start_date,
                        leave_request.end_date,
                    ],
                )
                if interviews:
                    # If interview exists then adding the leave request to the list
                    leave_requests_with_interview.append(leave_request)

        user_request_filter = UserLeaveRequestFilter(request.GET, queryset=queryset)
        page_obj = paginator_qry(user_request_filter.qs.order_by("-id"), page_number)
        request_ids = json.dumps(
            list(page_obj.object_list.values_list("id", flat=True))
        )
        user_leave = AvailableLeave.objects.filter(employee_id=user.id).exclude(
            leave_type_id__is_compensatory_leave=True
        )
        if (
            LeaveGeneralSetting.objects.first()
            and LeaveGeneralSetting.objects.first().compensatory_leave
        ):
            user_leave = AvailableLeave.objects.filter(employee_id=user.id)
        user_leaves = user_leave.select_related("leave_type_id")
        # Computed display for EL/CL monthly so cards show correct balance (e.g. 2.5 EL, 2 CL in Feb)
        computed_leave_display = {}
        for av in user_leaves:
            disp = _computed_leave_display_for_available(av)
            if disp is not None:
                computed_leave_display[av.id] = disp
        current_date = date.today()
        return render(
            request,
            "leave/user_leave/user_request_view.html",
            {
                "leave_requests": page_obj,
                "form": user_request_filter.form,
                "pd": previous_data,
                "current_date": current_date,
                "gp_fields": MyLeaveRequestReGroup.fields,
                "request_ids": request_ids,
                "user_leaves": user_leaves,
                "computed_leave_display": computed_leave_display,
                "leave_requests_with_interview": leave_requests_with_interview,
            },
        )
    except Exception as e:
        messages.error(request, _("User is not an employee.."))
        return redirect("/")


@login_required
@hx_request_required
def user_request_filter(request):
    """
    function used to filter user leave request.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET : return user leave request view template
    """
    try:
        user = request.user.employee_get
        queryset = user.leaverequest_set.all().order_by("-id")
        previous_data = request.GET.urlencode()
        page_number = request.GET.get("page")
        field = request.GET.get("field")

        # Fetching leave requests
        leave_requests = queryset

        leave_requests_with_interview = []
        if apps.is_installed("recruitment"):

            for leave_request in leave_requests:

                # Fetch interviews for the employee within the requested leave period
                InterviewSchedule = get_horilla_model_class(
                    app_label="recruitment", model="interviewschedule"
                )

                interviews = InterviewSchedule.objects.filter(
                    employee_id=leave_request.employee_id,
                    interview_date__range=[
                        leave_request.start_date,
                        leave_request.end_date,
                    ],
                )
                if interviews:
                    # If interview exists then adding the leave request to the list
                    leave_requests_with_interview.append(leave_request)

        queryset = sortby(request, queryset, "sortby")
        user_request_filter = UserLeaveRequestFilter(request.GET, queryset).qs
        template = ("leave/user_leave/user_requests.html",)

        if field != "" and field is not None:
            user_request_filter = group_by_queryset(
                user_request_filter, field, request.GET.get("page"), "page"
            )
            list_values = [entry["list"] for entry in user_request_filter]
            id_list = []
            for value in list_values:
                for instance in value.object_list:
                    id_list.append(instance.id)

            requests_ids = json.dumps(list(id_list))
            template = "leave/user_leave/group_by.html"

        else:
            user_request_filter = paginator_qry(
                user_request_filter, request.GET.get("page")
            )
            requests_ids = json.dumps(
                [instance.id for instance in user_request_filter.object_list]
            )

        data_dict = parse_qs(previous_data)
        get_key_instances(LeaveRequest, data_dict)
        if "status" in data_dict:
            status_list = data_dict["status"]
            if len(status_list) > 1:
                data_dict["status"] = [status_list[-1]]

        user_leave = AvailableLeave.objects.filter(employee_id=user.id).exclude(
            leave_type_id__is_compensatory_leave=True
        )
        if (
            LeaveGeneralSetting.objects.first()
            and LeaveGeneralSetting.objects.first().compensatory_leave
        ):
            user_leave = AvailableLeave.objects.filter(employee_id=user.id)
        user_leaves = user_leave

        context = {
            "leave_requests": user_request_filter,
            "pd": previous_data,
            "filter_dict": data_dict,
            "field": field,
            "current_date": date.today(),
            "request_ids": requests_ids,
            "user_leaves": user_leaves,
            "leave_requests_with_interview": leave_requests_with_interview,
        }
        return render(request, template, context=context)
    except Exception as e:
        messages.error(request, _("User is not an employee.."))
        return redirect("/")


@login_required
@hx_request_required
def user_request_one(request, id):
    """
    function used to view one user leave request.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET : return one user leave request view template
    """
    leave_request = LeaveRequest.objects.get(id=id)
    try:
        requests_ids_json = request.GET.get("instances_ids")
        if requests_ids_json:
            requests_ids = json.loads(requests_ids_json)
            previous_id, next_id = closest_numbers(requests_ids, id)
        return render(
            request,
            "leave/user_leave/user_request_one.html",
            {
                "leave_request": leave_request,
                "instances_ids": requests_ids_json,
                "previous": previous_id,
                "next": next_id,
            },
        )
    except Exception as e:
        messages.error(request, _("User has no leave request.."))
        return redirect("/")


@login_required
@manager_can_enter("leave.view_leaverequest")
def employee_leave(request):
    """
    function used to view employees are leave today.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET : return Json response of employee
    """
    leaves = LeaveRequest.employees_on_leave_today(status="approved")
    requests_ids = list(leaves.values_list("id", flat=True))
    today_holidays = Holidays.today_holidays()
    return render(
        request,
        "leave/dashboard/on_leave.html",
        {
            "leaves": leaves,
            "requests_ids": requests_ids,
            "today_holidays": today_holidays,
        },
    )


@login_required
def overall_leave(request):
    """
    function used to view overall leave in the company.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET : return Json response of labels, data
    """

    labels = []
    data = []
    departments = Department.objects.all()
    leave_requests = LeaveRequestFilter(request.GET).qs
    for department in departments:
        count = leave_requests.filter(
            employee_id__employee_work_info__department_id=department
        ).count()
        if count:
            labels.append(department.department)
            data.append(count)
    return JsonResponse({"labels": labels, "data": data})


@login_required
@permission_required("leave.delete_leaverequest")
def dashboard(request):
    """
    function used to view Admin dashboard in the leave module.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET : return Admin dasboard template.
    """
    today = date.today()
    requested = LeaveRequest.objects.filter(status="requested")
    approved = LeaveRequest.objects.filter(
        status="approved", start_date__month=today.month
    )
    rejected = LeaveRequest.objects.filter(
        status="rejected", start_date__month=today.month
    )
    holidays = Holidays.objects.filter(start_date__gte=today)
    next_holiday = holidays.order_by("start_date").first() if holidays else None

    context = {
        "requested": requested,
        "approved": approved,
        "rejected": rejected,
        "next_holiday": next_holiday,
        "dashboard": "dashboard",
        "today": today.strftime("%Y-%m-%d"),
        "first_day": today.replace(day=1).strftime("%Y-%m-%d"),
        "last_day": date(
            today.year, today.month, calendar.monthrange(today.year, today.month)[1]
        ).strftime("%Y-%m-%d"),
    }
    return render(request, "leave/dashboard.html", context)


@login_required
def employee_dashboard(request):
    """
    function used to view Employee dashboard in the leave module.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET : return Employee dasboard template.
    """
    today = date.today()
    user = Employee.objects.get(employee_user_id=request.user)
    leave_requests = LeaveRequest.objects.filter(employee_id=user)
    requested = leave_requests.filter(status="requested")
    approved = leave_requests.filter(status="approved")
    rejected = leave_requests.filter(status="rejected")

    holidays = Holidays.objects.filter(start_date__gte=today)
    next_holiday = (
        holidays.order_by("start_date").first() if holidays.exists() else None
    )

    context = {
        "leave_requests": leave_requests,
        "requested": requested,
        "approved": approved,
        "rejected": rejected,
        "next_holiday": next_holiday,
        "dashboard": "dashboard",
    }
    return render(request, "leave/employee_dashboard.html", context)


@login_required
@hx_request_required
def leave_dashboard_card_modal(request, card_type):
    """
    Return HTML fragment for leave dashboard card modal (admin or employee).
    card_type: 'requested' | 'approved' | 'rejected'
    Query param: dashboard=admin | employee
    """
    today = date.today()
    dashboard_mode = request.GET.get("dashboard", "admin")

    if dashboard_mode == "employee":
        try:
            employee = request.user.employee_get
        except Exception:
            employee = None
        if not employee:
            leave_requests = LeaveRequest.objects.none()
        else:
            leave_requests = LeaveRequest.objects.filter(employee_id=employee)
            if card_type == "requested":
                leave_requests = leave_requests.filter(status="requested")
            elif card_type == "approved":
                leave_requests = leave_requests.filter(status="approved")
            elif card_type == "rejected":
                leave_requests = leave_requests.filter(status="rejected")
            else:
                leave_requests = leave_requests.none()
        leave_requests = leave_requests.order_by("-start_date")[:50]
        show_employee = False
    else:
        # Admin dashboard: same logic as dashboard view
        if card_type == "requested":
            leave_requests = LeaveRequest.objects.filter(status="requested")
        elif card_type == "approved":
            leave_requests = LeaveRequest.objects.filter(
                status="approved", start_date__month=today.month
            )
        elif card_type == "rejected":
            leave_requests = LeaveRequest.objects.filter(
                status="rejected", start_date__month=today.month
            )
        else:
            leave_requests = LeaveRequest.objects.none()
        leave_requests = leave_requests.order_by("-start_date")[:50]
        show_employee = True

    context = {
        "leave_requests": leave_requests,
        "show_employee": show_employee,
    }
    html = render_to_string(
        "leave/dashboard/dashboard_card_modal_content.html",
        context,
        request=request,
    )
    return HttpResponse(html)


@login_required
def dashboard_leave_request(request):
    """
    function used to view leave request table.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET : return leave requests table.
    """
    requests_ids = []
    today = date.today()
    day = request.GET.get("date")
    employee = request.user.employee_get
    leave_requests = LeaveRequest.objects.filter(employee_id=employee)

    if day:
        day = datetime.strptime(day, "%Y-%m")
        leave_requests = leave_requests.filter(
            start_date__month=day.month, start_date__year=day.year
        )
    else:
        leave_requests = leave_requests.filter(
            start_date__month=today.month, start_date__year=today.year
        )

    requests_ids = [request.id for request in leave_requests]
    context = {
        "leave_requests": leave_requests,
        "dashboard": "dashboard",
        "requests_ids": requests_ids,
    }
    return render(request, "leave/leave_request/dashboard_leave_requests.html", context)


@login_required
def available_leave_chart(request):
    """
    Generate available leave chart data for employee dashboard.
    Uses same employee and computed-leave logic as My Leave Requests so that
    monthly/reset leave types (e.g. Probation Leave, Earned Leave, Casual Leave)
    show correct balance for all employees.
    """
    try:
        user = request.user.employee_get
    except Exception:
        return JsonResponse(
            {
                "labels": [],
                "dataset": [{"label": _("Total leaves available"), "data": []}],
                "message": _("Oops!! No leaves available for you this month..."),
            }
        )
    user_leave = AvailableLeave.objects.filter(employee_id=user.id).exclude(
        leave_type_id__is_compensatory_leave=True
    )
    if (
        LeaveGeneralSetting.objects.first()
        and LeaveGeneralSetting.objects.first().compensatory_leave
    ):
        user_leave = AvailableLeave.objects.filter(employee_id=user.id)
    user_leave = user_leave.select_related("leave_type_id")
    leave_count = []
    labels = []
    for av in user_leave:
        computed = _computed_leave_display_for_available(av)
        if computed is not None:
            total = (computed.get("available_days") or 0) + (
                computed.get("carryforward_days") or 0
            )
        else:
            total = (av.available_days or 0) + (av.carryforward_days or 0)
        if total > 0:
            leave_count.append(round(total, 3))
            labels.append(av.leave_type_id.name or "")
    dataset = [
        {
            "label": _("Total leaves available"),
            "data": leave_count,
        },
    ]
    response = {
        "labels": labels,
        "dataset": dataset,
        "message": _("Oops!! No leaves available for you this month..."),
    }
    return JsonResponse(response)


@login_required
def employee_leave_chart(request):
    """
    function used to generate employee leave chart in Admin dashboard.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET : return Json response of labels, dataset, message.
    """

    day = date.today()
    if request.GET.get("date"):
        day = request.GET.get("date")
        day = datetime.strptime(day, "%Y-%m")

    leave_requests = LeaveRequest.objects.filter(
        employee_id__is_active=True, status="approved"
    )
    leave_requests = leave_requests.filter(
        start_date__month=day.month, start_date__year=day.year
    )

    leave_types = leave_requests.values_list("leave_type_id__name", flat=True)

    labels = []
    dataset = []
    for employee in leave_requests.filter(
        start_date__month=day.month, start_date__year=day.year
    ):
        labels.append(employee.employee_id)

    for leave_type in list(set(leave_types)):
        dataset.append(
            {
                "label": leave_type,
                "data": [],
            }
        )

    labels = list(set(labels))
    total_leave_with_type = defaultdict(lambda: defaultdict(float))

    for label in labels:
        leaves = leave_requests.filter(
            employee_id=label, start_date__month=day.month, start_date__year=day.year
        )
        for leave in leaves:
            total_leave_with_type[leave.leave_type_id.name][label] += round(
                leave.requested_days, 2
            )

    for data in dataset:
        dataset_label = data["label"]
        data["data"] = [total_leave_with_type[dataset_label][label] for label in labels]

    employee_label = []
    for employee in list(set(labels)):
        employee_label.append(
            f"{employee.employee_first_name} {employee.employee_last_name}"
        )
    response = {
        "labels": employee_label,
        "dataset": dataset,
        "message": _("No leave request this month"),
    }
    return JsonResponse(response)


@login_required
def department_leave_chart(request):
    """
    function used to generate department leave chart in Admin dashboard.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET : return Json response of labels, dataset.
    """
    day = date.today()
    if request.GET.get("date"):
        day = request.GET.get("date")
        day = datetime.strptime(day, "%Y-%m")

    departments = Department.objects.all()
    department_counts = {dep.department: 0 for dep in departments}
    leave_request = LeaveRequest.objects.filter(status="approved")
    leave_request = leave_request.filter(
        start_date__month=day.month, start_date__year=day.year
    )
    leave_dates = []
    labels = []
    for leave in leave_request:
        for leave_date in leave.requested_dates():
            leave_dates.append(leave_date.strftime("%Y-%m-%d"))

        for dep in departments:
            if dep == leave.employee_id.employee_work_info.department_id:
                department_counts[dep.department] += leave.requested_days

    for department, count in department_counts.items():
        if count != 0:
            labels.append(department)
    values = list(department_counts.values())
    values = [value for value in values if value != 0]
    dataset = [
        {
            "label": _(""),
            "data": values,
        },
    ]
    response = {
        "labels": labels,
        "dataset": dataset,
        "message": _("No leave requests for this month."),
    }
    return JsonResponse(response)


@login_required
def leave_type_chart(request):
    """
    function used to generate leave type chart in Admin dashboard.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET : return Json response of labels, dataset.
    """
    day = date.today()
    if request.GET.get("date"):
        day = request.GET.get("date")
        day = datetime.strptime(day, "%Y-%m")

    leave_types = LeaveType.objects.all()
    leave_type_count = {types.name: 0 for types in leave_types}
    leave_request = LeaveRequest.objects.filter(status="approved")
    leave_request = leave_request.filter(
        start_date__month=day.month, start_date__year=day.year
    )
    for leave in leave_request:
        for lev in leave_types:
            if lev == leave.leave_type_id:
                leave_type_count[lev.name] += leave.requested_days

    # labels = [leave_type.name for leave_type in leave_types]
    labels = []
    for leave_type, count in leave_type_count.items():
        if count != 0:
            labels.append(leave_type)
    values = list(leave_type_count.values())
    values = [value for value in values if value != 0]

    response = {
        "labels": labels,
        "dataset": [
            {
                "data": values,
            },
        ],
        "message": _("No leave requests for any leave type this month."),
    }
    return JsonResponse(response)


@login_required
def leave_over_period(request):
    """
    function used to generate leave over a week chart in Admin dashboard.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET : return Json response of labels, dataset.
    """
    today = date.today()
    start_of_week = today - timedelta(days=today.weekday())
    week_dates = [start_of_week + timedelta(days=i) for i in range(6)]

    leave_in_week = []

    leave_request = LeaveRequest.objects.filter(status="approved")
    leave_dates = []
    for leave in leave_request:
        for leave_date in leave.requested_dates():
            leave_dates.append(leave_date)

    filtered_dates = [
        day
        for day in leave_dates
        if day.month == today.month and day.year == today.year
    ]
    for week_date in week_dates:
        days = []
        for filtered_date in filtered_dates:
            if filtered_date == week_date:
                days.append(filtered_date)
        leave_in_week.append(len(days))

    dataset = (
        {
            "label": _("Leave Trends"),
            "data": leave_in_week,
        },
    )

    labels = [week_date.strftime("%d-%m-%Y") for week_date in week_dates]

    response = {
        "labels": labels,
        "dataset": dataset,
    }
    return JsonResponse(response)


@login_required
@hx_request_required
def leave_request_create(request):
    """
    function used to create leave request from calendar.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET : return leave request form template
    POST : return leave request view
    """
    previous_data = unquote(request.GET.urlencode())[len("pd=") :]
    emp = request.user.employee_get
    emp_id = emp.id

    form = UserLeaveRequestCreationForm(employee=emp)
    if request.method == "POST":
        form = UserLeaveRequestCreationForm(request.POST, request.FILES, employee=emp)
        if int(form.data["employee_id"]) == int(emp_id):
            if form.is_valid():
                leave_request = form.save(commit=False)
                save = True

                if leave_request.leave_type_id.require_approval == "no":
                    # Deduct only from this request's leave type (Casual Leave never deducts from EL).
                    employee_id = leave_request.employee_id
                    leave_type_id = leave_request.leave_type_id
                    available_leave = AvailableLeave.objects.get(
                        leave_type_id=leave_type_id, employee_id=employee_id
                    )
                    if leave_request.requested_days > available_leave.available_days:
                        leave = (
                            leave_request.requested_days
                            - available_leave.available_days
                        )
                        leave_request.approved_available_days = (
                            available_leave.available_days
                        )
                        available_leave.available_days = 0
                        available_leave.carryforward_days = (
                            available_leave.carryforward_days - leave
                        )
                        leave_request.approved_carryforward_days = leave
                    else:
                        available_leave.available_days = (
                            available_leave.available_days
                            - leave_request.requested_days
                        )
                        leave_request.approved_available_days = (
                            leave_request.requested_days
                        )
                    leave_request.status = "approved"
                    available_leave.save()
                if save:
                    leave_request.created_by = request.user.employee_get
                    leave_request.save()

                    if multiple_approvals_check(leave_request.id):
                        conditional_requests = multiple_approvals_check(
                            leave_request.id
                        )
                        managers = []
                        for manager in conditional_requests["managers"]:
                            managers.append(manager.employee_user_id)
                        with contextlib.suppress(Exception):
                            notify.send(
                                request.user.employee_get,
                                recipient=managers[0],
                                verb="You have a new leave request to validate.",
                                verb_ar="لديك طلب إجازة جديد يجب التحقق منه.",
                                verb_de="Sie haben eine neue Urlaubsanfrage zur Validierung.",
                                verb_es="Tiene una nueva solicitud de permiso que debe validar.",
                                verb_fr="Vous avez une nouvelle demande de congé à valider.",
                                icon="people-circle",
                                redirect=f"/leave/request-view?id={leave_request.id}",
                            )

                    messages.success(request, _("Leave request created successfully.."))
                    with contextlib.suppress(Exception):
                        notify.send(
                            request.user.employee_get,
                            recipient=leave_request.employee_id.employee_work_info.reporting_manager_id.employee_user_id,
                            verb=f"New leave request created for {leave_request.employee_id}.",
                            verb_ar=f"تم إنشاء طلب إجازة جديد لـ {leave_request.employee_id}.",
                            verb_de=f"Neuer Urlaubsantrag für {leave_request.employee_id} erstellt.",
                            verb_es=f"Nueva solicitud de permiso creada para {leave_request.employee_id}.",
                            verb_fr=f"Nouvelle demande de congé créée pour {leave_request.employee_id}.",
                            icon="people-circle",
                            redirect=reverse("request-view")
                            + f"?id={leave_request.id}",
                        )

                    mail_thread = LeaveMailSendThread(
                        request, leave_request, type="request"
                    )
                    mail_thread.start()
                    form = UserLeaveRequestCreationForm(employee=emp)
                    if len(LeaveRequest.objects.filter(employee_id=emp_id)) == 1:
                        return HttpResponse(
                            "<script>window.location.reload();</script>"
                        )
            return render(
                request,
                "leave/user_leave/request_form.html",
                {
                    "form": form,
                    "pd": previous_data,
                },
            )
        else:
            messages.error(request, _("You don't have permission"))
            response = render(
                request, "leave/user_leave/request_form.html", {"form": form}
            )
            return HttpResponse(
                response.content.decode("utf-8") + "<script>location.reload();</script>"
            )
    return render(
        request,
        "leave/user_leave/request_form.html",
        {
            "form": form,
            "pd": previous_data,
        },
    )


@login_required
def leave_allocation_request_view(request):
    """
    function used to view leave allocation request.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET : return leave allocation request view template
    """
    employee = request.user.employee_get
    queryset = LeaveAllocationRequest.objects.all().order_by("-id")
    queryset = LeaveAllocationRequestFilter(request.GET, queryset).qs
    queryset = filtersubordinates(
        request, queryset, "leave.view_leaveallocationrequest"
    )
    page_number = request.GET.get("page")
    leave_allocation_requests = paginator_qry(queryset, page_number)
    requests_ids = json.dumps(
        list(leave_allocation_requests.object_list.values_list("id", flat=True))
    )
    my_leave_allocation_requests = LeaveAllocationRequest.objects.filter(
        employee_id=employee.id
    ).order_by("-id")
    my_leave_allocation_requests = LeaveAllocationRequestFilter(
        request.GET, my_leave_allocation_requests
    ).qs
    my_page_number = request.GET.get("m_page")
    my_leave_allocation_requests = paginator_qry(
        my_leave_allocation_requests, my_page_number
    )
    my_requests_ids = json.dumps(
        list(my_leave_allocation_requests.object_list.values_list("id", flat=True))
    )
    leave_allocation_request_filter = LeaveAllocationRequestFilter()
    previous_data = request.GET.urlencode()
    data_dict = parse_qs(previous_data)
    data_dict = get_key_instances(LeaveAllocationRequest, data_dict)
    context = {
        "leave_allocation_requests": leave_allocation_requests,
        "my_leave_allocation_requests": my_leave_allocation_requests,
        "pd": previous_data,
        "form": leave_allocation_request_filter.form,
        "filter_dict": data_dict,
        "gp_fields": LeaveAllocationRequestReGroup.fields,
        "requests_ids": requests_ids,
        "my_requests_ids": my_requests_ids,
    }
    return render(
        request,
        "leave/leave_allocation_request/leave_allocation_request_view.html",
        context=context,
    )


@login_required
@hx_request_required
def leave_allocation_request_single_view(request, req_id):
    """
    function used to present the leave allocation request detailed view.

    Parameters:
    request (HttpRequest): The HTTP request object.
    req_id : leave allocation request id

    Returns:
    return leave allocation request single view
    """
    my_request = False
    if request.GET.get("my_request") == "True":
        my_request = True
    requests_ids_json = request.GET.get("instances_ids")
    if requests_ids_json:
        requests_ids = json.loads(requests_ids_json)
        previous_id, next_id = closest_numbers(requests_ids, req_id)
    leave_allocation_request = LeaveAllocationRequest.find(req_id)
    context = {
        "leave_allocation_request": leave_allocation_request,
        "my_request": my_request,
        "instances_ids": requests_ids_json,
        "previous": previous_id,
        "next": next_id,
        "dashboard": request.GET.get("dashboard"),
    }
    return render(
        request,
        "leave/leave_allocation_request/leave_allocation_request_single_view.html",
        context=context,
    )


@login_required
@hx_request_required
def leave_allocation_request_create(request):
    """
    function used to create leave allocation request.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    GET : return leave allocation request form template
    POST : return leave allocation request view
    """
    employee = request.user.employee_get
    form = LeaveAllocationRequestForm(initial={"employee_id": employee})
    form = choosesubordinates(request, form, "leave.add_leaveallocationrequest")
    form.fields["employee_id"].queryset = form.fields[
        "employee_id"
    ].queryset | Employee.objects.filter(employee_user_id=request.user)
    if request.method == "POST":
        form = LeaveAllocationRequestForm(request.POST, request.FILES)
        if form.is_valid():
            leave_allocation_request = form.save(commit=False)
            leave_allocation_request.skip_history = False
            leave_allocation_request.save()
            messages.success(request, _("New Leave allocation request is created"))
            with contextlib.suppress(Exception):
                notify.send(
                    request.user.employee_get,
                    recipient=leave_allocation_request.employee_id.employee_work_info.reporting_manager_id.employee_user_id,
                    verb=f"New leave allocation request created for {leave_allocation_request.employee_id}.",
                    verb_ar=f"تم إنشاء طلب تخصيص إجازة جديد لـ {leave_allocation_request.employee_id}.",
                    verb_de=f"Neue Anfrage zur Urlaubszuweisung erstellt für {leave_allocation_request.employee_id}.",
                    verb_es=f"Nueva solicitud de asignación de permisos creada para {leave_allocation_request.employee_id}.",
                    verb_fr=f"Nouvelle demande d'allocation de congé créée pour {leave_allocation_request.employee_id}.",
                    icon="people-cicle",
                    redirect=reverse("leave-allocation-request-view")
                    + f"?id={leave_allocation_request.id}",
                )
            response = render(
                request,
                "leave/leave_allocation_request/leave_allocation_request_create.html",
                {"form": form},
            )
            return HttpResponse(
                response.content.decode("utf-8")
                + "<script>location. reload();</script>"
            )
    context = {"form": form}
    return render(
        request,
        "leave/leave_allocation_request/leave_allocation_request_create.html",
        context=context,
    )


@login_required
@hx_request_required
def leave_allocation_request_filter(request):
    field = request.GET.get("field")
    employee = request.user.employee_get
    page_number = request.GET.get("page")
    my_page_number = request.GET.get("m_page")
    previous_data = request.GET.urlencode()
    template = "leave/leave_allocation_request/leave_allocation_request_list.html"

    # Filter leave allocation requests
    leave_allocation_requests_filtered = LeaveAllocationRequestFilter(
        request.GET
    ).qs.order_by("-id")
    my_leave_allocation_requests_filtered = LeaveAllocationRequest.objects.filter(
        employee_id=employee.id
    ).order_by("-id")
    my_leave_allocation_requests_filtered = LeaveAllocationRequestFilter(
        request.GET, my_leave_allocation_requests_filtered
    ).qs
    leave_allocation_requests_filtered = filtersubordinates(
        request, leave_allocation_requests_filtered, "leave.view_leaveallocationrequest"
    )

    # Sort leave allocation requests if requested
    if request.GET.get("sortby"):
        leave_allocation_requests_filtered = sortby(
            request, leave_allocation_requests_filtered, "sortby"
        )

    # Group leave allocation requests if field parameter is provided
    if field:
        leave_allocation_requests = group_by_queryset(
            leave_allocation_requests_filtered, field, page_number, "page"
        )
        my_leave_allocation_requests = group_by_queryset(
            my_leave_allocation_requests_filtered, field, my_page_number, "m_page"
        )

        # Convert IDs to JSON format for details view
        list_values = [entry["list"] for entry in leave_allocation_requests]
        id_list = [
            instance.id for value in list_values for instance in value.object_list
        ]
        requests_ids = json.dumps(list(id_list))

        list_values = [entry["list"] for entry in my_leave_allocation_requests]
        id_list = [
            instance.id for value in list_values for instance in value.object_list
        ]
        my_requests_ids = json.dumps(list(id_list))
        template = (
            "leave/leave_allocation_request/leave_allocation_request_group_by.html"
        )
    else:
        leave_allocation_requests = paginator_qry(
            leave_allocation_requests_filtered, page_number
        )
        my_leave_allocation_requests = paginator_qry(
            my_leave_allocation_requests_filtered, my_page_number
        )
        requests_ids = json.dumps(
            list(leave_allocation_requests.object_list.values_list("id", flat=True))
        )
        my_requests_ids = json.dumps(
            list(my_leave_allocation_requests.object_list.values_list("id", flat=True))
        )

    # Parse previous data and construct context for filter tag
    data_dict = parse_qs(previous_data)
    data_dict = get_key_instances(LeaveAllocationRequest, data_dict)
    data_dict.pop("m_page", None)

    context = {
        "leave_allocation_requests": leave_allocation_requests,
        "my_leave_allocation_requests": my_leave_allocation_requests,
        "pd": previous_data,
        "filter_dict": data_dict,
        "field": field,
        "requests_ids": requests_ids,
        "my_requests_ids": my_requests_ids,
    }
    return render(request, template, context=context)


@login_required
@hx_request_required
@leave_allocation_change_permission()
def leave_allocation_request_update(request, req_id):
    """
    function used to update leave allocation request.

    Parameters:
    request (HttpRequest): The HTTP request object.
    req_id : leave allocation request id

    Returns:
    GET : return leave allocation request update template
    POST : return leave allocation request view
    """
    leave_allocation_request = LeaveAllocationRequest.objects.get(id=req_id)
    form = LeaveAllocationRequestForm(instance=leave_allocation_request)
    form = choosesubordinates(request, form, "leave.add_leaveallocationrequest")
    form.fields["employee_id"].queryset = form.fields[
        "employee_id"
    ].queryset | Employee.objects.filter(employee_user_id=request.user)
    if leave_allocation_request.status != "approved":
        if request.method == "POST":
            form = LeaveAllocationRequestForm(
                request.POST, request.FILES, instance=leave_allocation_request
            )
            if form.is_valid():
                leave_allocation_request = form.save(commit=False)
                leave_allocation_request.skip_history = False
                leave_allocation_request.save()
                messages.success(
                    request, _("Leave allocation request is updated successfully.")
                )
                with contextlib.suppress(Exception):
                    notify.send(
                        request.user.employee_get,
                        recipient=leave_allocation_request.employee_id.employee_work_info.reporting_manager_id.employee_user_id,
                        verb=f"Leave allocation request updated for {leave_allocation_request.employee_id}.",
                        verb_ar=f"تم تحديث طلب تخصيص الإجازة لـ {leave_allocation_request.employee_id}.",
                        verb_de=f"Urlaubszuteilungsanforderung aktualisiert für {leave_allocation_request.employee_id}.",
                        verb_es=f"Solicitud de asignación de licencia actualizada para {leave_allocation_request.employee_id}.",
                        verb_fr=f"Demande d'allocation de congé mise à jour pour {leave_allocation_request.employee_id}.",
                        icon="people-cicle",
                        redirect=reverse("leave-allocation-request-view")
                        + f"?id={leave_allocation_request.id}",
                    )
                response = render(
                    request,
                    "leave/leave_allocation_request/leave_allocation_request_update.html",
                    {"form": form, "req_id": req_id},
                )
                return HttpResponse(
                    response.content.decode("utf-8")
                    + "<script>location. reload();</script>"
                )
        return render(
            request,
            "leave/leave_allocation_request/leave_allocation_request_update.html",
            {"form": form, "req_id": req_id},
        )
    else:
        messages.error(request, _("You can't update this request..."))
        return HttpResponse("<script>window.location.reload();</script>")


@login_required
@leave_allocation_reject_permission()
def leave_allocation_request_approve(request, req_id):
    """
    function used to approve a leave allocation request.

    Parameters:
    request (HttpRequest): The HTTP request object.
    req_id : leave allocation request id

    Returns:
    GET :It returns to the default leave allocation request view template.
    """
    leave_allocation_request = LeaveAllocationRequest.objects.get(id=req_id)
    if leave_allocation_request.status == "requested":
        employee = leave_allocation_request.employee_id
        if (
            employee.available_leave.all()
            .filter(leave_type_id=leave_allocation_request.leave_type_id)
            .exists()
        ):
            available_leave = (
                employee.available_leave.all()
                .filter(leave_type_id=leave_allocation_request.leave_type_id)
                .first()
            )
        else:
            available_leave = AvailableLeave(
                leave_type_id=leave_allocation_request.leave_type_id,
                employee_id=employee,
            )
        available_leave.available_days += leave_allocation_request.requested_days
        available_leave.save()
        leave_allocation_request.status = "approved"
        leave_allocation_request.save()
        messages.success(request, _("Leave allocation request approved successfully"))
        with contextlib.suppress(Exception):
            notify.send(
                request.user.employee_get,
                recipient=leave_allocation_request.employee_id.employee_user_id,
                verb="Your leave allocation request has been approved",
                verb_ar="تمت الموافقة على طلب تخصيص إجازتك",
                verb_de="Ihr Antrag auf Urlaubszuweisung wurde genehmigt",
                verb_es="Se ha aprobado su solicitud de asignación de vacaciones",
                verb_fr="Votre demande d'allocation de congé a été approuvée",
                icon="people-circle",
                redirect=reverse("leave-allocation-request-view")
                + f"?id={leave_allocation_request.id}",
            )
    else:
        messages.error(request, _("The leave allocation request can't be approved"))
    return HttpResponseRedirect(request.META.get("HTTP_REFERER", "/"))


@login_required
@hx_request_required
@leave_allocation_reject_permission()
def leave_allocation_request_reject(request, req_id):
    """
    function used to Reject leave allocation request.

    Parameters:
    request (HttpRequest): The HTTP request object.
    id : leave allocation request id

    Returns:
    GET : It returns to the default leave allocation request view template.

    """
    leave_allocation_request = LeaveAllocationRequest.objects.get(id=req_id)
    if (
        leave_allocation_request.status == "requested"
        or leave_allocation_request.status == "approved"
    ):
        form = LeaveAllocationRequestRejectForm()
        if request.method == "POST":
            form = LeaveAllocationRequestRejectForm(request.POST)
            if form.is_valid():
                leave_allocation_request.reject_reason = form.cleaned_data["reason"]
                if leave_allocation_request.status == "approved":
                    leave_type = leave_allocation_request.leave_type_id
                    requested_days = leave_allocation_request.requested_days
                    available_leave = AvailableLeave.objects.filter(
                        leave_type_id=leave_type,
                        employee_id=leave_allocation_request.employee_id,
                    ).first()
                    available_leave.available_days = max(
                        0, available_leave.available_days - requested_days
                    )

                    available_leave.save()
                leave_allocation_request.status = "rejected"
                leave_allocation_request.save()
                messages.success(
                    request, _("Leave allocation request rejected successfully")
                )
                with contextlib.suppress(Exception):
                    notify.send(
                        request.user.employee_get,
                        recipient=leave_allocation_request.employee_id.employee_user_id,
                        verb="Your leave allocation request has been rejected",
                        verb_ar="تم رفض طلب تخصيص إجازتك",
                        verb_de="Ihr Antrag auf Urlaubszuweisung wurde abgelehnt",
                        verb_es="Se ha rechazado su solicitud de asignación de vacaciones",
                        verb_fr="Votre demande d'allocation de congé a été rejetée",
                        icon="people-circle",
                        redirect=reverse("leave-allocation-request-view")
                        + f"?id={leave_allocation_request.id}",
                    )
                return HttpResponse("<script>location.reload();</script>")
        return render(
            request,
            "leave/leave_allocation_request/leave_allocation_request_reject_form.html",
            {"form": form, "req_id": req_id},
        )
    else:
        messages.error(request, _("The leave allocation request can't be rejected"))
        return HttpResponse("<script>location.reload();</script>")


@login_required
@hx_request_required
@leave_allocation_delete_permission()
def leave_allocation_request_delete(request, req_id):
    """
    function used to delete leave allocation request.

    Parameters:
    request (HttpRequest): The HTTP request object.
    id : leave allocation request id

    Returns:
    GET : return leave allocation request view template
    """

    try:
        leave_allocation_request = LeaveAllocationRequest.objects.get(id=req_id)

        if leave_allocation_request.status != "approved":
            leave_allocation_request.delete()
            messages.success(
                request, _("Leave allocation request deleted successfully..")
            )
        else:
            messages.error(request, _("Approved request can't be deleted."))
    except LeaveAllocationRequest.DoesNotExist:
        messages.error(request, _("Leave allocation request not found."))
    except OverflowError:
        messages.error(request, _("Leave allocation request not found."))

    except ProtectedError:
        messages.error(request, _("Related entries exist"))
    hx_target = request.META.get("HTTP_HX_TARGET")
    previous_data = request.GET.urlencode()
    if hx_target and hx_target == "view-container":
        leave_allocations = LeaveAllocationRequest.objects.all()
        if leave_allocations.exists():
            return redirect(f"/leave/leave-allocation-request-filter?{previous_data}")
        else:
            return HttpResponse("<script>location.reload();</script>")
    elif hx_target and hx_target == "objectDetailsModalW25Target":
        instances_ids = request.GET.get("instances_ids")
        instances_list = json.loads(instances_ids)
        if req_id in instances_list:
            instances_list.remove(req_id)
        previous_instance, next_instance = closest_numbers(
            json.loads(instances_ids), req_id
        )
        return redirect(
            f"/leave/leave-allocation-request-single-view/{next_instance}?{previous_data}"
        )

    return redirect(leave_allocation_request_view)


@login_required
def assigned_leave_select(request):
    page_number = request.GET.get("page")

    if page_number == "all":
        if request.user.has_perm("leave.view_availableleave"):
            employees = AvailableLeave.objects.all()
        else:
            employees = AvailableLeave.objects.filter(
                employee_id__employee_work_info__reporting_manager_id__employee_user_id=request.user
            )

    employee_ids = [str(emp.id) for emp in employees]
    total_count = employees.count()

    context = {"employee_ids": employee_ids, "total_count": total_count}

    return JsonResponse(context, safe=False)


@login_required
def assigned_leave_select_filter(request):
    page_number = request.GET.get("page")
    filtered = request.GET.get("filter")
    filters = json.loads(filtered) if filtered else {}

    if page_number == "all":
        if request.user.has_perm("leave.view_availableleave"):
            employee_filter = AssignedLeaveFilter(
                filters, queryset=AvailableLeave.objects.all()
            )
        else:
            employee_filter = AssignedLeaveFilter(
                filters,
                queryset=AvailableLeave.objects.filter(
                    employee_id__employee_work_info__reporting_manager_id__employee_user_id=request.user
                ),
            )

        # Get the filtered queryset
        filtered_employees = employee_filter.qs

        employee_ids = [str(emp.id) for emp in filtered_employees]
        total_count = filtered_employees.count()

        context = {"employee_ids": employee_ids, "total_count": total_count}

        return JsonResponse(context)


@login_required
@require_http_methods(["POST"])
@manager_can_enter("leave.delete_leaverequest")
def leave_request_bulk_delete(request):
    """
    This method is used to delete a bulk of leave requests.
    """
    ids = request.POST["ids"]
    ids = json.loads(ids)
    count = 0  # To track the number of successfully deleted requests
    for leave_request_id in ids:
        try:
            leave_request = LeaveRequest.objects.get(id=leave_request_id)
            employee = leave_request.employee_id
            if leave_request.status == "requested":
                leave_request.delete()
                count += 1
            else:
                messages.error(
                    request,
                    _("{}'s leave request cannot be deleted.".format(employee)),
                )
        except Exception as e:
            messages.error(request, _("An error occurred: {}.".format(str(e))))

    if count > 0:
        messages.success(
            request,
            _("{count}  leave request(s) successfully deleted.".format(count=count)),
        )

    return JsonResponse({"message": "Success"})


@login_required
def leave_request_select(request):
    page_number = request.GET.get("page")

    if page_number == "all":
        if request.user.has_perm("leave.view_leaverequest"):
            employees = LeaveRequest.objects.all()
        else:
            employees = LeaveRequest.objects.filter(
                employee_id__employee_work_info__reporting_manager_id__employee_user_id=request.user
            )

    employee_ids = [str(emp.id) for emp in employees]
    total_count = employees.count()

    context = {"employee_ids": employee_ids, "total_count": total_count}

    return JsonResponse(context, safe=False)


@login_required
def leave_request_select_filter(request):
    page_number = request.GET.get("page")
    filtered = request.GET.get("filter")
    filters = json.loads(filtered) if filtered else {}

    if page_number == "all":
        if request.user.has_perm("leave.view_leaverequest"):
            employee_filter = LeaveRequestFilter(
                filters, queryset=LeaveRequest.objects.all()
            )
        else:
            employee_filter = LeaveRequestFilter(
                filters,
                queryset=LeaveRequest.objects.filter(
                    employee_id__employee_work_info__reporting_manager_id__employee_user_id=request.user
                ),
            )

        # Get the filtered queryset
        filtered_employees = employee_filter.qs

        employee_ids = [str(emp.id) for emp in filtered_employees]
        total_count = filtered_employees.count()

        context = {"employee_ids": employee_ids, "total_count": total_count}

        return JsonResponse(context)


@require_http_methods(["POST"])
@login_required
def user_request_bulk_delete(request):
    """
    This method is used to delete bulk of leaves requests
    """
    ids = request.POST["ids"]
    ids = json.loads(ids)
    for leave_request_id in ids:
        try:
            leave_request = LeaveRequest.objects.get(id=leave_request_id)
            status = leave_request.status
            if leave_request.status == "requested":
                leave_request.delete()
                messages.success(
                    request,
                    _("Leave request deleted."),
                )
            else:
                messages.error(
                    request,
                    _("You cannot delete leave request with status {}.".format(status)),
                )
        except Exception as e:
            messages.error(request, _("Leave request not found."))
    return JsonResponse({"message": "Success"})


@login_required
def user_request_select(request):
    page_number = request.GET.get("page")
    user = request.user.employee_get

    if page_number == "all":
        employees = LeaveRequest.objects.filter(employee_id=user)

    employee_ids = [str(emp.id) for emp in employees]
    total_count = employees.count()

    context = {"employee_ids": employee_ids, "total_count": total_count}

    return JsonResponse(context, safe=False)


@login_required
def user_request_select_filter(request):
    page_number = request.GET.get("page")
    filtered = request.GET.get("filter")
    filters = json.loads(filtered) if filtered else {}
    user = request.user.employee_get

    if page_number == "all":
        employee_filter = UserLeaveRequestFilter(
            filters, queryset=LeaveRequest.objects.filter(employee_id=user)
        )

        # Get the filtered queryset
        filtered_employees = employee_filter.qs

        employee_ids = [str(emp.id) for emp in filtered_employees]
        total_count = filtered_employees.count()

        context = {"employee_ids": employee_ids, "total_count": total_count}

        return JsonResponse(context)


@login_required
@hx_request_required
def employee_available_leave_count(request):
    leave_type_id = request.GET.get("leave_type_id")
    hx_target = request.META.get("HTTP_HX_TARGET")
    start_date_str = request.GET.get("start_date")

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        start_date = None

    if not leave_type_id:
        return render(
            request,
            "leave/leave_request/employee_available_leave_count.html",
            {"hx_target": hx_target},
        )

    employee_id = request.GET.getlist("employee_id")
    employee_id = employee_id[0] if employee_id else None
    # For user's own leave request form, ensure we have employee (HTMX may not always send hidden field)
    if not employee_id and leave_type_id and request.user.is_authenticated:
        try:
            employee_id = request.user.employee_get.id
        except Exception:
            pass

    available_leave = (
        AvailableLeave.objects.select_related("leave_type_id", "employee_id")
        .filter(leave_type_id=leave_type_id, employee_id=employee_id)
        .first()
    )

    total_leave_days = 0
    forcasted_days = 0
    pending_requests_days = 0

    def _pending_leave_days(emp, lt_id):
        if not emp:
            return 0
        return (
            emp.leaverequest_set.filter(
                status="requested",
                leave_type_id=lt_id,
            ).aggregate(total=Sum("requested_days"))["total"]
            or 0
        )

    from leave.intern_leave import intern_isl_balance

    try:
        isl_type = LeaveType.objects.get(name="Intern Sick Leave (ISL)")
    except LeaveType.DoesNotExist:
        isl_type = None

    employee = None
    if employee_id:
        from employee.models import Employee
        employee = Employee.objects.filter(id=employee_id).first()

    is_intern_isl = (
        isl_type
        and str(leave_type_id) == str(isl_type.id)
        and employee
        and is_intern(employee)
    )

    if is_intern_isl:
        balance, _, _ = intern_isl_balance(employee)
        total_leave_days = round(max(balance + 0, 0), 3)
        if employee:
            pending_requests_days = _pending_leave_days(employee, leave_type_id)
        if available_leave is None:
            available_leave = object()
    elif available_leave:
        leave_type = available_leave.leave_type_id
        # Earned Leave (monthly 1.25): show computed balance as of today so it's correct even if scheduler hasn't run
        EL_EPOCH_DATE = date(2026, 1, 1)
        lt_name = (leave_type.name or "").strip()
        reset_based = getattr(leave_type, "reset_based", None) or ""
        total_days_val = round(float(leave_type.total_days or 0), 2)
        # Earned Leave: monthly 1.25 — match by name + monthly reset, or by 1.25/month (even if reset_based missing in DB)
        is_el_monthly = (
            (reset_based == "monthly" and ("Earned Leave" in lt_name or total_days_val == 1.25))
            or ("Earned Leave" in lt_name and total_days_val == 1.25)
        )
        # Casual Leave: monthly 1/day, carryforward cap 12 (exclude Probation Casual Leave)
        is_cl_monthly = (
            reset_based == "monthly"
            and total_days_val == 1.0
            and "probation" not in lt_name.lower()
            and ("Casual Leave" in lt_name or "Casual" in lt_name)
        )
        # Probation Leave / Probation Casual Leave / Interns Leave: monthly 1/day, no carryforward
        lt_name_lower_av = lt_name.lower()
        is_pl_monthly = (
            total_days_val == 1.0
            and (
                (reset_based == "monthly" and ("probation" in lt_name_lower_av or "interns leave" in lt_name_lower_av or "intern" in lt_name_lower_av))
                or ("probation" in lt_name_lower_av or "interns leave" in lt_name_lower_av or "intern" in lt_name_lower_av)
            )
        )
        if is_el_monthly and leave_type.total_days:
            # Accrual starts from 1st of month: max(Jan 1 2026, 1st of assign month)
            raw_start = getattr(available_leave, "assigned_date", None)
            if raw_start:
                first_of_assign = raw_start.replace(day=1)
                accrual_start = max(EL_EPOCH_DATE, first_of_assign)
            else:
                accrual_start = EL_EPOCH_DATE
            today = date.today()
            months = (today.year - accrual_start.year) * 12 + (today.month - accrual_start.month) + 1
            months = max(0, months)
            accrual_amount = float(leave_type.total_days)
            total_accrued = accrual_amount * months
            cap = leave_type.carryforward_max if leave_type.carryforward_max is not None else 9999
            approved_taken = LeaveRequest.objects.filter(
                employee_id=available_leave.employee_id,
                leave_type_id=leave_type,
                status="approved",
            ).aggregate(total=Sum("requested_days"))["total"] or 0
            balance = total_accrued - approved_taken
            balance = min(max(0.0, balance), cap)
            total_leave_days = round(balance, 3)
        elif is_cl_monthly and leave_type.total_days:
            # Casual Leave: months accrued minus PCL taken minus CL taken
            from leave.methods import get_casual_leave_balance_stats

            total_leave_days = get_casual_leave_balance_stats(available_leave)[
                "available_days"
            ]
        else:
            # Probation Leave & Interns Leave: same logic — use computed balance (cap 2, no carry forward). Case-insensitive.
            is_pl_by_name = (
                total_days_val == 1.0
                and ("probation" in lt_name_lower_av or "interns leave" in lt_name_lower_av or "intern" in lt_name_lower_av)
            )
            computed = computed_balance_for_validation(available_leave, as_of_date=date.today())
            if computed is not None and is_pl_by_name:
                # Probation/Interns: max 1 day per calendar month — cap display by usable days in the relevant month
                ref_date = start_date if start_date else date.today()
                days_in_month = probation_leave_days_in_month(
                    available_leave.employee_id_id,
                    leave_type.id,
                    ref_date.year,
                    ref_date.month,
                )
                usable_in_month = max(0.0, 1.0 - float(days_in_month))
                total_leave_days = round(min(computed, usable_in_month), 3)
            elif is_pl_monthly and leave_type.total_days:
                # Fallback if computed returned None — still cap by per-month usable
                raw_start = getattr(available_leave, "assigned_date", None)
                accrual_start = (
                    raw_start.replace(day=1) if raw_start else date.today().replace(day=1)
                )
                today = date.today()
                months = (today.year - accrual_start.year) * 12 + (today.month - accrual_start.month) + 1
                months = max(0, months)
                total_accrued = 1.0 * months
                approved_taken = LeaveRequest.objects.filter(
                    employee_id=available_leave.employee_id,
                    leave_type_id=leave_type,
                    status="approved",
                ).aggregate(total=Sum("requested_days"))["total"] or 0
                balance = min(max(0.0, total_accrued - approved_taken), 2.0)
                ref_date = start_date if start_date else date.today()
                days_in_month = probation_leave_days_in_month(
                    available_leave.employee_id_id,
                    leave_type.id,
                    ref_date.year,
                    ref_date.month,
                )
                usable_in_month = max(0.0, 1.0 - float(days_in_month))
                total_leave_days = round(min(balance, usable_in_month), 3)
            else:
                # Other leave types: show stored balance
                total_leave_days = round(
                    (available_leave.available_days or 0) + (available_leave.carryforward_days or 0),
                    3,
                )
        next_reset = leave_type.leave_type_next_reset_date()
        if next_reset and start_date and start_date >= next_reset:
            forcasted_days = available_leave.forcasted_leaves(start_date)
            if forcasted_days and forcasted_days > 0:
                pass
        if available_leave.employee_id_id:
            pending_requests_days = _pending_leave_days(
                available_leave.employee_id, leave_type_id
            )

    context = {
        "hx_target": hx_target,
        "leave_type_id": leave_type_id,
        "available_leave": available_leave,
        "total_leave_days": total_leave_days,
        "forcasted_days": forcasted_days,
        "pending_requests": pending_requests_days,
    }
    return render(
        request, "leave/leave_request/employee_available_leave_count.html", context
    )


@login_required
@hx_request_required
@manager_can_enter("base.add_penaltyaccounts")
def cut_available_leave(request, instance_id):
    """
    This method is used to create the penalties
    """
    previous_data = request.GET.urlencode()
    instance = LeaveRequest.objects.get(id=instance_id)
    form = PenaltyAccountForm(employee=instance.employee_id)
    available = AvailableLeave.objects.filter(employee_id=instance.employee_id)
    if request.method == "POST":
        form = PenaltyAccountForm(request.POST)
        if form.is_valid():
            penalty_instance = form.instance
            penalty = PenaltyAccounts()
            penalty.leave_request_id = instance
            penalty.deduct_from_carry_forward = (
                penalty_instance.deduct_from_carry_forward
            )
            penalty.employee_id = instance.employee_id
            penalty.leave_type_id = penalty_instance.leave_type_id
            penalty.minus_leaves = penalty_instance.minus_leaves
            penalty.penalty_amount = penalty_instance.penalty_amount
            penalty.save()
            form = PenaltyAccountForm()
            messages.success(request, "Penalty/Fine added")
    return render(
        request,
        "leave/leave_request/penalty/form.html",
        {
            "available": available,
            "form": form,
            "instance": instance,
            "pd": previous_data,
        },
    )


@login_required
@hx_request_required
def create_leaverequest_comment(request, leave_id):
    """
    This method renders form and template to create Leave request comments
    """
    leave = LeaveRequest.objects.filter(id=leave_id).first()
    emp = request.user.employee_get
    form = LeaverequestcommentForm(
        initial={"employee_id": emp.id, "request_id": leave_id}
    )
    target = request.GET.get("target")
    url = "request-filter" if target == "leaveRequest" else "user-request-filter"
    previous_data = request.GET.urlencode()
    if request.method == "POST":
        form = LeaverequestcommentForm(request.POST)
        if form.is_valid():
            form.instance.employee_id = emp
            form.instance.request_id = leave
            form.save()
            comments = LeaverequestComment.objects.filter(request_id=leave_id).order_by(
                "-created_at"
            )
            no_comments = False
            if not comments.exists():
                no_comments = True
            form = LeaverequestcommentForm(
                initial={"employee_id": emp.id, "request_id": leave_id}
            )
            messages.success(request, _("Comment added successfully!"))
            work_info = EmployeeWorkInformation.objects.filter(
                employee_id=leave.employee_id
            )
            if work_info.exists():
                if (
                    leave.employee_id.employee_work_info.reporting_manager_id
                    is not None
                ):
                    if request.user.employee_get.id == leave.employee_id.id:
                        rec = (
                            leave.employee_id.employee_work_info.reporting_manager_id.employee_user_id
                        )
                        notify.send(
                            request.user.employee_get,
                            recipient=rec,
                            verb=f"{leave.employee_id}'s leave request has received a comment.",
                            verb_ar=f"تلقت طلب إجازة {leave.employee_id} تعليقًا.",
                            verb_de=f"{leave.employee_id}s Urlaubsantrag hat einen Kommentar erhalten.",
                            verb_es=f"La solicitud de permiso de {leave.employee_id} ha recibido un comentario.",
                            verb_fr=f"La demande de congé de {leave.employee_id} a reçu un commentaire.",
                            redirect=reverse("request-view") + f"?id={leave.id}",
                            icon="chatbox-ellipses",
                        )
                    elif (
                        request.user.employee_get.id
                        == leave.employee_id.employee_work_info.reporting_manager_id.id
                    ):
                        rec = leave.employee_id.employee_user_id
                        notify.send(
                            request.user.employee_get,
                            recipient=rec,
                            verb="Your leave request has received a comment.",
                            verb_ar="تلقى طلب إجازتك تعليقًا.",
                            verb_de="Ihr Urlaubsantrag hat einen Kommentar erhalten.",
                            verb_es="Tu solicitud de permiso ha recibido un comentario.",
                            verb_fr="Votre demande de congé a reçu un commentaire.",
                            redirect=reverse("user-request-view") + f"?id={leave.id}",
                            icon="chatbox-ellipses",
                        )
                    else:
                        rec = [
                            leave.employee_id.employee_user_id,
                            leave.employee_id.employee_work_info.reporting_manager_id.employee_user_id,
                        ]
                        notify.send(
                            request.user.employee_get,
                            recipient=rec,
                            verb=f"{leave.employee_id}'s leave request has received a comment.",
                            verb_ar=f"تلقت طلب إجازة {leave.employee_id} تعليقًا.",
                            verb_de=f"{leave.employee_id}s Urlaubsantrag hat einen Kommentar erhalten.",
                            verb_es=f"La solicitud de permiso de {leave.employee_id} ha recibido un comentario.",
                            verb_fr=f"La demande de congé de {leave.employee_id} a reçu un commentaire.",
                            redirect=reverse("request-view") + f"?id={leave.id}",
                            icon="chatbox-ellipses",
                        )
                else:
                    rec = leave.employee_id.employee_user_id
                    notify.send(
                        request.user.employee_get,
                        recipient=rec,
                        verb="Your leave request has received a comment.",
                        verb_ar="تلقى طلب إجازتك تعليقًا.",
                        verb_de="Ihr Urlaubsantrag hat einen Kommentar erhalten.",
                        verb_es="Tu solicitud de permiso ha recibido un comentario.",
                        verb_fr="Votre demande de congé a reçu un commentaire.",
                        redirect=reverse("user-request-view") + f"?id={leave.id}",
                        icon="chatbox-ellipses",
                    )
            return render(
                request,
                "leave/leave_request/leave_comment.html",
                {
                    "comments": comments,
                    "no_comments": no_comments,
                    "request_id": leave_id,
                    "leave_request": leave,
                },
            )
    return render(
        request,
        "leave/leave_request/leave_comment.html",
        {
            "form": form,
            "request_id": leave_id,
            "pd": previous_data,
            "target": target,
            "url": url,
            "leave_request": leave,
        },
    )


@login_required
@hx_request_required
def view_leaverequest_comment(request, leave_id):
    """
    This method is used to show Leave request comments
    """
    leave_request = LeaveRequest.find(leave_id)
    if not (
        request.user.employee_get == leave_request.employee_id
        or request.user.has_perm("leave.view_leaverequestcomment")
        or is_reportingmanager(request)
    ):
        messages.warning(request, _("You don't have permission"))
        return render(request, "decorator_404.html")

    comments = LeaverequestComment.objects.filter(request_id=leave_id).order_by(
        "-created_at"
    )
    no_comments = False
    if not comments.exists():
        no_comments = True

    if request.FILES:
        files = request.FILES.getlist("files")
        comment_id = request.GET["comment_id"]
        comment = LeaverequestComment.objects.get(id=comment_id)
        attachments = []
        for file in files:
            file_instance = LeaverequestFile()
            file_instance.file = file
            file_instance.save()
            attachments.append(file_instance)
        comment.files.add(*attachments)

    return render(
        request,
        "leave/leave_request/leave_comment.html",
        {
            "comments": comments,
            "no_comments": no_comments,
            "leave_request": leave_request,
        },
    )


@login_required
@hx_request_required
def create_allocationrequest_comment(request, leave_id):
    """
    This method renders form and template to create Allocation request comments
    """
    previous_data = request.GET.urlencode()
    leave = LeaveAllocationRequest.objects.filter(id=leave_id).first()
    emp = request.user.employee_get
    form = LeaveallocationrequestcommentForm(
        initial={"employee_id": emp.id, "request_id": leave_id}
    )

    if request.method == "POST":
        form = LeaveallocationrequestcommentForm(request.POST)
        if form.is_valid():
            form.instance.employee_id = emp
            form.instance.request_id = leave
            form.save()
            comments = LeaveallocationrequestComment.objects.filter(
                request_id=leave_id
            ).order_by("-created_at")
            no_comments = False
            if not comments.exists():
                no_comments = True
            form = LeaveallocationrequestcommentForm(
                initial={"employee_id": emp.id, "request_id": leave_id}
            )
            messages.success(request, _("Comment added successfully!"))
            work_info = EmployeeWorkInformation.objects.filter(
                employee_id=leave.employee_id
            )
            if work_info.exists():
                if (
                    leave.employee_id.employee_work_info.reporting_manager_id
                    is not None
                ):
                    if request.user.employee_get.id == leave.employee_id.id:
                        rec = (
                            leave.employee_id.employee_work_info.reporting_manager_id.employee_user_id
                        )
                        notify.send(
                            request.user.employee_get,
                            recipient=rec,
                            verb=f"{leave.employee_id}'s leave allocation request has received a comment.",
                            verb_ar=f"تلقت طلب تخصيص الإجازة لـ {leave.employee_id} تعليقًا.",
                            verb_de=f"{leave.employee_id}s Anfrage zur Urlaubszuweisung hat einen Kommentar erhalten.",
                            verb_es=f"La solicitud de asignación de permisos de {leave.employee_id} ha recibido un comentario.",
                            verb_fr=f"La demande d'allocation de congé de {leave.employee_id} a reçu un commentaire.",
                            redirect=reverse("leave-allocation-request-view")
                            + f"?id={leave.id}",
                            icon="chatbox-ellipses",
                        )
                    elif (
                        request.user.employee_get.id
                        == leave.employee_id.employee_work_info.reporting_manager_id.id
                    ):
                        rec = leave.employee_id.employee_user_id
                        notify.send(
                            request.user.employee_get,
                            recipient=rec,
                            verb="Your leave allocation request has received a comment.",
                            verb_ar="تلقى طلب تخصيص الإجازة الخاص بك تعليقًا.",
                            verb_de="Ihr Antrag auf Urlaubszuweisung hat einen Kommentar erhalten.",
                            verb_es="Tu solicitud de asignación de permisos ha recibido un comentario.",
                            verb_fr="Votre demande d'allocation de congé a reçu un commentaire.",
                            redirect=reverse("leave-allocation-request-view")
                            + f"?id={leave.id}",
                            icon="chatbox-ellipses",
                        )
                    else:
                        rec = [
                            leave.employee_id.employee_user_id,
                            leave.employee_id.employee_work_info.reporting_manager_id.employee_user_id,
                        ]
                        notify.send(
                            request.user.employee_get,
                            recipient=rec,
                            verb=f"{leave.employee_id}'s leave allocation request has received a comment.",
                            verb_ar=f"تلقت طلب تخصيص الإجازة لـ {leave.employee_id} تعليقًا.",
                            verb_de=f"{leave.employee_id}s Anfrage zur Urlaubszuweisung hat einen Kommentar erhalten.",
                            verb_es=f"La solicitud de asignación de permisos de {leave.employee_id} ha recibido un comentario.",
                            verb_fr=f"La demande d'allocation de congé de {leave.employee_id} a reçu un commentaire.",
                            redirect=reverse("leave-allocation-request-view")
                            + f"?id={leave.id}",
                            icon="chatbox-ellipses",
                        )
                else:
                    rec = leave.employee_id.employee_user_id
                    notify.send(
                        request.user.employee_get,
                        recipient=rec,
                        verb="Your leave allocation request has received a comment.",
                        verb_ar="تلقى طلب تخصيص الإجازة الخاص بك تعليقًا.",
                        verb_de="Ihr Antrag auf Urlaubszuweisung hat einen Kommentar erhalten.",
                        verb_es="Tu solicitud de asignación de permisos ha recibido un comentario.",
                        verb_fr="Votre demande d'allocation de congé a reçu un commentaire.",
                        redirect=reverse("leave-allocation-request-view")
                        + f"?id={leave.id}",
                        icon="chatbox-ellipses",
                    )
            return render(
                request,
                "leave/leave_allocation_request/leave_allocation_comment.html",
                {
                    "comments": comments,
                    "no_comments": no_comments,
                    "request_id": leave_id,
                },
            )
    return render(
        request,
        "leave/leave_allocation_request/leave_allocation_comment.html",
        {"form": form, "request_id": leave_id, "pd": previous_data},
    )


@login_required
@hx_request_required
def view_allocationrequest_comment(request, leave_id):
    """
    This method is used to show Allocation request comments
    """
    leave_alloc_request = LeaveAllocationRequest.find(leave_id)
    if not (
        request.user.employee_get == leave_alloc_request.employee_id
        or request.user.has_perm("leave.view_leaveallocationrequestcomment")
        or is_reportingmanager(request)
    ):
        messages.warning(request, _("You don't have permission"))
        return render(request, "decorator_404.html")
    comments = LeaveallocationrequestComment.objects.filter(
        request_id=leave_id
    ).order_by("-created_at")
    no_comments = False
    if not comments.exists():
        no_comments = True

    if request.FILES:
        files = request.FILES.getlist("files")
        comment_id = request.GET["comment_id"]
        comment = LeaveallocationrequestComment.objects.get(id=comment_id)
        attachments = []
        for file in files:
            file_instance = LeaverequestFile()
            file_instance.file = file
            file_instance.save()
            attachments.append(file_instance)
        comment.files.add(*attachments)

    return render(
        request,
        "leave/leave_allocation_request/leave_allocation_comment.html",
        {
            "comments": comments,
            "no_comments": no_comments,
            "request_id": leave_id,
            "leave_alloc_request": leave_alloc_request,
        },
    )


@login_required
@hx_request_required
def delete_allocationrequest_comment(request, comment_id):
    """
    This method is used to delete Allocation request comments
    """
    script = ""
    comment = LeaveallocationrequestComment.find(comment_id)
    request_id = comment.request_id.id
    if (
        request.user.employee_get == comment.employee_id
        or request.user.has_perm("leave.delete_leaveallocationrequestcomment")
        or is_reportingmanager(request)
    ):
        comment.delete()
        messages.success(request, _("Comment deleted successfully!"))
    else:
        script = f"""
                    <span hx-get="/leave/allocation-request-view-comment/{request_id}/" hx-target="#commentContainer" hx-trigger="load"></span>
                """
        messages.warning(request, _("You don't have permission"))
    return HttpResponse(script)


@login_required
def delete_allocation_comment_file(request):
    """
    Used to delete attachment
    """
    script = ""
    ids = request.GET.getlist("ids")
    leave_id = request.GET["leave_id"]
    comment_id = request.GET["comment_id"]
    comment = LeaveallocationrequestComment.find(comment_id)
    if (
        request.user.employee_get == comment.employee_id
        or request.user.has_perm("leave.delete_leaverequestfile")
        or is_reportingmanager(request)
    ):
        LeaverequestFile.objects.filter(id__in=ids).delete()
        messages.success(request, _("File deleted successfully"))
    else:
        messages.warning(request, _("You don't have permission"))
        script = f"""
                <span hx-get='/leave/allocation-request-view-comment/{leave_id}/' hx-target='#commentContainer' hx-trigger='load'></span>
                """
    return HttpResponse(script)


@login_required
@hx_request_required
def view_clashes(request, leave_request_id):
    """
    This method is used to filter or view the leave clashes
    """
    record = get_object_or_404(LeaveRequest, id=leave_request_id)

    if record.status == "rejected" or record.status == "cancelled":
        overlapping_requests = LeaveRequest.objects.none()
        clashed_due_to_department = LeaveRequest.objects.none()
        clashed_due_to_job_position = LeaveRequest.objects.none()
    else:
        overlapping_requests = (
            LeaveRequest.objects.filter(
                (
                    Q(
                        employee_id__employee_work_info__department_id=record.employee_id.employee_work_info.department_id
                    )
                    | Q(
                        employee_id__employee_work_info__job_position_id=record.employee_id.employee_work_info.job_position_id
                    )
                )
                & Q(
                    employee_id__employee_work_info__company_id=record.employee_id.employee_work_info.company_id
                ),
                start_date__lte=record.end_date,
                end_date__gte=record.start_date,
            )
            .exclude(id=leave_request_id)
            .exclude(Q(status="cancelled") | Q(status="rejected"))
        )

        clashed_due_to_department = overlapping_requests.filter(
            employee_id__employee_work_info__department_id=record.employee_id.employee_work_info.department_id
        )

        clashed_due_to_job_position = overlapping_requests.filter(
            employee_id__employee_work_info__job_position_id=record.employee_id.employee_work_info.job_position_id
        )

    leave_request_filter = LeaveRequestFilter(request.GET, overlapping_requests).qs
    leave_request_filter = paginator_qry(leave_request_filter, request.GET.get("page"))

    requests_ids = json.dumps(
        [instance.id for instance in leave_request_filter.object_list]
    )

    return render(
        request,
        "leave/leave_request/leave_clashes.html",
        {
            "leave_request": record,
            "records": overlapping_requests,
            "current_date": date.today(),
            "requests_ids": requests_ids,
            "clashed_due_to_department": clashed_due_to_department,
            "clashed_due_to_job_position": clashed_due_to_job_position,
        },
    )


@login_required
@permission_required("leave.view_leavegeneralsetting")
def compensatory_leave_settings_view(request):
    enabled_compensatory = (
        LeaveGeneralSetting.objects.exists()
        and LeaveGeneralSetting.objects.first().compensatory_leave
    )
    leave_type, create = LeaveType.objects.get_or_create(
        is_compensatory_leave=True,
        defaults={"name": "Compensatory Leave Type", "payment": "paid"},
    )
    context = {"enabled_compensatory": enabled_compensatory, "leave_type": leave_type}
    return render(request, "compensatory_settings.html", context)


@login_required
@permission_required("leave.add_leavegeneralsetting")
def enable_compensatory_leave(request):
    """
    This method is used to enable/disable the compensatory leave feature
    """
    compensatory_leave = LeaveGeneralSetting.objects.first()
    compensatory_leave = (
        compensatory_leave if compensatory_leave else LeaveGeneralSetting()
    )
    compensatory_leave.compensatory_leave = "compensatory_leave" in request.GET.keys()
    compensatory_leave.save()
    if "compensatory_leave" in request.GET.keys():
        messages.success(request, _("Compensatory leave is enabled successfully!"))
    else:
        messages.success(request, _("Compensatory leave is disabled successfully!"))
    return HttpResponse("")


@login_required
@hx_request_required
def delete_leaverequest_comment(request, comment_id):
    """
    This method is used to delete Leave request comments
    """
    script = ""
    comment = LeaverequestComment.find(comment_id)
    if (
        request.user.employee_get == comment.employee_id
        or request.user.has_perm("leave.delete_leaverequestcomment")
        or is_reportingmanager(request)
    ):
        comment.delete()
        messages.success(request, _("Comment deleted successfully!"))
    else:
        messages.warning(request, _("You don't have permission"))
        script = f"""
            <span hx-get="/leave/leave-request-view-comment/{comment.request_id.id}/?&amp;target=leaveRequest" hx-target="#commentContainer" hx-trigger="load"></span>
        """
    return HttpResponse(script)


@login_required
def delete_leave_comment_file(request):
    """
    Used to delete attachment
    """
    script = ""
    ids = request.GET.getlist("ids")
    leave_id = request.GET["leave_id"]
    comment_id = request.GET["comment_id"]
    comment = LeaverequestComment.find(comment_id)
    if (
        request.user.employee_get == comment.employee_id
        or request.user.has_perm("leave.delete_leaverequestfile")
        or is_reportingmanager(request)
    ):
        LeaverequestFile.objects.filter(id__in=ids).delete()
        messages.success(request, _("File deleted successfully"))
    else:
        messages.warning(request, _("You don't have permission"))
        script = f"""
            <span hx-get="/leave/leave-request-view-comment/{leave_id}/?&amp;target=leaveRequest" hx-target="#commentContainer" hx-trigger="load"></span>
        """
    return HttpResponse(script)


if apps.is_installed("attendance"):
    from leave.models import CompensatoryLeaveRequest, CompensatoryLeaverequestComment

    @login_required
    def get_leave_attendance_dates(request):
        """
        function used to return attendance dates that taken on leave days .

        Parameters:
        request (HttpRequest): The HTTP request object.

        Returns:
        GET : return attendance dates
        """
        if request.GET.get("employee_id"):
            employee = Employee.objects.get(id=request.GET.get("employee_id"))
            holiday_attendance = get_leave_day_attendance(employee)
            # Get a list of tuples containing (id, attendance_date)
            attendance_dates = list(
                holiday_attendance.values_list("id", "attendance_date")
            )
            form = CompensatoryLeaveForm()
            form.fields["attendance_id"].choices = attendance_dates
            attendance_id = render_to_string(
                "leave/compensatory_leave/attendance_id.html",
                {
                    "form": form,
                },
            )
            return HttpResponse(f"{attendance_id}")

    @login_required
    def delete_comment_compensatory_file(request):
        """
        Used to delete attachment
        """
        ids = request.GET.getlist("ids")
        LeaverequestFile.objects.filter(id__in=ids).delete()
        leave_id = request.GET["leave_id"]
        comments = CompensatoryLeaverequestComment.objects.all()
        if not request.user.has_perm("leave.delete_compensatoryleaverequestcomment"):
            comments = comments.filter(employee_id__employee_user_id=request.user)
        if request.GET.get("compensatory"):
            comments = comments.filter(request_id=leave_id).order_by("-created_at")
            template = "leave/compensatory_leave/compensatory_leave_comment.html"
        else:
            comments = comments.filter(request_id=leave_id).order_by("-created_at")
            template = "leave/leave_request/leave_comment.html"
        return render(
            request,
            template,
            {
                "comments": comments,
                "request_id": leave_id,
            },
        )

    @login_required
    @hx_request_required
    def delete_leaverequest_compensatory_comment(request, comment_id):
        """
        This method is used to delete Leave request comments
        """
        if request.GET.get("compensatory"):
            comment = CompensatoryLeaverequestComment.objects.filter(id=comment_id)
            if not request.user.has_perm(
                "leave.delete_compensatoryleaverequestcomment"
            ):
                comment = comment.filter(employee_id__employee_user_id=request.user)
            redirect_url = "view-compensatory-leave-comment"
        else:
            comment = LeaverequestComment.objects.filter(id=comment_id)
            if not request.user.has_perm("leave.delete_leaverequestcomment"):
                comment = comment.filter(employee_id__employee_user_id=request.user)
            redirect_url = "leave-request-view-comment"
        leave_id = comment.first().request_id.id
        comment.delete()
        messages.success(request, _("Comment deleted successfully!"))
        return redirect(redirect_url, leave_id)

    @login_required
    @is_compensatory_leave_enabled()
    def view_compensatory_leave(request):
        """
        function used to view compensatory leave requests.

        Parameters:
        request (HttpRequest): The HTTP request object.

        Returns:
        GET : return compensetory leave request view template
        """
        employee = request.user.employee_get
        queryset = CompensatoryLeaveRequest.objects.all().order_by("-id")
        queryset = CompensatoryLeaveRequestFilter(request.GET, queryset).qs
        queryset = filtersubordinates(
            request, queryset, "leave.view_compensatoryleaverequest"
        )
        page_number = request.GET.get("page")
        comp_leave_requests = paginator_qry(queryset, page_number)
        requests_ids = json.dumps(
            list(comp_leave_requests.object_list.values_list("id", flat=True))
        )
        my_comp_leave_requests = CompensatoryLeaveRequest.objects.filter(
            employee_id=employee.id
        ).order_by("-id")
        my_comp_leave_requests = CompensatoryLeaveRequestFilter(
            request.GET, my_comp_leave_requests
        ).qs
        my_page_number = request.GET.get("m_page")
        my_comp_leave_requests = paginator_qry(my_comp_leave_requests, my_page_number)
        my_requests_ids = json.dumps(
            list(my_comp_leave_requests.object_list.values_list("id", flat=True))
        )
        comp_leave_requests_filter = CompensatoryLeaveRequestFilter()
        previous_data = request.GET.urlencode()
        data_dict = parse_qs(previous_data)
        data_dict = get_key_instances(CompensatoryLeaveRequest, data_dict)
        context = {
            "my_comp_leave_requests": my_comp_leave_requests,
            "comp_leave_requests": comp_leave_requests,
            "pd": previous_data,
            "form": comp_leave_requests_filter.form,
            "filter_dict": data_dict,
            "gp_fields": LeaveAllocationRequestReGroup.fields,
            "requests_ids": requests_ids,
            "my_requests_ids": my_requests_ids,
        }
        return render(
            request, "leave/compensatory_leave/compensatory_leave_view.html", context
        )

    @login_required
    @is_compensatory_leave_enabled()
    @hx_request_required
    def filter_compensatory_leave(request):
        """
        function used to view compensatory leave requests.
        """
        field = request.GET.get("field")
        employee = request.user.employee_get
        page_number = request.GET.get("page")
        my_page_number = request.GET.get("m_page")
        previous_data = request.GET.urlencode()
        template = "leave/compensatory_leave/compensatory_leave_req_list.html"

        # Filter compensatory leave requests
        comp_leave_requests_filtered = CompensatoryLeaveRequestFilter(
            request.GET
        ).qs.order_by("-id")
        my_comp_leave_requests_filtered = CompensatoryLeaveRequest.objects.filter(
            employee_id=employee.id
        ).order_by("-id")
        my_comp_leave_requests_filtered = CompensatoryLeaveRequestFilter(
            request.GET, my_comp_leave_requests_filtered
        ).qs
        comp_leave_requests_filtered = filtersubordinates(
            request, comp_leave_requests_filtered, "leave.view_leaveallocationrequest"
        )

        # Sort compensatory leave requests if requested
        if request.GET.get("sortby"):
            comp_leave_requests_filtered = sortby(
                request, comp_leave_requests_filtered, "sortby"
            )
            my_comp_leave_requests_filtered = sortby(
                request, my_comp_leave_requests_filtered, "sortby"
            )

        # Group compensatory leave requests if field parameter is provided
        if field:
            comp_leave_requests = group_by_queryset(
                comp_leave_requests_filtered, field, page_number, "page"
            )
            my_comp_leave_requests = group_by_queryset(
                my_comp_leave_requests_filtered, field, my_page_number, "m_page"
            )

            # Convert IDs to JSON format for details view
            list_values = [entry["list"] for entry in comp_leave_requests]
            id_list = [
                instance.id for value in list_values for instance in value.object_list
            ]
            requests_ids = json.dumps(list(id_list))

            list_values = [entry["list"] for entry in my_comp_leave_requests]
            id_list = [
                instance.id for value in list_values for instance in value.object_list
            ]
            my_requests_ids = json.dumps(list(id_list))
            template = (
                "leave/leave_allocation_request/leave_allocation_request_group_by.html"
            )
        else:
            comp_leave_requests = paginator_qry(
                comp_leave_requests_filtered, page_number
            )
            my_comp_leave_requests = paginator_qry(
                my_comp_leave_requests_filtered, my_page_number
            )
            requests_ids = json.dumps(
                list(comp_leave_requests.object_list.values_list("id", flat=True))
            )
            my_requests_ids = json.dumps(
                list(my_comp_leave_requests.object_list.values_list("id", flat=True))
            )

        # Parse previous data and construct context for filter tag
        data_dict = parse_qs(previous_data)
        data_dict = get_key_instances(CompensatoryLeaveRequest, data_dict)
        data_dict.pop("m_page", None)

        context = {
            "comp_leave_requests": comp_leave_requests,
            "my_comp_leave_requests": my_comp_leave_requests,
            "pd": previous_data,
            "filter_dict": data_dict,
            "field": field,
            "requests_ids": requests_ids,
            "my_requests_ids": my_requests_ids,
        }
        return render(request, template, context=context)

    @login_required
    @is_compensatory_leave_enabled()
    @hx_request_required
    def create_compensatory_leave(request, comp_id=None):
        """
        function used to create or update compensatory leave request.

        Parameters:
        request (HttpRequest): The HTTP request object.

        Returns:
        GET : return leave allocation request form template
        POST : return leave allocation request view
        """
        employee = request.user.employee_get
        template = "leave/compensatory_leave/comp_leave_form.html"
        instance = None
        if comp_id != None:
            instance = CompensatoryLeaveRequest.objects.get(id=comp_id)
        form = CompensatoryLeaveForm(instance=instance)
        if request.method == "POST":
            form = CompensatoryLeaveForm(request.POST, instance=instance)
            if form.is_valid():
                comp_req = form.save()
                comp_req.requested_days = attendance_days(
                    comp_req.employee_id, comp_req.attendance_id.all()
                )
                comp_req.save()
                if comp_id != None:
                    messages.success(request, _("Compensatory Leave updated."))
                else:
                    messages.success(request, _("Compensatory Leave created."))
                return HttpResponse("<script>window.location.reload();</script>")

        context = {
            "employee": employee,
            "form": form,
        }
        return render(request, template, context)

    @login_required
    @is_compensatory_leave_enabled()
    @hx_request_required
    @owner_can_enter(
        perm="leave.delete_compensatoryleaverequest",
        model=CompensatoryLeaveRequest,
        manager_access=True,
    )
    def delete_compensatory_leave(request, comp_id):
        """
        function used to delete compensatory leave request,
        and reload the list view of compensatory leave requests.
        """
        try:
            comp_leave_req = CompensatoryLeaveRequest.objects.get(id=comp_id).delete()
            messages.success(request, _("Compensatory leave request deleted."))

        except:
            messages.error(request, _("Sorry, something went wrong!"))
        if request.GET.get("list") == "True":
            return redirect(filter_compensatory_leave)
        else:
            return HttpResponse("<script>location.reload();</script>")

    @login_required
    @is_compensatory_leave_enabled()
    @hx_request_required
    @manager_can_enter(perm="leave.change_compensatoryleaverequest")
    def approve_compensatory_leave(request, comp_id):
        """
        function used to approve compensatory leave request,
        and reload the list view of compensatory leave requests.
        """
        try:
            comp_leave_req = CompensatoryLeaveRequest.objects.get(id=comp_id)
            if comp_leave_req.status == "requested":
                comp_leave_req.status = "approved"
                comp_leave_req.assign_compensatory_leave_type()
                comp_leave_req.save()
                messages.success(request, _("Compensatory leave request approved."))
                with contextlib.suppress(Exception):
                    notify.send(
                        request.user.employee_get,
                        recipient=comp_leave_req.employee_id.employee_user_id,
                        verb="Your compensatory leave request has been approved",
                        verb_ar="تمت الموافقة على طلب إجازة الاعتذار الخاص بك",
                        verb_de="Ihr Antrag auf Freizeitausgleich wurde genehmigt",
                        verb_es="Su solicitud de permiso compensatorio ha sido aprobada",
                        verb_fr="Votre demande de congé compensatoire a été approuvée",
                        redirect=reverse("view-compensatory-leave")
                        + f"?id={comp_leave_req.id}",
                    )
            else:
                messages.info(
                    request,
                    _(
                        "The compensatory leave request is not in the 'requested' status."
                    ),
                )
        except:
            messages.error(request, _("Sorry, something went wrong!"))
        if request.GET.get("individual"):
            return HttpResponse("<script>location.reload();</script>")
        return redirect(filter_compensatory_leave)

    @login_required
    @is_compensatory_leave_enabled()
    @hx_request_required
    @manager_can_enter(perm="leave.delete_compensatoryleaverequest")
    def reject_compensatory_leave(request, comp_id):
        """
        function used to Reject compensatoey leave request.

        Parameters:
        request (HttpRequest): The HTTP request object.
        comp_id : compensatory leave request id

        Returns:
        GET : It returns to the default compensatory leave request view template.

        """
        comp_leave_req = CompensatoryLeaveRequest.objects.get(id=comp_id)
        if comp_leave_req.status == "requested" or comp_leave_req.status == "approved":
            form = CompensatoryLeaveRequestRejectForm()
            if request.method == "POST":
                form = CompensatoryLeaveRequestRejectForm(request.POST)
                if form.is_valid():
                    comp_leave_req.reject_reason = form.cleaned_data["reason"]
                    comp_leave_req.status = "rejected"
                    comp_leave_req.exclude_compensatory_leave()
                    comp_leave_req.save()
                    messages.success(request, _("Compensatory Leave request rejected."))
                    with contextlib.suppress(Exception):
                        notify.send(
                            request.user.employee_get,
                            recipient=comp_leave_req.employee_id.employee_user_id,
                            verb="Your compensatory leave request has been rejected",
                            verb_ar="تم رفض طلبك للإجازة التعويضية",
                            verb_de="Ihr Antrag auf Freizeitausgleich wurde abgelehnt",
                            verb_es="Se ha rechazado su solicitud de permiso compensatorio",
                            verb_fr="Votre demande de congé compensatoire a été rejetée",
                            redirect=reverse("view-compensatory-leave")
                            + f"?id={comp_leave_req.id}",
                        )
                    return HttpResponse("<script>location.reload();</script>")
            return render(
                request,
                "leave/compensatory_leave/compensatory_leave_reject_form..html",
                {"form": form, "comp_id": comp_id},
            )
        else:
            messages.error(request, _("The leave allocation request can't be rejected"))
            return HttpResponse("<script>location.reload();</script>")

    @login_required
    @is_compensatory_leave_enabled()
    @hx_request_required
    def compensatory_leave_individual_view(request, comp_leave_id):
        """
        function used to present the compensatory leave request detailed view.

        Parameters:
        request (HttpRequest): The HTTP request object.
        comp_leave_id : compensatory leave request id

        Returns:
        return compensatory leave request single view
        """
        requests_ids_json = request.GET.get("instances_ids")
        if requests_ids_json:
            requests_ids = json.loads(requests_ids_json)
            previous_id, next_id = closest_numbers(requests_ids, comp_leave_id)
        comp_leave_req = CompensatoryLeaveRequest.objects.get(id=comp_leave_id)
        context = {
            "comp_leave_req": comp_leave_req,
            "my_request": eval_validate(request.GET.get("my_request")),
            "instances_ids": requests_ids_json,
            "previous": previous_id,
            "next": next_id,
        }
        return render(
            request,
            "leave/compensatory_leave/individual_view_compensatory.html",
            context=context,
        )

    @login_required
    @is_compensatory_leave_enabled()
    @hx_request_required
    def view_compensatory_leave_comment(request, comp_leave_id):
        """
        This method is used to show Leave request comments
        """
        comments = CompensatoryLeaverequestComment.objects.filter(
            request_id=comp_leave_id
        ).order_by("-created_at")
        no_comments = False
        if not comments.exists():
            no_comments = True

        if request.FILES:
            files = request.FILES.getlist("files")
            comment_id = request.GET["comment_id"]
            comment = CompensatoryLeaverequestComment.objects.get(id=comment_id)
            attachments = []
            for file in files:
                file_instance = LeaverequestFile()
                file_instance.file = file
                file_instance.save()
                attachments.append(file_instance)
            comment.files.add(*attachments)

        return render(
            request,
            "leave/compensatory_leave/compensatory_leave_comment.html",
            {
                "comments": comments,
                "no_comments": no_comments,
                "request_id": comp_leave_id,
            },
        )

    @login_required
    @is_compensatory_leave_enabled()
    @hx_request_required
    def create_compensatory_leave_comment(request, comp_leave_id):
        """
        This method renders form and template to create Compensatory leave comments
        """
        comp_leave = CompensatoryLeaveRequest.objects.filter(id=comp_leave_id).first()
        emp = request.user.employee_get
        form = CompensatoryLeaveRequestcommentForm(
            initial={"employee_id": emp.id, "request_id": comp_leave}
        )
        target = request.GET.get("target")
        url = "request-filter" if target == "leaveRequest" else "user-request-filter"
        previous_data = request.GET.urlencode()
        if request.method == "POST":
            form = CompensatoryLeaveRequestcommentForm(request.POST)
            if form.is_valid():
                form.instance.employee_id = emp
                form.instance.request_id = comp_leave
                form.save()
                comments = CompensatoryLeaverequestComment.objects.filter(
                    request_id=comp_leave
                ).order_by("-created_at")
                no_comments = False
                if not comments.exists():
                    no_comments = True
                form = CompensatoryLeaveRequestcommentForm(
                    initial={"employee_id": emp.id, "request_id": comp_leave}
                )
                messages.success(request, _("Comment added successfully!"))
                work_info = EmployeeWorkInformation.objects.filter(
                    employee_id=comp_leave.employee_id
                )
                if work_info.exists():
                    if (
                        comp_leave.employee_id.employee_work_info.reporting_manager_id
                        is not None
                    ):
                        if request.user.employee_get.id == comp_leave.employee_id.id:
                            rec = (
                                comp_leave.employee_id.employee_work_info.reporting_manager_id.employee_user_id
                            )
                            notify.send(
                                request.user.employee_get,
                                recipient=rec,
                                verb=f"{comp_leave.employee_id}'s Compensatory leave request has received a comment.",
                                verb_ar=f"تلقى طلب إجازة الاعتذار لـ {comp_leave.employee_id} تعليقًا.",
                                verb_de=f"Der Antrag auf Freizeitausgleich von {comp_leave.employee_id} hat einen Kommentar erhalten.",
                                verb_es=f"La solicitud de permiso compensatorio de {comp_leave.employee_id} ha recibido un comentario.",
                                verb_fr=f"La demande de congé compensatoire de {comp_leave.employee_id} a reçu un commentaire.",
                                redirect=reverse("view-compensatory-leave")
                                + f"?id={comp_leave.id}",
                                icon="chatbox-ellipses",
                            )
                        elif (
                            request.user.employee_get.id
                            == comp_leave.employee_id.employee_work_info.reporting_manager_id.id
                        ):
                            rec = comp_leave.employee_id.employee_user_id
                            notify.send(
                                request.user.employee_get,
                                recipient=rec,
                                verb="Your compensatory leave request has received a comment.",
                                verb_ar="تلقى طلب إجازة العوض الخاص بك تعليقًا.",
                                verb_de="Ihr Antrag auf Freizeitausgleich hat einen Kommentar erhalten.",
                                verb_es="Su solicitud de permiso compensatorio ha recibido un comentario.",
                                verb_fr="Votre demande de congé compensatoire a reçu un commentaire.",
                                redirect=reverse("view-compensatory-leave")
                                + f"?id={comp_leave.id}",
                                icon="chatbox-ellipses",
                            )
                        else:
                            rec = [
                                comp_leave.employee_id.employee_user_id,
                                comp_leave.employee_id.employee_work_info.reporting_manager_id.employee_user_id,
                            ]
                            notify.send(
                                request.user.employee_get,
                                recipient=rec,
                                verb=f"{comp_leave.employee_id}'s compensatory leave request has received a comment.",
                                verb_ar=f"تلقى طلب إجازة التعويض لـ {comp_leave.employee_id} تعليقًا.",
                                verb_de=f"Der Antrag auf Freizeitausgleich von {comp_leave.employee_id} hat einen Kommentar erhalten.",
                                verb_es=f"El pedido de permiso compensatorio de {comp_leave.employee_id} ha recibido un comentario.",
                                verb_fr=f"La demande de congé compensatoire de {comp_leave.employee_id} a reçu un commentaire.",
                                redirect=reverse("view-compensatory-leave")
                                + f"?id={comp_leave.id}",
                                icon="chatbox-ellipses",
                            )
                    else:
                        rec = comp_leave.employee_id.employee_user_id
                        notify.send(
                            request.user.employee_get,
                            recipient=rec,
                            verb="Your compensatory leave request has received a comment.",
                            verb_ar="تلقى طلب إجازة العوض الخاص بك تعليقًا.",
                            verb_de="Ihr Antrag auf Freizeitausgleich hat einen Kommentar erhalten.",
                            verb_es="Su solicitud de permiso compensatorio ha recibido un comentario.",
                            verb_fr="Votre demande de congé compensatoire a reçu un commentaire.",
                            redirect=reverse("view-compensatory-leave")
                            + f"?id={comp_leave.id}",
                            icon="chatbox-ellipses",
                        )
                return render(
                    request,
                    "leave/compensatory_leave/compensatory_leave_comment.html",
                    {
                        "comments": comments,
                        "no_comments": no_comments,
                        "request_id": comp_leave_id,
                    },
                )
        return render(
            request,
            "leave/compensatory_leave/compensatory_leave_comment.html",
            {
                "form": form,
                "request_id": comp_leave_id,
                "pd": previous_data,
                "target": target,
                "url": url,
            },
        )


if apps.is_installed("recruitment"):

    def check_interview_conflicts(request):
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")
        employee_id = request.GET.get("employee_id")

        try:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d")
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d")
            delta = start_date_obj - end_date_obj
            date_list = [
                start_date_obj + timedelta(days=i) for i in range(delta.days + 1)
            ]
            InterviewSchedule = get_horilla_model_class(
                app_label="recruitment", model="interviewschedule"
            )

            interviews = InterviewSchedule.objects.filter(
                employee_id=employee_id, interview_date__in=date_list
            )

            response = {
                "interviews": list(
                    interviews.values_list("candidate_id__name", flat=True)
                ),
            }
            return JsonResponse(response)
        except Exception as e:
            logger.error(e)
            return JsonResponse(e)


@login_required
@permission_required("leave.view_leavegeneralsetting")
def employee_past_leave_restriction(request):
    enabled_restriction = EmployeePastLeaveRestrict.objects.first()
    if not enabled_restriction:
        enabled_restriction = EmployeePastLeaveRestrict.objects.create(enabled=True)
    if request.method == "POST":
        enabled = request.POST.get("enabled")
        if enabled:
            enabled_restriction.enabled = True
        else:
            enabled_restriction.enabled = False
        enabled_restriction.save()

    return render(
        request,
        "leave/settings/past_leave_restrict_view.html",
        {"enabled_restriction": enabled_restriction},
    )


@login_required
def employee_profile_leave_tab(request):
    """
    This method is used to view own profile of employee.
    """
    user = request.user
    employee = request.user.employee_get
    if apps.is_installed("leave"):
        instances = LeaveRequest.objects.filter(employee_id=employee)
        user_leaves = employee.available_leave.all().exclude(
            leave_type_id__is_compensatory_leave=True
        )
        if (
            LeaveGeneralSetting.objects.first()
            and LeaveGeneralSetting.objects.first().compensatory_leave
        ):
            user_leaves = employee.available_leave.all()
    else:
        user_leaves = None
        instances = None

    leave_request_ids = (
        json.dumps([instance.id for instance in instances])
        if instances
        else json.dumps([])
    )
    today = datetime.today()
    now = timezone.now()
    return render(
        request,
        "employee/profile/profile_view.html",
        {
            "employee": employee,
            "user_leaves": user_leaves,
            "leave_request_ids": leave_request_ids,
            "current_date": today,
            "now": now,
        },
    )


@login_required
def employee_view_individual_leave_tab(request, obj_id, **kwargs):
    """
    This method is used to view profile of an employee.
    """
    employee = Employee.objects.get(id=obj_id)
    instances = (
        LeaveRequest.objects.filter(employee_id=employee)
        if apps.is_installed("leave")
        else None
    )
    leave_request_ids = (
        json.dumps([instance.id for instance in instances])
        if instances
        else json.dumps([])
    )
    employee_leaves = employee.available_leave.all()
    filtered_employee_ids = request.session.get("filtered_employees", [])
    filtered_employees = Employee.objects.filter(id__in=filtered_employee_ids)

    request_ids_str = json.dumps(
        [
            instance.id
            for instance in paginator_qry(
                filtered_employees, request.GET.get("page")
            ).object_list
        ]
    )

    # Convert the string to an actual list of integers
    requests_ids = (
        ast.literal_eval(request_ids_str)
        if isinstance(request_ids_str, str)
        else request_ids_str
    )

    employee_id = employee.id
    previous_id = None
    next_id = None

    for index, req_id in enumerate(requests_ids):
        if req_id == employee_id:

            if index == len(requests_ids) - 1:
                next_id = None
            else:
                next_id = requests_ids[index + 1]
            if index == 0:
                previous_id = None
            else:
                previous_id = requests_ids[index - 1]
            break

    # Use same computed balance as My Leave Requests (EL/CL monthly accrual)
    computed_leave_display = {}
    for av in employee_leaves:
        disp = _computed_leave_display_for_available(av)
        if disp is not None:
            computed_leave_display[av.id] = disp

    context = {
        "employee": employee,
        "previous": previous_id,
        "next": next_id,
        "requests_ids": requests_ids,
        "current_date": date.today(),
        "leave_request_ids": leave_request_ids,
        "computed_leave_display": computed_leave_display,
    }
    # if the requesting user opens own data
    if request.user.employee_get == employee:
        context["user_leaves"] = employee_leaves
    else:
        context["employee_leaves"] = employee_leaves

    return render(request, "tabs/leave-tab.html", context=context)


@login_required
def leave_request_and_approve(request):
    previous_data = request.GET.urlencode()
    page_number = request.GET.get("page")
    leave_requests = LeaveRequest.objects.filter(
        status="requested",
        employee_id__is_active=True,
        start_date__gte=date.today(),
    )
    leave_requests = filtersubordinates(
        request, leave_requests, "leave.change_leaverequest"
    )

    # Filter the initial query set for multi_approve_requests
    multi_approve_requests = LeaveRequestConditionApproval.objects.filter(
        is_approved=False, is_rejected=False
    )
    if multi_approve_requests:
        multi_ids = [request.leave_request_id.id for request in multi_approve_requests]

        # Create a new list excluding leave requests with IDs in multi_ids
        leave_requests = [
            leave for leave in leave_requests if leave.id not in multi_ids
        ]

    leave_requests = paginator_qry(leave_requests, page_number)
    leave_requests_ids = json.dumps([instance.id for instance in leave_requests])
    return render(
        request,
        "leave/dashboard/leave_request_approve.html",
        {
            "leave_requests": leave_requests,
            "requests_ids": leave_requests_ids,
            "pd": previous_data,
            # "current_date":date.today(),
        },
    )


def _leave_configuration_filter_options(request):
    """Dropdown options for export filter modal."""
    from base.models import Department

    all_available = filtersubordinates(
        request, AvailableLeave.objects.all(), "leave.view_availableleave"
    )
    return {
        "filter_leave_type_options": LeaveType.objects.all().order_by("name"),
        "filter_department_options": Department.objects.all().order_by("department"),
        "filter_employee_options": (
            Employee.objects.filter(available_leave__in=all_available)
            .distinct()
            .order_by("employee_first_name", "employee_last_name")
        ),
    }


def _build_leave_configuration_data(request, apply_filters=False):
    """
    Build leave configuration dashboard data (shared by view + Excel export).
    When apply_filters=True, reads GET: leave_type, department, employee, search.
    """
    from base.models import Department
    from leave.methods import (
        computed_balance_for_validation,
        get_casual_leave_balance_stats,
        get_leave_configuration_date_display,
        get_probation_period_leave_balance_stats,
        get_recent_approved_probation_leaves,
        get_probation_casual_leave_type,
        get_sick_leave_balance_stats,
        get_yearly_entitlement,
        is_casual_leave_type,
        is_probation_casual_leave_type,
        is_probation_or_interns_monthly,
        is_probation_sick_leave_type,
        is_sick_leave_type,
        leave_type_uses_yearly_balance_reset,
    )

    filter_leave_type = ""
    filter_department = ""
    filter_employee = ""
    filter_search = ""
    if apply_filters:
        filter_leave_type = (request.GET.get("leave_type") or "").strip()
        filter_department = (request.GET.get("department") or "").strip()
        filter_employee = (request.GET.get("employee") or "").strip()
        filter_search = (request.GET.get("search") or "").strip()

    leave_types = LeaveType.objects.all().order_by("name")
    if apply_filters and filter_leave_type.isdigit():
        leave_types = leave_types.filter(id=int(filter_leave_type))

    available_leaves_queryset = filtersubordinates(
        request, AvailableLeave.objects.all(), "leave.view_availableleave"
    ).select_related(
        "employee_id",
        "employee_id__employee_user_id",
        "employee_id__employee_work_info__department_id",
        "leave_type_id",
    )

    if apply_filters and filter_employee.isdigit():
        available_leaves_queryset = available_leaves_queryset.filter(
            employee_id_id=int(filter_employee)
        )
    if apply_filters and filter_department.isdigit():
        available_leaves_queryset = available_leaves_queryset.filter(
            employee_id__employee_work_info__department_id_id=int(filter_department)
        )
    if apply_filters and filter_search:
        available_leaves_queryset = available_leaves_queryset.filter(
            Q(employee_id__employee_first_name__icontains=filter_search)
            | Q(employee_id__employee_last_name__icontains=filter_search)
            | Q(employee_id__badge_id__icontains=filter_search)
            | Q(leave_type_id__name__icontains=filter_search)
        )

    employees = (
        Employee.objects.filter(available_leave__in=available_leaves_queryset)
        .distinct()
        .order_by("employee_first_name", "employee_last_name")
    )

    configuration_data = []
    for leave_type in leave_types:
        assigned_employees = available_leaves_queryset.filter(
            leave_type_id=leave_type
        )

        reset_date_column_label = "Reset Date"
        sick_leave_display = is_sick_leave_type(leave_type)
        casual_leave_display = is_casual_leave_type(leave_type)
        probation_period_leave_display = is_probation_or_interns_monthly(leave_type)
        employee_data = []
        for av_leave in assigned_employees:
            employee = av_leave.employee_id
            probation_leave_taken = None
            sick_leave_taken = None
            probation_casual_leave_taken = None
            casual_leave_taken = None

            computed_balance = computed_balance_for_validation(av_leave)
            leave_taken = (
                LeaveRequest.objects.filter(
                    employee_id=employee,
                    leave_type_id=leave_type,
                    status="approved",
                ).aggregate(total=Sum("requested_days"))["total"]
                or 0
            )

            if sick_leave_display:
                sl_stats = get_sick_leave_balance_stats(av_leave)
                total_leave_days = sl_stats["yearly_total"]
                accrued_days = sl_stats["accrued_days"]
                available_days = sl_stats["available_days"]
                probation_leave_taken = sl_stats["probation_leave_taken"]
                sick_leave_taken = sl_stats["sick_leave_taken"]
                leave_taken = sick_leave_taken
                carryforward_days = 0
            elif casual_leave_display:
                cl_stats = get_casual_leave_balance_stats(av_leave)
                total_leave_days = cl_stats["yearly_total"]
                accrued_days = cl_stats["accrued_days"]
                available_days = cl_stats["available_days"]
                probation_casual_leave_taken = cl_stats["probation_casual_leave_taken"]
                casual_leave_taken = cl_stats["casual_leave_taken"]
                leave_taken = cl_stats["leave_taken"]
                carryforward_days = 0
            elif probation_period_leave_display:
                prob_stats = get_probation_period_leave_balance_stats(av_leave)
                total_leave_days = prob_stats["yearly_total"]
                accrued_days = prob_stats["accrued_days"]
                available_days = prob_stats["bucket_days"]
                leave_taken = prob_stats["leave_taken"]
                carryforward_days = 0
            elif computed_balance is not None:
                available_days = computed_balance
                total_leave_days = get_yearly_entitlement(leave_type)
                accrued_days = round(available_days + float(leave_taken), 2)
                carryforward_days = 0
            else:
                available_days = av_leave.available_days or 0
                carryforward_days = av_leave.carryforward_days or 0
                total_leave_days = get_yearly_entitlement(leave_type) or (
                    av_leave.total_leave_days
                    or (available_days + carryforward_days)
                )
                accrued_days = round(available_days + float(leave_taken), 2)

            recent_leaves = LeaveRequest.objects.filter(
                employee_id=employee,
                leave_type_id=leave_type,
                status="approved",
            ).order_by("-start_date")

            recent_leave_timeline = []
            if sick_leave_display:
                for leave in get_recent_approved_probation_leaves(employee):
                    recent_leave_timeline.append(
                        {"leave": leave, "kind": "probation"}
                    )
                for leave in recent_leaves:
                    recent_leave_timeline.append({"leave": leave, "kind": "sick"})
                recent_leave_timeline.sort(
                    key=lambda entry: entry["leave"].start_date, reverse=True
                )
            elif casual_leave_display:
                pcl_type = get_probation_casual_leave_type()
                if pcl_type:
                    for leave in LeaveRequest.objects.filter(
                        employee_id=employee,
                        leave_type_id=pcl_type,
                        status="approved",
                    ).order_by("-start_date"):
                        recent_leave_timeline.append(
                            {"leave": leave, "kind": "probation_casual"}
                        )
                for leave in recent_leaves:
                    recent_leave_timeline.append({"leave": leave, "kind": "casual"})
                recent_leave_timeline.sort(
                    key=lambda entry: entry["leave"].start_date, reverse=True
                )

            display_reset_date, date_column_label = (
                get_leave_configuration_date_display(leave_type, av_leave)
            )
            reset_date_column_label = date_column_label

            employee_data.append(
                {
                    "employee": employee,
                    "available_leave": av_leave,
                    "available_days": round(available_days, 2),
                    "carryforward_days": round(carryforward_days, 2),
                    "total_leave_days": round(total_leave_days, 2),
                    "accrued_days": round(accrued_days, 2),
                    "leave_taken": round(leave_taken, 2),
                    "probation_leave_taken": probation_leave_taken,
                    "sick_leave_taken": sick_leave_taken,
                    "probation_casual_leave_taken": probation_casual_leave_taken,
                    "casual_leave_taken": casual_leave_taken,
                    "assigned_date": av_leave.assigned_date,
                    "reset_date": display_reset_date,
                    "expired_date": av_leave.expired_date,
                    "recent_leaves": recent_leaves,
                    "recent_leave_timeline": recent_leave_timeline,
                }
            )

        rules = []
        if leave_type.reset_based:
            if leave_type.reset_based == "monthly":
                accrual = leave_type.total_days or 0
                rules.append(f"Accrual: {accrual} day(s) per month")
                if leave_type.reset_day:
                    if leave_type.carryforward_type == "no carryforward":
                        rules.append(
                            f"Accrual Date: {leave_type.reset_day} of each month (no carry forward)"
                        )
                    else:
                        rules.append(
                            f"Accrual Date: {leave_type.reset_day} of each month (carries forward)"
                        )
                if leave_type_uses_yearly_balance_reset(leave_type):
                    rules.append(
                        "Yearly Reset: 1 Jan (unused balance cleared for all employees)"
                    )
            elif leave_type.reset_based == "yearly":
                rules.append(
                    f"Accrual: {leave_type.total_days or 0} day(s) per year"
                )
                if leave_type.reset_month and leave_type.reset_day:
                    month_names = [
                        "",
                        "Jan",
                        "Feb",
                        "Mar",
                        "Apr",
                        "May",
                        "Jun",
                        "Jul",
                        "Aug",
                        "Sep",
                        "Oct",
                        "Nov",
                        "Dec",
                    ]
                    month_name = (
                        month_names[int(leave_type.reset_month)]
                        if leave_type.reset_month.isdigit()
                        else leave_type.reset_month
                    )
                    rules.append(
                        f"Reset Date: {month_name} {leave_type.reset_day} (balance resets)"
                    )

        if sick_leave_display:
            rules.append(
                "Probation Sick Leave taken is deducted from yearly Sick Leave (e.g. 7 − 3 = 4)"
            )
        if casual_leave_display:
            rules.append(
                "Probation Casual Leave taken is deducted from Casual Leave "
                "(e.g. 4 months − 1 PCL = 3 available)"
            )
        if probation_period_leave_display:
            if is_probation_sick_leave_type(leave_type) or is_probation_casual_leave_type(
                leave_type
            ):
                rules.append(
                    "Total: 3 days (3-month probation × 1 day/month); "
                    "bucket holds max 2 at once (current + previous month)"
                )
            else:
                rules.append(
                    "Max 2 days in bucket (current + previous month; no further carry forward)"
                )
            if is_probation_sick_leave_type(leave_type):
                rules.append(
                    "On confirm: taken days are deducted from yearly Sick Leave"
                )
            elif is_probation_casual_leave_type(leave_type):
                rules.append(
                    "On confirm: taken days are deducted from Casual Leave"
                )

        if leave_type.carryforward_type == "no carryforward":
            rules.append("Carry Forward: No")
        elif leave_type.carryforward_type == "carryforward":
            max_cf = leave_type.carryforward_max or "Unlimited"
            rules.append(f"Carry Forward: Yes (Max: {max_cf})")
        elif leave_type.carryforward_type == "carryforward expire":
            max_cf = leave_type.carryforward_max or "Unlimited"
            expire_in = leave_type.carryforward_expire_in or 0
            expire_period = leave_type.carryforward_expire_period or "year"
            rules.append(
                f"Carry Forward: Yes (Max: {max_cf}, Expires: {expire_in} {expire_period}(s))"
            )

        if leave_type.payment == "paid":
            rules.append("Payment: Paid")
        else:
            rules.append("Payment: Unpaid (LOP)")

        assigned_count = assigned_employees.count()
        if assigned_count == 0:
            continue

        configuration_data.append(
            {
                "leave_type": leave_type,
                "rules": rules,
                "assigned_count": assigned_count,
                "employees": employee_data,
                "reset_date_column_label": reset_date_column_label,
                "sick_leave_display": sick_leave_display,
                "casual_leave_display": casual_leave_display,
                "probation_period_leave_display": probation_period_leave_display,
            }
        )

    filter_options = _leave_configuration_filter_options(request)
    query_string = request.GET.urlencode() if apply_filters else ""
    return {
        "configuration_data": configuration_data,
        "total_leave_types": (
            len(configuration_data) if apply_filters else LeaveType.objects.count()
        ),
        "total_employees_with_leaves": employees.count(),
        "total_assignments": available_leaves_queryset.count(),
        "filter_leave_type": filter_leave_type,
        "filter_department": filter_department,
        "filter_employee": filter_employee,
        "filter_search": filter_search,
        **filter_options,
        "pd": query_string,
        "has_filters": apply_filters and bool(
            filter_leave_type or filter_department or filter_employee or filter_search
        ),
    }


def _format_leave_period(leave, full_year=False):
    if not leave.start_date:
        return ""
    fmt = "%d-%b-%Y" if full_year else "%d %b"
    start = leave.start_date.strftime(fmt)
    if leave.end_date and leave.end_date != leave.start_date:
        start = f"{start} - {leave.end_date.strftime(fmt)}"
    days = leave.requested_days or 0
    return f"{start} ({days}d)"


def _format_all_leaves_for_excel(config, emp_data):
    """Format approved leave history for Excel (one entry per line)."""
    lines = []
    if config["sick_leave_display"] and emp_data["recent_leave_timeline"]:
        for entry in emp_data["recent_leave_timeline"]:
            label = (
                str(_("Probation"))
                if entry["kind"] == "probation"
                else str(_("Sick Leave"))
            )
            lines.append(f"{_format_leave_period(entry['leave'], full_year=True)} — {label}")
    elif emp_data["recent_leaves"]:
        for leave in emp_data["recent_leaves"]:
            lines.append(_format_leave_period(leave, full_year=True))
    if not lines:
        return str(_("No leaves taken"))
    return "\n".join(lines)


@login_required
@permission_required("leave.view_availableleave")
def leave_configuration_view(request):
    """
    Comprehensive Leave Configuration Dashboard
    Shows leave types, employee assignments, balances, and leave history.
    """
    return render(
        request,
        "leave/leave_configuration/leave_configuration.html",
        _build_leave_configuration_data(request, apply_filters=False),
    )


@login_required
@permission_required("leave.view_availableleave")
def export_leave_configuration_excel(request):
    """Export Leave Configuration dashboard data to a formatted Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    data = _build_leave_configuration_data(request, apply_filters=True)
    configuration_data = data["configuration_data"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Leave Configuration"

    headers = [
        str(_("Employee")),
        str(_("Department")),
        str(_("Available")),
        str(_("Accrued")),
        str(_("Total")),
        str(_("Taken")),
        str(_("Probation Taken")),
        str(_("Assigned Date")),
        str(_("Reset / Expiry Date")),
        str(_("All Leaves")),
    ]
    col_count = len(headers)
    num_cols = {3, 4, 5, 6, 7}
    date_cols = {8, 9}

    thin = Border(
        left=Side(style="thin", color="D0D5DD"),
        right=Side(style="thin", color="D0D5DD"),
        top=Side(style="thin", color="D0D5DD"),
        bottom=Side(style="thin", color="D0D5DD"),
    )
    title_font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    subtitle_font = Font(name="Calibri", size=11, color="E8F0FE")
    section_font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    body_font = Font(name="Calibri", size=10)
    title_fill = PatternFill("solid", fgColor="1F4E79")
    header_fill = PatternFill("solid", fgColor="2E75B6")
    section_fill = PatternFill("solid", fgColor="D6E4F0")
    alt_fill = PatternFill("solid", fgColor="F5F8FC")
    amount_fill = PatternFill("solid", fgColor="EAF3DE")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    title_cell = ws.cell(row=1, column=1, value=str(_("Leave Configuration Report")))
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, col_count + 1):
        ws.cell(row=1, column=col).fill = title_fill
    ws.row_dimensions[1].height = 28

    subtitle = _(
        "Leave Types: %(types)s  |  Employees: %(emps)s  |  Assignments: %(assigns)s"
    ) % {
        "types": data["total_leave_types"],
        "emps": data["total_employees_with_leaves"],
        "assigns": data["total_assignments"],
    }
    if data.get("has_filters"):
        filter_bits = []
        if data.get("filter_search"):
            filter_bits.append(f"{_('Search')}: {data['filter_search']}")
        selected_lt = next(
            (
                lt.name
                for lt in data.get("filter_leave_type_options", [])
                if str(lt.id) == str(data.get("filter_leave_type"))
            ),
            None,
        )
        if selected_lt:
            filter_bits.append(f"{_('Leave Type')}: {selected_lt}")
        selected_dept = next(
            (
                dept.department
                for dept in data.get("filter_department_options", [])
                if str(dept.id) == str(data.get("filter_department"))
            ),
            None,
        )
        if selected_dept:
            filter_bits.append(f"{_('Department')}: {selected_dept}")
        selected_emp = next(
            (
                emp.get_full_name()
                for emp in data.get("filter_employee_options", [])
                if str(emp.id) == str(data.get("filter_employee"))
            ),
            None,
        )
        if selected_emp:
            filter_bits.append(f"{_('Employee')}: {selected_emp}")
        if filter_bits:
            subtitle = f"{subtitle}  |  {_('Filtered')}: {', '.join(filter_bits)}"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
    subtitle_cell = ws.cell(row=2, column=1, value=str(subtitle))
    subtitle_cell.font = subtitle_font
    subtitle_cell.fill = title_fill
    for col in range(1, col_count + 1):
        ws.cell(row=2, column=col).fill = title_fill
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 8

    header_row = 4
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin
    ws.row_dimensions[header_row].height = 26

    row_idx = header_row + 1
    for config in configuration_data:
        leave_type_name = config["leave_type"].name
        assigned_count = config["assigned_count"]

        ws.merge_cells(
            start_row=row_idx, start_column=1, end_row=row_idx, end_column=col_count
        )
        section_cell = ws.cell(
            row=row_idx,
            column=1,
            value=str(
                _("%(name)s  —  %(count)s Employee(s)")
                % {"name": leave_type_name, "count": assigned_count}
            ),
        )
        section_cell.font = section_font
        section_cell.fill = section_fill
        section_cell.alignment = Alignment(horizontal="left", vertical="center")
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = section_fill
            cell.border = thin
            cell.font = section_font
        ws.row_dimensions[row_idx].height = 22
        row_idx += 1

        for idx, emp_data in enumerate(config["employees"]):
            employee = emp_data["employee"]
            department = ""
            work_info = getattr(employee, "employee_work_info", None)
            if work_info and work_info.department_id:
                department = str(work_info.department_id)

            probation_taken = emp_data["probation_leave_taken"]
            if config["sick_leave_display"] and probation_taken is not None:
                probation_value = float(probation_taken)
            else:
                probation_value = ""

            all_leaves_text = _format_all_leaves_for_excel(config, emp_data)
            leave_line_count = max(1, all_leaves_text.count("\n") + 1)

            values = [
                employee.get_full_name(),
                department,
                float(emp_data["available_days"]),
                float(emp_data["accrued_days"]),
                float(emp_data["total_leave_days"]),
                float(emp_data["leave_taken"] or 0),
                probation_value,
                (
                    emp_data["assigned_date"].strftime("%d-%b-%Y")
                    if emp_data["assigned_date"]
                    else ""
                ),
                (
                    emp_data["reset_date"].strftime("%d-%b-%Y")
                    if emp_data["reset_date"]
                    else ""
                ),
                all_leaves_text,
            ]
            use_alt = idx % 2 == 1
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.font = body_font
                cell.border = thin
                cell.fill = alt_fill if use_alt else PatternFill()
                if col in num_cols:
                    if value != "":
                        cell.number_format = "0.00"
                        if col in (3, 4, 5, 6):
                            cell.fill = amount_fill if not use_alt else PatternFill(
                                "solid", fgColor="E2EED8"
                            )
                    cell.alignment = right
                elif col in date_cols:
                    cell.alignment = center
                elif col == col_count:
                    cell.alignment = left_top
                else:
                    cell.alignment = left
            ws.row_dimensions[row_idx].height = max(20, min(leave_line_count * 16, 120))
            row_idx += 1

        row_idx += 1

    widths = [24, 20, 12, 12, 12, 12, 14, 14, 16, 48]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A5"
    ws.print_title_rows = "1:4"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_view.showGridLines = False

    today = date.today().strftime("%Y%m%d")
    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    )
    response["Content-Disposition"] = (
        f'attachment; filename="leave_configuration_{today}.xlsx"'
    )
    wb.save(response)
    return response


def leave_allocation_approve(request):
    previous_data = request.GET.urlencode()
    page_number = request.GET.get("page")
    allocation_reqests = LeaveAllocationRequest.objects.filter(
        status="requested", employee_id__is_active=True
    )
    allocation_reqests = filtersubordinates(
        request, allocation_reqests, "leave.view_leaveallocationrequest"
    )
    allocation_reqests = paginator_qry(allocation_reqests, page_number)
    allocation_reqests_ids = json.dumps(
        [instance.id for instance in allocation_reqests]
    )
    return render(
        request,
        "leave/dashboard/leave_allocation_approve.html",
        {
            "allocation_reqests": allocation_reqests,
            "reqests_ids": allocation_reqests_ids,
            "pd": previous_data,
            # "current_date":date.today(),
        },
    )
