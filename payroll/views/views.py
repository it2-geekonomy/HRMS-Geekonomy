"""
views.py

This module is used to define the method for the path in the urls
"""

import calendar
import json
import os
import logging
import subprocess
import tempfile
from io import BytesIO
from collections import defaultdict
from datetime import date, datetime, timedelta
from itertools import groupby
from urllib.parse import parse_qs

from django.contrib.staticfiles.finders import find
from django.core.signing import BadSignature, TimestampSigner
from django.templatetags.static import static
from xhtml2pdf import pisa

import pandas as pd
import pdfkit
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import F, ProtectedError, Q, Sum
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.utils.translation import gettext_lazy as _

from base.methods import (
    closest_numbers,
    eval_validate,
    export_data,
    generate_colors,
    generate_pdf,
    get_key_instances,
    sortby,
)
from base.models import Company
from employee.models import Employee, EmployeeWorkInformation
from horilla.decorators import (
    hx_request_required,
    login_required,
    owner_can_enter,
    permission_required,
)
from horilla.group_by import group_by_queryset
from horilla.horilla_settings import HORILLA_DATE_FORMATS
from notifications.signals import notify
from payroll.context_processors import get_active_employees
from payroll.filters import ContractFilter, ContractReGroup, PayslipFilter
from payroll.forms.component_forms import (
    ContractExportFieldForm,
    GenerateSalaryDataForm,
    MonthlySalaryDataEditForm,
    PayrollSettingsForm,
    PayslipAutoGenerateForm,
)
from payroll.methods.methods import paginator_qry, save_payslip
from payroll.methods.salary_data import (
    compute_salary_data_for_employee,
    get_lop_amount,
    get_ph_wo_paid_leave_counts,
)


def _get_paid_days_lop_days_from_salary_data(payslip):
    """
    Get Paid Days, LOP Days and Working Days using the same logic as Salary Data
    (WorkRecords: days_worked = P + L, working_days = month minus PH/WO).
    Returns (paid_days, lop_days, working_days). Falls back to (None, None, None) if unavailable.
    """
    try:
        month = payslip.start_date.month
        year = payslip.start_date.year
        employee = payslip.employee_id
        salary_data = compute_salary_data_for_employee(employee, month, year)
        days_worked = salary_data["days_worked"]
        working_days = salary_data["working_days"]
        paid_days = days_worked
        lop_days = round(float(working_days or 0) - float(days_worked or 0), 1)
        return paid_days, lop_days, working_days
    except Exception:
        return None, None, None


def _sum_payslip_deduction_lines(data):
    """Sum pretax/post-tax/tax/federal deductions (same basis as payslip generation)."""
    total = float(data.get("federal_tax") or 0)
    for key in ("pretax_deductions", "post_tax_deductions", "tax_deductions"):
        for item in data.get(key) or []:
            if isinstance(item, dict):
                total += float(item.get("amount") or 0)
    return total


def _sync_payslip_lop_from_salary_data(data, payslip):
    """
    Refresh LOP amount and day counts from Salary Data / WorkRecords.
    Recomputes total_deductions from deduction lines + LOP once (no double-count).
    """
    paid_days, lop_days, working_days = _get_paid_days_lop_days_from_salary_data(payslip)
    if paid_days is None:
        data["lop_days"] = data.get("unpaid_days", 0)
        return data

    data["paid_days"] = paid_days
    data["unpaid_days"] = lop_days
    data["lop_days"] = lop_days
    if working_days is not None:
        data["total_working_days"] = int(working_days)

    if not working_days or float(working_days) <= 0:
        _apply_pt_threshold_to_payslip_data(data)
        return data

    month = payslip.start_date.month
    year = payslip.start_date.year
    salary_data = compute_salary_data_for_employee(payslip.employee_id, month, year)
    monthly_salary = salary_data.get("monthly_salary") or data.get("monthly_salary")
    lop_amount = get_lop_amount(
        monthly_salary,
        working_days,
        paid_days,
    )
    data["loss_of_pay"] = lop_amount
    data["basic_pay"] = salary_data.get("basic_salary") or data.get("basic_pay")

    data["total_deductions"] = round(
        _sum_payslip_deduction_lines(data) + lop_amount, 2
    )

    gross = float(data.get("gross_pay") or payslip.gross_pay or 0)
    data["net_pay"] = round(gross - data["total_deductions"], 2)

    _apply_pt_threshold_to_payslip_data(data)
    return data


PT_MONTHLY_THRESHOLD = 20000  # PT ₹200 only when monthly/gross > this (20000 no PT, 20001+ PT)


def _apply_pt_threshold_to_payslip_data(data):
    """
    If gross_pay (monthly salary) <= 20000, zero out Professional Tax deduction;
    PT applied only when gross > 20000 (20001+). Modifies data in place.
    """
    gross = float(data.get("gross_pay") or 0)
    if gross > PT_MONTHLY_THRESHOLD:
        return
    deduction_list_keys = [
        "pretax_deductions",
        "post_tax_deductions",
        "tax_deductions",
        "basic_pay_deductions",
        "gross_pay_deductions",
    ]
    pt_titles = ("professional tax", "pt", "professional tax (pt)")
    amount_zeroed = 0
    for key in deduction_list_keys:
        for item in data.get(key) or []:
            if not isinstance(item, dict):
                continue
            title = (item.get("title") or "").strip().lower()
            if title in pt_titles or "professional tax" in title:
                amt = float(item.get("amount") or 0)
                if amt:
                    amount_zeroed += amt
                    item["amount"] = 0
    if amount_zeroed:
        data["total_deductions"] = float(data.get("total_deductions") or 0) - amount_zeroed
        data["net_pay"] = float(data.get("net_pay") or 0) + amount_zeroed


from payroll.models.models import (
    Contract,
    FilingStatus,
    MonthlySalaryData,
    PayrollGeneralSetting,
    Payslip,
    PayslipAutoGenerate,
    Reimbursement,
    ReimbursementFile,
    ReimbursementrequestComment,
    SalaryDataArrearsLog,
    SalaryDataAuditLog,
)
from payroll.models.tax_models import PayrollSettings

# Create your views here.

status_choices = {
    "draft": _("Draft"),
    "review_ongoing": _("Review Ongoing"),
    "confirmed": _("Confirmed"),
    "paid": _("Paid"),
}


def get_payslip_display_totals(payslip):
    """
    Return (display_deduction, display_net_pay) for a payslip using the same LOP and PT
    logic as the individual payslip view, so the list "view all" matches "view payslip".
    """
    data = payslip.pay_head_data.copy() if payslip.pay_head_data else {}
    data.setdefault("total_deductions", float(payslip.deduction or 0))
    data.setdefault("net_pay", float(payslip.net_pay or 0))
    data.setdefault("gross_pay", float(payslip.gross_pay or 0))
    data.setdefault("basic_pay", float(payslip.basic_pay or 0))
    data.setdefault("contract_wage", float(payslip.contract_wage or 0))

    paid_days, lop_days, working_days = _get_paid_days_lop_days_from_salary_data(payslip)
    total_working_days = 30
    total_leaves = 0
    if payslip.pay_head_data and "total_leaves" in payslip.pay_head_data:
        total_leaves = payslip.pay_head_data.get("total_leaves", 0)
    if paid_days is None:
        lop_days = 0
        paid_days = 30
        working_days = None
        if payslip.pay_head_data:
            pay_data = payslip.pay_head_data
            if "unpaid_days" in pay_data:
                lop_days = pay_data.get("unpaid_days", 0)
            elif "loss_of_pay" in pay_data:
                lop_amount = pay_data.get("loss_of_pay", 0)
                actual_basic_pay = pay_data.get(
                    "actual_basic_pay", payslip.basic_pay + lop_amount
                )
                if actual_basic_pay > 0:
                    daily_rate = actual_basic_pay / 30
                    lop_days = round(lop_amount / daily_rate, 1) if daily_rate > 0 else 0
            if "total_working_days" in pay_data:
                total_working_days = pay_data.get("total_working_days", 30)
            paid_days = (
                pay_data.get("paid_days", total_working_days - lop_days)
                if "paid_days" in pay_data
                else total_working_days - lop_days
            )
            total_leaves = (
                pay_data.get("total_leaves", 0) if "total_leaves" in pay_data else 0
            )

    _sync_payslip_lop_from_salary_data(data, payslip)
    return (data["total_deductions"], data["net_pay"])


def get_language_code(request):
    scale_x_text = _("Name of Employees")
    scale_y_text = _("Amount")
    response = {"scale_x_text": scale_x_text, "scale_y_text": scale_y_text}
    return JsonResponse(response)


@login_required
@permission_required("payroll.add_contract")
def contract_create(request):
    """
    Contract create view
    """
    from payroll.forms.forms import ContractForm

    form = ContractForm()
    if request.method == "POST":
        form = ContractForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, _("Contract Created"))
            return redirect(contract_view)
    return render(request, "payroll/common/form.html", {"form": form})


@login_required
@permission_required("payroll.change_contract")
def contract_update(request, contract_id, **kwargs):
    """
    Update an existing contract.

    Args:
        request: The HTTP request object.
        contract_id: The ID of the contract to update.

    Returns:
        If the request method is POST and the form is valid, redirects to the contract view.
        Otherwise, renders the contract update form.

    """
    from payroll.forms.forms import ContractForm

    contract = Contract.objects.filter(id=contract_id).first()
    if not contract:
        messages.info(request, _("The contract could not be found."))
        return redirect(contract_view)
    contract_form = ContractForm(instance=contract)
    if request.method == "POST":
        contract_form = ContractForm(request.POST, request.FILES, instance=contract)
        if contract_form.is_valid():
            contract_form.save()
            messages.success(request, _("Contract updated"))
            return redirect(contract_view)
    return render(
        request,
        "payroll/common/form.html",
        {
            "form": contract_form,
        },
    )


@login_required
@hx_request_required
@permission_required("payroll.change_contract")
def contract_status_update(request, contract_id):
    from payroll.forms.forms import ContractForm

    previous_data = request.GET.urlencode()
    if request.method == "POST":
        contract = Contract.objects.get(id=contract_id)
        if request.POST.get("view"):
            status = request.POST.get("status")
            if status in dict(contract.CONTRACT_STATUS_CHOICES).keys():
                save = True
                if status in ["active", "draft"]:
                    active_contract = Contract.objects.filter(
                        contract_status="active", employee_id=contract.employee_id
                    ).exists()
                    draft_contract = Contract.objects.filter(
                        contract_status="draft", employee_id=contract.employee_id
                    ).exists()
                    if (status == "active" and active_contract) or (
                        status == "draft" and draft_contract
                    ):
                        save = False
                        messages.info(
                            request,
                            _("An {} contract already exists for {}").format(
                                status, contract.employee_id
                            ),
                        )
                if save:
                    contract.contract_status = status
                    contract.save()
                    messages.success(
                        request, _("The contract status has been updated successfully.")
                    )
            else:
                messages.warning(
                    request, _("You selected the wrong option for contract status.")
                )

            return redirect(f"/payroll/contract-filter?{previous_data}")

        contract_form = ContractForm(request.POST, request.FILES, instance=contract)
        if contract_form.is_valid():
            contract_form.save()
            messages.success(request, _("Contract status updated"))
        else:
            for errors in contract_form.errors.values():
                for error in errors:
                    messages.error(request, error)
        return HttpResponse("<script>$('#reloadMessagesButton').click()</script>")


@login_required
@permission_required("payroll.change_contract")
def bulk_contract_status_update(request):
    status = request.POST.get("status")
    ids = eval_validate(request.POST.get("ids"))
    all_contracts = Contract.objects.all()
    contracts = all_contracts.filter(id__in=ids)

    for contract in contracts:
        save = True
        if status in ["active", "draft"]:
            active_contract = all_contracts.filter(
                contract_status="active", employee_id=contract.employee_id
            ).exists()
            draft_contract = all_contracts.filter(
                contract_status="draft", employee_id=contract.employee_id
            ).exists()
            if (status == "active" and active_contract) or (
                status == "draft" and draft_contract
            ):
                save = False
                messages.info(
                    request,
                    _("An {} contract already exists for {}").format(
                        status, contract.employee_id
                    ),
                )
        if save:
            contract.contract_status = status
            contract.save()
            messages.success(
                request, _("The contract status has been updated successfully.")
            )
    return HttpResponse("success")


@login_required
@permission_required("payroll.change_contract")
def update_contract_filing_status(request, contract_id):
    if request.method == "POST":
        contract = get_object_or_404(Contract, id=contract_id)
        filing_status_id = request.POST.get("filing_status")
        try:
            filing_status = (
                FilingStatus.objects.get(id=int(filing_status_id))
                if filing_status_id
                else None
            )
            contract.filing_status = filing_status
            messages.success(
                request, _("The employee filing status has been updated successfully.")
            )
        except (ValueError, OverflowError, FilingStatus.DoesNotExist):
            messages.warning(
                request, _("You selected the wrong option for filing status.")
            )
        contract.save()
        return redirect(contract_filter)


@login_required
@hx_request_required
@permission_required("payroll.delete_contract")
def contract_delete(request, contract_id):
    """
    Delete a contract.

    Args:
        contract_id: The ID of the contract to delete.

    Returns:
        Redirects to the contract view after successfully deleting the contract.

    """
    try:
        Contract.objects.get(id=contract_id).delete()
        messages.success(request, _("Contract deleted"))
        request_path = request.path.split("/")
        if "delete-contract-modal" in request_path:
            if instances_ids := request.GET.get("instances_ids"):
                get_data = request.GET.copy()
                get_data.pop("instances_ids", None)
                previous_data = get_data.urlencode()
                instances_list = json.loads(instances_ids)
                previous_instance, next_instance = closest_numbers(
                    instances_list, contract_id
                )
                if contract_id in instances_list:
                    instances_list.remove(contract_id)
                urls = f"/payroll/single-contract-view/{next_instance}/"
                params = f"?{previous_data}&instances_ids={instances_list}"
                return redirect(urls + params)
            return HttpResponse("<script>window.location.reload();</script>")
        else:
            return redirect(f"/payroll/contract-filter?{request.GET.urlencode()}")
    except Contract.DoesNotExist:
        messages.error(request, _("Contract not found."))
    except ProtectedError:
        messages.error(request, _("You cannot delete this contract."))
    return HttpResponseRedirect(request.META.get("HTTP_REFERER", "/"))


@login_required
@permission_required("payroll.view_contract")
def contract_view(request):
    """
    Contract view method
    """

    contracts = Contract.objects.all()
    if contracts.exists():
        template = "payroll/contract/contract_view.html"
    else:
        template = "payroll/contract/contract_empty.html"

    contracts = paginator_qry(contracts, request.GET.get("page"))
    contract_ids_json = json.dumps([instance.id for instance in contracts.object_list])
    filter_form = ContractFilter(request.GET)
    context = {
        "contracts": contracts,
        "f": filter_form,
        "contract_ids": contract_ids_json,
        "gp_fields": ContractReGroup.fields,
    }

    return render(request, template, context)


@login_required
# @hx_request_required         #this function is also used in payroll dashboard which uses ajax
@owner_can_enter("payroll.view_contract", Contract)
def view_single_contract(request, contract_id):
    """
    Renders a single contract view page.
    """
    get_data = request.GET.copy()
    get_data.pop("instances_ids", None)
    previous_data = get_data.urlencode()
    dashboard = request.GET.get("dashboard", "")

    HTTP_REFERERS = request.META.get("HTTP_REFERER", "").split("/")
    delete_hx_target = (
        "#personal_target"
        if "employee-view" in HTTP_REFERERS or "employee-profile" in HTTP_REFERERS
        else "#objectDetailsModalTarget"
    )

    contract = Contract.find(contract_id)

    # Fallback to employee work_info when contract fields are None (for display)
    try:
        work_info = contract.employee_id.employee_work_info
    except Exception:
        work_info = None
    context = {
        "contract": contract,
        "dashboard": dashboard,
        "delete_hx_target": delete_hx_target,
        "pd": previous_data,
        "department_display": (
            contract.department.department
            if contract.department
            else (work_info.department_id.department if work_info and work_info.department_id else None)
        ),
        "job_position_display": (
            contract.job_position
            if contract.job_position
            else (work_info.job_position_id if work_info and work_info.job_position_id else None)
        ),
        "job_role_display": (
            contract.job_role.job_role
            if contract.job_role
            else (work_info.job_role_id.job_role if work_info and work_info.job_role_id else None)
        ),
        "shift_display": (
            contract.shift
            if contract.shift
            else (work_info.shift_id if work_info and work_info.shift_id else None)
        ),
        "work_type_display": (
            contract.work_type.work_type
            if contract.work_type
            else (work_info.work_type_id.work_type if work_info and work_info.work_type_id else None)
        ),
    }

    contract_ids_json = request.GET.get("instances_ids")
    if contract_ids_json:
        contract_ids = json.loads(contract_ids_json)
        previous_id, next_id = closest_numbers(contract_ids, contract_id)
        context.update(
            {
                "previous": previous_id,
                "next": next_id,
                "contract_ids": contract_ids_json,
            }
        )
    return render(request, "payroll/contract/contract_single_view.html", context)


@login_required
@hx_request_required
@permission_required("payroll.view_contract")
def contract_filter(request):
    """
    Filter contracts based on the provided query parameters.

    Args:
        request: The HTTP request object containing the query parameters.

    Returns:
        Renders the contract list template with the filtered contracts.

    """
    query_string = request.GET.urlencode()
    contracts_filter = ContractFilter(request.GET)
    template = "payroll/contract/contract_list.html"
    contracts = contracts_filter.qs
    field = request.GET.get("field")

    if field != "" and field is not None:
        contracts = group_by_queryset(contracts, field, request.GET.get("page"), "page")
        list_values = [entry["list"] for entry in contracts]
        id_list = []
        for value in list_values:
            for instance in value.object_list:
                id_list.append(instance.id)

        contract_ids_json = json.dumps(list(id_list))
        template = "payroll/contract/group_by.html"

    else:
        contracts = sortby(request, contracts, "orderby")
        contracts = paginator_qry(contracts, request.GET.get("page"))
        contract_ids_json = json.dumps(
            [instance.id for instance in contracts.object_list]
        )

    data_dict = parse_qs(query_string)
    get_key_instances(Contract, data_dict)
    keys_to_remove = [key for key, value in data_dict.items() if value == ["unknown"]]
    for key in keys_to_remove:
        data_dict.pop(key)
    if "contract_status" in data_dict:
        status_list = data_dict["contract_status"]
        if len(status_list) > 1:
            data_dict["contract_status"] = [status_list[-1]]
    return render(
        request,
        template,
        {
            "contracts": contracts,
            "pd": query_string,
            "filter_dict": data_dict,
            "contract_ids": contract_ids_json,
            "field": field,
        },
    )


@login_required
@permission_required("payroll.view_payrollsettings")
def settings(request):
    """
    This method is used to render settings template
    """
    instance = PayrollSettings.objects.first()
    currency_form = PayrollSettingsForm(instance=instance)
    selected_company_id = request.session.get("selected_company")

    if selected_company_id == "all" or not selected_company_id:
        companies = Company.objects.all()
    else:
        companies = Company.objects.filter(id=selected_company_id)

    if request.method == "POST":

        currency_form = PayrollSettingsForm(request.POST, instance=instance)
        if currency_form.is_valid():

            currency_form.save()
            messages.success(request, _("Payroll settings updated."))
            return HttpResponseRedirect(request.META.get("HTTP_REFERER", "/"))
    return render(
        request,
        "payroll/settings/payroll_settings.html",
        {
            "currency_form": currency_form,
            "companies": companies,
            "selected_company_id": selected_company_id,
        },
    )


@login_required
@permission_required("payroll.change_payslip")
def update_payslip_status(request, payslip_id):
    """
    This method is used to update the payslip confirmation status
    """
    status = request.POST.get("status")
    view = request.POST.get("view")
    payslip = Payslip.objects.filter(id=payslip_id).first()
    if payslip:
        payslip.status = status
        payslip.save()
        messages.success(request, _("Payslip status updated"))
    else:
        messages.error(request, _("Payslip not found"))
    if view:
        from .component_views import filter_payslip

        return redirect(filter_payslip)
    data = payslip.pay_head_data
    data["employee"] = payslip.employee_id
    data["payslip"] = payslip
    data["json_data"] = data.copy()
    data["json_data"]["employee"] = payslip.employee_id.id
    data["json_data"]["payslip"] = payslip.id
    data["instance"] = payslip
    return render(request, "payroll/payslip/individual_payslip_summery.html", data)


def update_payslip_status_no_id(request):
    """
    This method is used to update the payslip confirmation status
    """
    message = {"type": "success", "message": "Payslip status updated."}
    if request.method == "POST":
        ids_json = request.POST["ids"]
        ids = json.loads(ids_json)
        status = request.POST["status"]
        slips = Payslip.objects.filter(id__in=ids)
        slips.update(status=status)
        message = {
            "type": "success",
            "message": f"{slips.count()} Payslips status updated.",
        }
    return JsonResponse(message)


@login_required
@permission_required("payroll.change_payslip")
def bulk_update_payslip_status(request):
    """
    This method is used to update payslip status when generating payslip through
    generate payslip method
    """
    json_data = request.GET["json_data"]
    pay_data = json.loads(json_data)
    status = request.GET["status"]

    for json_entry in pay_data:
        data = json.loads(json_entry)
        emp_id = data["employee"]
        employee = Employee.objects.get(id=emp_id)

        payslip_kwargs = {
            "employee_id": employee,
            "start_date": data["start_date"],
            "end_date": data["end_date"],
        }
        filtered_instance = Payslip.objects.filter(**payslip_kwargs).first()
        instance = filtered_instance if filtered_instance is not None else Payslip()

        instance.employee_id = employee
        instance.start_date = data["start_date"]
        instance.end_date = data["end_date"]
        instance.status = status
        instance.basic_pay = data["basic_pay"]
        instance.contract_wage = data["contract_wage"]
        instance.gross_pay = data["gross_pay"]
        instance.deduction = data["total_deductions"]
        instance.net_pay = data["net_pay"]
        instance.pay_head_data = data
        instance.save()

    return JsonResponse({"type": "success", "message": "Payslips status updated"})


def view_payslip_pdf(request, payslip_id):
    """
    Renders the payslip HTML (same page you see in HRMS).
    Accepts ?pdf_token=... for one-time access so email can fetch this exact URL and convert to PDF.
    """
    from .component_views import filter_payslip

    if not Payslip.objects.filter(id=payslip_id).exists():
        return render(request, "405.html")
    payslip = Payslip.objects.get(id=payslip_id)

    # One-time token: PDF table layout for email (for_pdf=True); skips auto-print JS
    pdf_token = request.GET.get("pdf_token")
    if pdf_token:
        try:
            signer = TimestampSigner()
            decoded = signer.unsign(pdf_token, max_age=600)
            if decoded == str(payslip_id):
                data = get_view_payslip_pdf_context(
                    payslip, request=request, for_pdf=False, pdf_kit_render=True
                )
                return render(request, "payroll/payslip/payslip_pdf.html", context=data)
        except (BadSignature, ValueError):
            pass

    if not request.user.is_authenticated:
        return redirect(reverse("login"))
    if not (
        request.user.has_perm("payroll.view_payslip")
        or payslip.employee_id.employee_user_id == request.user
    ):
        return redirect(filter_payslip)
    data = get_view_payslip_pdf_context(
        payslip, request=request, for_pdf=False, pdf_kit_render=False
    )
    return render(request, "payroll/payslip/payslip_pdf.html", context=data)


@login_required
# @permission_required("payroll.view_payslip")
def view_created_payslip(request, payslip_id, **kwargs):
    """
    This method is used to view the saved payslips
    """
    payslip = Payslip.objects.filter(id=payslip_id).first()
    if payslip is not None and (
        request.user.has_perm("payroll.view_payslip")
        or payslip.employee_id.employee_user_id == request.user
    ):
        # the data must be dictionary in the payslip model for the json field
        data = payslip.pay_head_data.copy()
        _sync_payslip_lop_from_salary_data(data, payslip)
        data["employee"] = payslip.employee_id
        data["payslip"] = payslip
        data["json_data"] = data.copy()
        data["json_data"]["employee"] = payslip.employee_id.id
        data["json_data"]["payslip"] = payslip.id
        data["instance"] = payslip
        return render(request, "payroll/payslip/individual_payslip.html", data)
    return render(request, "404.html")


@login_required
@permission_required("payroll.delete_payslip")
def delete_payslip(request, payslip_id):
    """
    This method is used to delete payslip instances
    Args:
        payslip_id (int): Payslip model instance id
    """
    from .component_views import filter_payslip

    try:
        Payslip.objects.get(id=payslip_id).delete()
        messages.success(request, _("Payslip deleted"))
    except Payslip.DoesNotExist:
        messages.error(request, _("Payslip not found."))
    except ProtectedError:
        messages.error(request, _("Something went wrong"))
    if not Payslip.objects.filter():
        return HttpResponse("<script>window.location.reload()</script>")
    return redirect(filter_payslip)


@login_required
@permission_required("payroll.add_contract")
def contract_info_initial(request):
    """
    This is an ajax method to return json response to auto fill the contract
    form fields
    """
    employee_id = request.GET["employee_id"]
    work_info = EmployeeWorkInformation.objects.filter(employee_id=employee_id).first()
    
    if work_info is None:
        response_data = {
            "department": "",
            "job_position": "",
            "job_role": "",
            "shift": "",
            "work_type": "",
            "wage": "",
            "contract_start_date": "",
            "contract_end_date": "",
        }
        return JsonResponse(response_data)
    
    response_data = {
        "department": (
            work_info.department_id.id if work_info.department_id is not None else ""
        ),
        "job_position": (
            work_info.job_position_id.id
            if work_info.job_position_id is not None
            else ""
        ),
        "job_role": (
            work_info.job_role_id.id if work_info.job_role_id is not None else ""
        ),
        "shift": work_info.shift_id.id if work_info.shift_id is not None else "",
        "work_type": (
            work_info.work_type_id.id if work_info.work_type_id is not None else ""
        ),
        "wage": work_info.basic_salary or "",
        "contract_start_date": (
            work_info.date_joining.strftime("%Y-%m-%d") 
            if work_info.date_joining else ""
        ),
        "contract_end_date": (
            work_info.contract_end_date.strftime("%Y-%m-%d") 
            if work_info.contract_end_date else ""
        ),
    }
    return JsonResponse(response_data)


@login_required
@permission_required("payroll.view_contract")
def view_payroll_dashboard(request):
    """
    Dashboard rendering views
    """
    from payroll.forms.forms import DashboardExport

    paid = Payslip.objects.filter(status="paid")
    posted = Payslip.objects.filter(status="confirmed")
    review_ongoing = Payslip.objects.filter(status="review_ongoing")
    draft = Payslip.objects.filter(status="draft")
    export_form = DashboardExport()
    context = {
        "paid": paid,
        "posted": posted,
        "review_ongoing": review_ongoing,
        "draft": draft,
        "export_form": export_form,
    }
    return render(request, "payroll/dashboard.html", context=context)


@login_required
@permission_required("payroll.view_contract")
def view_salary_data(request):
    """
    List monthly salary data with optional month/year filter.
    Archived data is hidden by default and only shown when filtering by year/month.
    """
    import calendar
    from base.methods import get_key_instances

    # Use .filter() not .all() so we get company-filtered list without employee_id__is_active=True
    # (otherwise salary data for inactive employees would not show after generate).
    qs = MonthlySalaryData.objects.filter().select_related("employee_id")
    year = request.GET.get("year")
    month = request.GET.get("month")
    show_archived_only = request.GET.get("archived") in ("1", "true", "yes")
    # When archived=1: show only archived. When no year/month: show only non-archived. When filtering by month/year: show all.
    if show_archived_only:
        qs = qs.filter(archived=True)
    elif not year and not month:
        qs = qs.filter(archived=False)
    if year:
        qs = qs.filter(year=int(year))
    if month:
        qs = qs.filter(month=int(month))
    qs = qs.order_by("-year", "-month", "employee_id__employee_first_name")
    # Total = sum of (final_salary + arrears_amount) for displayed rows
    total_final_salary = qs.aggregate(
        total=Sum(F("final_salary") + F("arrears_amount"))
    )["total"] or 0
    # 25 records per page for Salary Data
    from django.core.paginator import Paginator
    salary_data_paginator = Paginator(qs, 25)
    salary_data_list = salary_data_paginator.get_page(request.GET.get("page"))
    previous_data = request.GET.urlencode()
    data_dict = parse_qs(previous_data)
    get_key_instances(MonthlySalaryData, data_dict)
    generate_form = GenerateSalaryDataForm()
    month_choices = [(i, calendar.month_abbr[i]) for i in range(1, 13)]
    context = {
        "salary_data_list": salary_data_list,
        "salary_data_start_index": salary_data_list.start_index(),
        "generate_form": generate_form,
        "filter_dict": data_dict,
        "month_choices": month_choices,
        "total_final_salary": total_final_salary,
        "show_archived_only": show_archived_only,
    }
    return render(request, "payroll/salary_data/salary_data_list.html", context)


@login_required
@permission_required("payroll.change_contract")
def archive_salary_data_month(request):
    """
    Archive all salary data for a given month/year. Archived data is hidden from the
    default list and only shown when the user filters by that month/year.
    """
    if request.method != "POST":
        return redirect("view-salary-data")
    year = request.POST.get("year")
    month = request.POST.get("month")
    if not year or not month:
        messages.warning(request, _("Please select both year and month to archive."))
        return redirect("view-salary-data")
    try:
        year = int(year)
        month = int(month)
    except (TypeError, ValueError):
        messages.error(request, _("Invalid year or month."))
        return redirect("view-salary-data")
    if not (1 <= month <= 12):
        messages.error(request, _("Month must be between 1 and 12."))
        return redirect("view-salary-data")
    updated = MonthlySalaryData.objects.filter(year=year, month=month).update(
        archived=True
    )
    import calendar
    month_name = calendar.month_abbr[month]
    if updated:
        messages.success(
            request,
            _("Salary data for %(month)s %(year)s archived (%(count)s record(s)). It will only appear when you filter by that month/year.")
            % {"month": month_name, "year": year, "count": updated},
        )
    else:
        messages.info(
            request,
            _("No salary data found for %(month)s %(year)s to archive.")
            % {"month": month_name, "year": year},
        )
    return redirect("view-salary-data")


@login_required
@permission_required("payroll.view_contract")
def export_salary_data_pdf(request):
    """
    Export Salary Data list as PDF. Uses same year/month filters as the list view.
    Archived data excluded when no filter; included when year/month or archived=1.
    Use .entire() so archived/inactive employees' records are included.
    """
    import calendar

    qs = MonthlySalaryData.objects.entire().filter().select_related("employee_id")
    year = request.GET.get("year")
    month = request.GET.get("month")
    show_archived_only = request.GET.get("archived") in ("1", "true", "yes")
    if show_archived_only:
        qs = qs.filter(archived=True)
    elif not year and not month:
        qs = qs.filter(archived=False)
    if year:
        qs = qs.filter(year=int(year))
    if month:
        qs = qs.filter(month=int(month))
    qs = qs.order_by("-year", "-month", "employee_id__employee_first_name")

    total_final_salary = qs.aggregate(
        total=Sum(F("final_salary") + F("arrears_amount"))
    )["total"] or 0

    if year and month:
        title = f"{calendar.month_abbr[int(month)]} {year}"
    elif year:
        title = str(year)
    else:
        title = _("All")

    context = {
        "salary_data_list": list(qs),
        "total_final_salary": total_final_salary,
        "title": title,
    }
    html_content = render_to_string(
        "payroll/salary_data/salary_data_export_pdf.html", context
    )
    result = BytesIO()
    pdf_status = pisa.CreatePDF(src=html_content, dest=result, encoding="utf-8")
    if pdf_status.err:
        return HttpResponse(_("Error generating PDF"), status=500)
    result.seek(0)
    response = HttpResponse(
        result.getvalue(), content_type="application/pdf"
    )
    filename = f"salary_data_{title.replace(' ', '_')}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
@permission_required("payroll.view_contract")
def export_salary_data_excel(request):
    """
    Export Salary Data list as Excel. Uses same year/month filters as the list view.
    Archived data excluded when no filter; include with archived=1. Use .entire() so
    archived/inactive employees (who may not match company filter) are included.
    """
    import calendar

    qs = MonthlySalaryData.objects.entire().filter().select_related("employee_id")
    year = request.GET.get("year")
    month = request.GET.get("month")
    show_archived_only = request.GET.get("archived") in ("1", "true", "yes")
    if show_archived_only:
        qs = qs.filter(archived=True)
    elif not year and not month:
        qs = qs.filter(archived=False)
    if year:
        qs = qs.filter(year=int(year))
    if month:
        qs = qs.filter(month=int(month))
    qs = qs.order_by("-year", "-month", "employee_id__employee_first_name")

    if year and month:
        title = f"{calendar.month_abbr[int(month)]}_{year}"
    elif year:
        title = str(year)
    else:
        title = _("All")
    if show_archived_only:
        title = f"{title}_archived"

    rows = []
    for row in qs:
        emp = row.employee_id
        emp_name = emp.get_full_name()
        if getattr(emp, "employee_badge_id", None):
            emp_name = f"{emp_name} ({emp.employee_badge_id})"
        lop = (float(row.working_days or 0) - float(row.days_worked or 0))
        rows.append({
            _("Employee"): emp_name,
            _("Year"): row.year,
            _("Month"): calendar.month_abbr[row.month] if row.month else "",
            _("Working Days"): row.working_days,
            _("Days Worked"): row.days_worked,
            _("LOP (Days)"): round(lop, 1),
            _("Basic Salary"): round(row.basic_salary or 0, 0),
            _("Monthly Salary"): round(row.monthly_salary or 0, 0),
            _("Calculated (before PT)"): round(row.calculated_salary or 0, 0),
            _("PT"): round(row.pt_deduction or 0, 0),
            _("Final Salary"): round(row.total_pay() or 0, 0),
            _("Arrears"): round(row.arrears_amount or 0, 0),
        })
    df = pd.DataFrame(rows)
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"salary_data_{title}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    with pd.ExcelWriter(response, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Salary Data", index=False)
    return response


def _log_salary_data_audit(request, action, year, month, summary):
    """Record who generated or deleted salary data and when."""
    user = getattr(request, "user", None)
    SalaryDataAuditLog.objects.create(
        user=user if user is not None and user.is_authenticated else None,
        action=action,
        year=year,
        month=month,
        summary=(summary or "")[:4000],
    )


@login_required
@permission_required("payroll.add_contract")
def generate_salary_data(request):
    """
    Generate monthly salary data for selected employee(s).
    Uses Attendance Calendar: days_worked (P+L), working_days (excl. PH, WO),
    monthly_salary = basic*2, calculated_salary = (days_worked/working_days)*monthly_salary.
    """
    if request.method != "POST":
        form = GenerateSalaryDataForm()
        return render(
            request,
            "payroll/salary_data/salary_data_generate_form.html",
            {"form": form},
        )
    # So form's employee_id queryset includes inactive employees; otherwise selected
    # inactive employees are not in queryset and cleaned_data["employee_id"] is empty.
    setattr(request, "is_filtering", True)
    form = GenerateSalaryDataForm(request.POST)
    if not form.is_valid():
        return render(
            request,
            "payroll/salary_data/salary_data_generate_form.html",
            {"form": form},
        )
    year = form.cleaned_data["year"]
    month = form.cleaned_data["month"]
    employees = list(form.cleaned_data["employee_id"])
    created = 0
    updated = 0
    skipped = []
    for employee in employees:
        try:
            data = compute_salary_data_for_employee(employee, month, year)
        except Exception:
            skipped.append(employee)
            continue
        # Avoid update_or_create (it uses select_for_update) which can trigger
        # "FOR UPDATE cannot be applied to the nullable side of an outer join"
        # when HorillaCompanyManager adds company joins.
        obj = MonthlySalaryData.objects.filter(
            employee_id=data["employee"],
            year=data["year"],
            month=data["month"],
        ).first()
        defaults = {
            "days_worked": data["days_worked"],
            "working_days": data["working_days"],
            "basic_salary": data["basic_salary"],
            "monthly_salary": data["monthly_salary"],
            "calculated_salary": data["calculated_salary"],
            "pt_deduction": data.get("pt_deduction", 200),
            "final_salary": data.get("final_salary", 0),
        }
        if obj:
            for key, value in defaults.items():
                setattr(obj, key, value)
            obj.save()
            updated += 1
        else:
            MonthlySalaryData.objects.create(
                employee_id=data["employee"],
                year=data["year"],
                month=data["month"],
                **defaults,
            )
            created += 1
    if created or updated:
        msg = _("%(created)s created, %(updated)s updated.") % {
            "created": created,
            "updated": updated,
        }
        if skipped:
            msg += " " + _("%(count)s employee(s) skipped.") % {"count": len(skipped)}
        messages.success(request, msg)
        month_label = calendar.month_name[int(month)] if month else ""
        detail = _(
            "User: %(user)s. %(created)s record(s) created, %(updated)s updated for %(month)s %(year)s."
        ) % {
            "user": request.user.get_full_name() or request.user.get_username(),
            "created": created,
            "updated": updated,
            "month": month_label,
            "year": year,
        }
        if skipped:
            detail += " " + _("%(count)s employee(s) skipped (error).") % {
                "count": len(skipped)
            }
        _log_salary_data_audit(
            request,
            SalaryDataAuditLog.ACTION_GENERATED,
            year,
            month,
            detail,
        )
    else:
        if skipped:
            messages.warning(
                request,
                _("No records saved. %(count)s employee(s) skipped.") % {"count": len(skipped)},
            )
        else:
            messages.info(request, _("No records to save."))
    return redirect(
        f"{reverse('view-salary-data')}?year={year}&month={month}"
    )


@login_required
@permission_required("payroll.view_contract")
def view_salary_data_detail(request, salary_data_id):
    """Detail view for a single MonthlySalaryData record with full breakdown."""
    try:
        salary_data = MonthlySalaryData.objects.select_related("employee_id").get(
            id=salary_data_id
        )
    except MonthlySalaryData.DoesNotExist:
        messages.error(request, _("Salary data not found."))
        return redirect("view-salary-data")
    counts = get_ph_wo_paid_leave_counts(
        salary_data.month, salary_data.year, salary_data.employee_id_id
    )
    import calendar
    from datetime import date

    _, last = calendar.monthrange(salary_data.year, salary_data.month)
    total_days_in_month = last
    month_name = date(salary_data.year, salary_data.month, 1).strftime("%B")
    context = {
        "salary_data": salary_data,
        "ph_count": counts["ph_count"],
        "wo_count": counts["wo_count"],
        "paid_leave_count": counts["paid_leave_count"],
        "total_days_in_month": total_days_in_month,
        "month_name": month_name,
    }
    return render(
        request,
        "payroll/salary_data/salary_data_detail.html",
        context,
    )


@login_required
@hx_request_required
@permission_required("payroll.change_monthlysalarydata")
def edit_salary_data(request, salary_data_id):
    """
    Edit a MonthlySalaryData row (modal).
    Recomputes monthly_salary/calculated_salary/final_salary from edited inputs.
    """
    try:
        salary_data = MonthlySalaryData.objects.select_related("employee_id").get(
            id=salary_data_id
        )
    except MonthlySalaryData.DoesNotExist:
        messages.error(request, _("Salary data not found."))
        return HttpResponse("<script>window.location.reload()</script>")

    form = MonthlySalaryDataEditForm(instance=salary_data)
    if request.method == "POST":
        form = MonthlySalaryDataEditForm(request.POST, instance=salary_data)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.monthly_salary = (obj.basic_salary or 0) * 2
            if obj.working_days and float(obj.working_days) > 0:
                obj.calculated_salary = (
                    float(obj.days_worked or 0)
                    / float(obj.working_days)
                    * float(obj.monthly_salary or 0)
                )
            else:
                obj.calculated_salary = 0
            obj.final_salary = float(obj.calculated_salary or 0) - float(
                obj.pt_deduction or 0
            )
            obj.save()
            messages.success(request, _("Salary data updated."))
            return HttpResponse("<script>window.location.reload()</script>")

    return render(
        request,
        "payroll/salary_data/salary_data_edit_form.html",
        {"form": form, "salary_data": salary_data},
    )


@login_required
@permission_required("payroll.view_contract")
def delete_salary_data(request, salary_data_id):
    """Delete a single MonthlySalaryData record."""
    try:
        obj = MonthlySalaryData.objects.select_related("employee_id").get(id=salary_data_id)
        emp_label = str(obj.employee_id)
        yr, mo = obj.year, obj.month
        month_label = calendar.month_name[int(mo)] if mo else ""
        detail = _(
            "User: %(user)s. Deleted salary data for %(employee)s — %(month)s %(year)s."
        ) % {
            "user": request.user.get_full_name() or request.user.get_username(),
            "employee": emp_label,
            "month": month_label,
            "year": yr,
        }
        obj.delete()
        _log_salary_data_audit(
            request,
            SalaryDataAuditLog.ACTION_DELETED_SINGLE,
            yr,
            mo,
            detail,
        )
        messages.success(request, _("Salary data deleted."))
    except MonthlySalaryData.DoesNotExist:
        messages.error(request, _("Salary data not found."))
    year = request.GET.get("year", "")
    month = request.GET.get("month", "")
    params = []
    if year:
        params.append(f"year={year}")
    if month:
        params.append(f"month={month}")
    qs = "&".join(params)
    return redirect(reverse("view-salary-data") + ("?" + qs if qs else ""))


@login_required
@permission_required("payroll.view_contract")
def bulk_delete_salary_data(request):
    """Delete multiple MonthlySalaryData records. Expects POST with ids (list)."""
    if request.method != "POST":
        return redirect("view-salary-data")
    ids = request.POST.getlist("ids")
    if not ids:
        messages.warning(request, _("No salary data selected."))
    else:
        rows = list(
            MonthlySalaryData.objects.filter(id__in=ids).select_related("employee_id")
        )
        snapshots = [
            f"{row.employee_id} ({row.year}-{row.month:02d})" for row in rows
        ]
        years = {row.year for row in rows}
        months = {row.month for row in rows}
        # .delete() total includes CASCADE (e.g. SalaryDataArrearsLog); use len(rows) for salary rows.
        _total_deleted, _by_model = MonthlySalaryData.objects.filter(
            id__in=ids
        ).delete()
        n_salary_rows = len(rows)
        if n_salary_rows and snapshots:
            detail = _(
                "User: %(user)s. Bulk deleted %(count)s salary row(s): %(rows)s"
            ) % {
                "user": request.user.get_full_name() or request.user.get_username(),
                "count": n_salary_rows,
                "rows": "; ".join(snapshots)[:3500],
            }
            yr = years.pop() if len(years) == 1 else None
            mo = months.pop() if len(months) == 1 else None
            _log_salary_data_audit(
                request,
                SalaryDataAuditLog.ACTION_DELETED_BULK,
                yr,
                mo,
                detail,
            )
        messages.success(
            request,
            _("%(count)s salary data record(s) deleted.") % {"count": n_salary_rows},
        )
    year = request.GET.get("year", "") or request.POST.get("year", "")
    month = request.GET.get("month", "") or request.POST.get("month", "")
    params = []
    if year:
        params.append(f"year={year}")
    if month:
        params.append(f"month={month}")
    qs = "&".join(params)
    return redirect(reverse("view-salary-data") + ("?" + qs if qs else ""))


@login_required
@permission_required("payroll.view_contract")
def view_salary_data_audit_log(request):
    """Who generated or deleted salary data, with date and time."""
    logs_qs = SalaryDataAuditLog.objects.select_related("user").order_by("-performed_at")
    paginator = Paginator(logs_qs, 40)
    page = paginator.get_page(request.GET.get("page"))
    return render(
        request,
        "payroll/salary_data/salary_data_audit_log.html",
        {"page_obj": page},
    )


@login_required
@permission_required("payroll.view_contract")
def salary_data_arrears_log(request, salary_data_id):
    """Return HTML for arrears log modal (who added/updated, amount, when)."""
    try:
        salary_data = MonthlySalaryData.objects.select_related("employee_id").get(
            id=salary_data_id
        )
    except MonthlySalaryData.DoesNotExist:
        return HttpResponse("", status=404)
    logs = SalaryDataArrearsLog.objects.filter(salary_data_id=salary_data_id).select_related("user").order_by("-created_at")
    return render(
        request,
        "payroll/salary_data/salary_data_arrears_log.html",
        {"salary_data": salary_data, "logs": logs},
    )


@login_required
@permission_required("payroll.view_contract")
def update_salary_data_arrears(request, salary_data_id):
    """Update arrears amount and description for a MonthlySalaryData record."""
    if request.method != "POST":
        return redirect("view-salary-data")
    try:
        obj = MonthlySalaryData.objects.get(id=salary_data_id)
    except MonthlySalaryData.DoesNotExist:
        messages.error(request, _("Salary data not found."))
        return redirect("view-salary-data")
    try:
        arrears_amount = float(request.POST.get("arrears_amount") or 0)
    except (ValueError, TypeError):
        arrears_amount = 0
    arrears_description = (request.POST.get("arrears_description") or "")[:255]
    prev_amount = obj.arrears_amount or 0
    obj.arrears_amount = max(0, arrears_amount)
    obj.arrears_description = arrears_description
    obj.save(update_fields=["arrears_amount", "arrears_description"])
    # Log who added/updated arrears
    action = "added" if prev_amount == 0 and obj.arrears_amount else "updated"
    SalaryDataArrearsLog.objects.create(
        salary_data=obj,
        user=request.user,
        action=action,
        amount=obj.arrears_amount,
        description=obj.arrears_description or "",
    )
    messages.success(request, _("Arrears updated."))
    year = request.GET.get("year", "") or request.POST.get("year", "")
    month = request.GET.get("month", "") or request.POST.get("month", "")
    params = []
    if year:
        params.append(f"year={year}")
    if month:
        params.append(f"month={month}")
    qs = "&".join(params)
    return redirect(reverse("view-salary-data") + ("?" + qs if qs else ""))


@login_required
@permission_required("payroll.add_contract")
def generate_salary_data_form(request):
    """Return only the generate form fragment for HTMX modal."""
    # So Employee multi-select table includes both active and inactive (HorillaCompanyManager.all()
    # otherwise restricts to is_active=True). Filter "Is Active? = No" then has rows to show.
    setattr(request, "is_filtering", True)
    form = GenerateSalaryDataForm()
    return render(
        request,
        "payroll/salary_data/salary_data_generate_form.html",
        {"form": form},
    )


@login_required
@permission_required("payroll.view_contract")
def dashboard_employee_chart(request):
    """
    payroll dashboard employee chart data
    """

    date = request.GET.get("period")
    year = date.split("-")[0]
    month = date.split("-")[1]
    dataset = []

    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    if is_ajax and request.method == "GET":
        employee_list = Payslip.objects.filter(
            Q(start_date__month=month) & Q(start_date__year=year)
        )
        labels = []
        for employee in employee_list:
            labels.append(employee.employee_id)

        colors = [
            "rgba(255, 99, 132, 1)",  # Red
            "rgba(255, 206, 86, 1)",  # Yellow
            "rgba(54, 162, 235, 1)",  # Blue
            "rgba(75, 242, 182, 1)",  # green
        ]

        for choice, color in zip(Payslip.status_choices, colors):
            dataset.append(
                {
                    "label": choice[0],
                    "data": [],
                    "backgroundColor": color,
                }
            )

        employees = [employee.employee_id for employee in employee_list]

        employees = list(set(employees))
        total_pay_with_status = defaultdict(lambda: defaultdict(float))

        for label in employees:
            payslips = employee_list.filter(employee_id=label)
            for payslip in payslips:
                total_pay_with_status[payslip.status][label] += round(
                    payslip.net_pay, 2
                )

        for data in dataset:
            dataset_label = data["label"]
            data["data"] = [
                total_pay_with_status[dataset_label][label] for label in employees
            ]

        employee_label = []
        for employee in employees:
            employee_label.append(
                f"{employee.employee_first_name} {employee.employee_last_name}"
            )

        for value, choice in zip(dataset, Payslip.status_choices):
            if value["label"] == choice[0]:
                value["label"] = choice[1]

        list_of_employees = list(
            Employee.objects.values_list(
                "id", "employee_first_name", "employee_last_name"
            )
        )
        response = {
            "dataset": dataset,
            "labels": employee_label,
            "employees": list_of_employees,
            "message": _("No payslips generated for this month."),
        }
        return JsonResponse(response)


def payslip_details(request):
    """
    payroll dashboard payslip details data
    """

    date = request.GET.get("period")
    year = date.split("-")[0]
    month = date.split("-")[1]
    employee_list = []
    employee_list = Payslip.objects.filter(
        Q(start_date__month=month) & Q(start_date__year=year)
    )
    total_amount = 0
    for employee in employee_list:
        total_amount += employee.net_pay

    response = {
        "no_of_emp": len(employee_list),
        "total_amount": round(total_amount, 2),
    }
    return JsonResponse(response)


@login_required
def dashboard_department_chart(request):
    """
    payroll dashboard department chart data
    """

    date = request.GET.get("period")
    year = date.split("-")[0]
    month = date.split("-")[1]
    dataset = [
        {
            "label": "",
            "data": [],
            "backgroundColor": ["#8de5b3", "#f0a8a6", "#8ed1f7", "#f8e08e", "#c2c7cc"],
        }
    ]
    department = []
    department_total = []

    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    if is_ajax and request.method == "GET":
        employee_list = Payslip.objects.filter(
            Q(start_date__month=month) & Q(start_date__year=year)
        )

        for employee in employee_list:
            department.append(
                employee.employee_id.employee_work_info.department_id.department
            )

        department = list(set(department))
        for depart in department:
            department_total.append({"department": depart, "amount": 0})

        for employee in employee_list:
            employee_department = (
                employee.employee_id.employee_work_info.department_id.department
            )

            for depart in department_total:
                if depart["department"] == employee_department:
                    depart["amount"] += round(employee.net_pay, 2)

        colors = generate_colors(len(department))

        dataset = [
            {
                "label": "",
                "data": [],
                "backgroundColor": colors,
            }
        ]

        for depart_total, depart in zip(department_total, department):
            if depart == depart_total["department"]:
                dataset[0]["data"].append(depart_total["amount"])

        response = {
            "dataset": dataset,
            "labels": department,
            "department_total": department_total,
            "message": _("No payslips generated for this month."),
        }
        return JsonResponse(response)


def contract_ending(request):
    """
    payroll dashboard contract ending details data
    """

    date = request.GET.get("period")
    month = date.split("-")[1]
    year = date.split("-")[0]

    if request.GET.get("initialLoad") == "true":
        if month == "12":
            month = 0
            year = int(year) + 1

        contract_end = Contract.objects.filter(
            contract_end_date__month=int(month) + 1, contract_end_date__year=int(year)
        )
    else:
        contract_end = Contract.objects.filter(
            contract_end_date__month=int(month), contract_end_date__year=int(year)
        )

    ending_contract = []
    for contract in contract_end:
        ending_contract.append(
            {"contract_name": contract.get_display_name(), "contract_id": contract.id}
        )

    response = {
        "contract_end": ending_contract,
        "message": _("No contracts ending this month"),
    }
    return JsonResponse(response)


def payslip_export(request):
    """
    payroll dashboard exporting to excell data

    Args:
    - request (HttpRequest): The HTTP request object.
    - contract_id (int): The ID of the contract to view.

    """

    start_date = request.POST.get("start_date")
    end_date = request.POST.get("end_date")
    employee = request.POST.getlist("employees")
    status = request.POST.get("status")
    contributions = (
        request.POST.getlist("contributions")
        if request.POST.getlist("contributions")
        else get_active_employees(None)["get_active_employees"].values_list(
            "id", flat=True
        )
    )
    department = []
    total_amount = 0

    table1_data = []
    table2_data = []
    table3_data = []
    table4_data = []
    table5_data = []

    employee_payslip_list = Payslip.objects.all()

    if start_date:
        employee_payslip_list = employee_payslip_list.filter(start_date__gte=start_date)

    if end_date:
        employee_payslip_list = employee_payslip_list.filter(end_date__lte=end_date)

    if employee:
        employee_payslip_list = employee_payslip_list.filter(employee_id__in=employee)

    if status:
        employee_payslip_list = employee_payslip_list.filter(status=status)

    for employ in contributions:
        payslips = Payslip.objects.filter(employee_id__id=employ)
        if end_date:
            payslips = Payslip.objects.filter(
                employee_id__id=employ, end_date__lte=end_date
            )
        if start_date:
            payslips = Payslip.objects.filter(
                employee_id__id=employ, start_date__gte=start_date
            )
            if end_date:
                payslips = payslips.filter(end_date__lte=end_date)
        pay_heads = payslips.values_list("pay_head_data", flat=True)
        # contribution_deductions = []
        deductions = []
        for head in pay_heads:
            for deduction in head["gross_pay_deductions"]:
                if deduction.get("deduction_id"):
                    deductions.append(deduction)
            for deduction in head["basic_pay_deductions"]:
                if deduction.get("deduction_id"):
                    deductions.append(deduction)
            for deduction in head["pretax_deductions"]:
                if deduction.get("deduction_id"):
                    deductions.append(deduction)
            for deduction in head["post_tax_deductions"]:
                if deduction.get("deduction_id"):
                    deductions.append(deduction)
            for deduction in head["tax_deductions"]:
                if deduction.get("deduction_id"):
                    deductions.append(deduction)
            for deduction in head["net_deductions"]:
                deductions.append(deduction)

        deductions.sort(key=lambda x: x["deduction_id"])
        grouped_deductions = {
            key: list(group)
            for key, group in groupby(deductions, key=lambda x: x["deduction_id"])
        }

        for deduction_id, group in grouped_deductions.items():
            employee_contribution = sum(item["amount"] for item in group)
            try:
                employer_contribution = sum(
                    item["employer_contribution_amount"] for item in group
                )
            except:
                employer_contribution = 0
            if employer_contribution > 0:
                table5_data.append(
                    {
                        "Employee": Employee.objects.get(id=employ),
                        "Employer Contribution": employer_contribution,
                        "Employee Contribution": employee_contribution,
                    }
                )

    emp = request.user.employee_get
    if employee_payslip_list:
        for payslip in employee_payslip_list:
            # Taking the company_name of the user
            info = EmployeeWorkInformation.objects.filter(employee_id=emp).first()

            if info:
                employee_company = info.company_id
                company_name = Company.objects.filter(company=employee_company).first()
                date_format = (
                    company_name.date_format
                    if company_name and company_name.date_format
                    else "MMM. D, YYYY"
                )
            else:
                date_format = "MMM. D, YYYY"

            start_date_str = str(payslip.start_date)
            end_date_str = str(payslip.end_date)

            # Convert the string to a datetime.date object
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()

            for format_name, format_string in HORILLA_DATE_FORMATS.items():
                if format_name == date_format:
                    formatted_start_date = start_date.strftime(format_string)

            for format_name, format_string in HORILLA_DATE_FORMATS.items():
                if format_name == date_format:
                    formatted_end_date = end_date.strftime(format_string)

            table1_data.append(
                {
                    "employee": f"{payslip.employee_id.employee_first_name} {payslip.employee_id.employee_last_name}",
                    "start_date": formatted_start_date,
                    "end_date": formatted_end_date,
                    "basic_pay": round(payslip.basic_pay, 2),
                    "deduction": round(payslip.deduction, 2),
                    "allowance": round(payslip.gross_pay - payslip.basic_pay, 2),
                    "gross_pay": round(payslip.gross_pay, 2),
                    "net_pay": round(payslip.net_pay, 2),
                    "status": status_choices.get(payslip.status),
                },
            )
    else:
        table1_data.append(
            {
                "employee": "None",
                "start_date": "None",
                "end_date": "None",
                "basic_pay": "None",
                "deduction": "None",
                "allowance": "None",
                "gross_pay": "None",
                "net_pay": "None",
                "status": "None",
            },
        )

    for employee in employee_payslip_list:
        department.append(
            employee.employee_id.employee_work_info.department_id.department
        )

    department = list(set(department))

    for depart in department:
        table2_data.append({"Department": depart, "Amount": 0})

    for employee in employee_payslip_list:
        employee_department = (
            employee.employee_id.employee_work_info.department_id.department
        )

        for depart in table2_data:
            if depart["Department"] == employee_department:
                depart["Amount"] += round(employee.net_pay, 2)

    if not employee_payslip_list:
        table2_data.append({"Department": "None", "Amount": 0})

    contract_end = Contract.objects.all()
    if not start_date and not end_date:
        contract_end = contract_end.filter(
            Q(contract_end_date__month=datetime.now().month)
            & Q(contract_end_date__year=datetime.now().year)
        )
    if end_date:
        contract_end = contract_end.filter(contract_end_date__lte=end_date)

    if start_date:
        if not end_date:
            contract_end = contract_end.filter(
                Q(contract_end_date__gte=start_date)
                & Q(contract_end_date__lte=datetime.now())
            )
        else:
            contract_end = contract_end.filter(contract_end_date__gte=start_date)

    table3_data = {"contract_ending": []}

    for contract in contract_end:
        table3_data["contract_ending"].append(contract.get_display_name())

    if not contract_end:
        table3_data["contract_ending"].append("None")

    for employee in employee_payslip_list:
        total_amount += round(employee.net_pay, 2)

    table4_data = {
        "no_of_payslip_generated": len(employee_payslip_list),
        "total_amount": [total_amount],
    }

    df_table1 = pd.DataFrame(table1_data)
    df_table2 = pd.DataFrame(table2_data)
    df_table3 = pd.DataFrame(table3_data)
    df_table4 = pd.DataFrame(table4_data)
    df_table5 = pd.DataFrame(table5_data)

    df_table1 = df_table1.rename(
        columns={
            "employee": "Employee",
            "start_date": "Start Date",
            "end_date": "End Date",
            "deduction": "Deduction",
            "allowance": "Allowance",
            "gross_pay": "Gross Pay",
            "net_pay": "Net Pay",
            "status": "Status",
        }
    )

    df_table3 = df_table3.rename(
        columns={
            "contract_ending": (
                f"Contract Ending {start_date} to {end_date}"
                if start_date and end_date
                else f"Contract Ending"
            ),
        }
    )

    df_table4 = df_table4.rename(
        columns={
            "no_of_payslip_generated": "Number of payslips generated",
            "total_amount": "Total Amount",
        }
    )

    df_table5 = df_table5.rename(
        columns={
            "contract_ending": (
                f"Employee - Employer Contributions {start_date} to {end_date}"
                if start_date and end_date
                else f"Contract Ending"
            ),
        }
    )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=payslip.xlsx"

    writer = pd.ExcelWriter(response, engine="xlsxwriter")
    df_table1.to_excel(
        writer, sheet_name="Payroll Dashboard details", index=False, startrow=3
    )
    df_table2.to_excel(
        writer,
        sheet_name="Payroll Dashboard details",
        index=False,
        startrow=len(df_table1) + 3 + 3,
    )
    df_table3.to_excel(
        writer,
        sheet_name="Payroll Dashboard details",
        index=False,
        startrow=len(df_table1) + 3 + len(df_table2) + 6,
    )
    df_table5.to_excel(
        writer,
        sheet_name="Payroll Dashboard details",
        index=False,
        startrow=len(df_table1) + 3 + len(df_table2) + len(df_table3) + 9,
    )
    df_table4.to_excel(
        writer,
        sheet_name="Payroll Dashboard details",
        index=False,
        startrow=len(df_table1)
        + 3
        + len(df_table2)
        + len(df_table3)
        + len(df_table5)
        + 12,
    )

    workbook = writer.book
    worksheet = writer.sheets["Payroll Dashboard details"]
    max_columns = max(
        len(df_table1.columns),
        len(df_table2.columns),
        len(df_table3.columns),
        len(df_table4.columns),
        len(df_table5.columns),
    )

    heading_format = workbook.add_format(
        {
            "bold": True,
            "font_size": 14,
            "align": "center",
            "valign": "vcenter",
            "bg_color": "#eb7968",
            "font_size": 20,
        }
    )

    worksheet.set_row(0, 30)
    worksheet.merge_range(
        0,
        0,
        0,
        max_columns - 1,
        (
            f"Payroll details {start_date} to {end_date}"
            if start_date and end_date
            else f"Payroll details"
        ),
        heading_format,
    )

    header_format = workbook.add_format(
        {"bg_color": "#eb7968", "bold": True, "text_wrap": True}
    )

    for col_num, value in enumerate(df_table1.columns.values):
        worksheet.write(3, col_num, value, header_format)
        col_letter = chr(65 + col_num)

        header_width = max(len(value) + 2, len(df_table1[value].astype(str).max()) + 2)
        worksheet.set_column(f"{col_letter}:{col_letter}", header_width)

    for col_num, value in enumerate(df_table2.columns.values):
        worksheet.write(len(df_table1) + 3 + 3, col_num, value, header_format)
        col_letter = chr(65 + col_num)

        header_width = max(len(value) + 2, len(df_table2[value].astype(str).max()) + 2)
        worksheet.set_column(f"{col_letter}:{col_letter}", header_width)

    for col_num, value in enumerate(df_table3.columns.values):
        worksheet.write(
            len(df_table1) + 3 + len(df_table2) + 6, col_num, value, header_format
        )
        col_letter = chr(65 + col_num)

        header_width = max(len(value) + 2, len(df_table3[value].astype(str).max()) + 2)
        worksheet.set_column(f"{col_letter}:{col_letter}", header_width)

    for col_num, value in enumerate(df_table5.columns.values):
        worksheet.write(
            len(df_table1) + 3 + len(df_table2) + len(df_table3) + 9,
            col_num,
            value,
            header_format,
        )
        col_letter = chr(65 + col_num)

    for col_num, value in enumerate(df_table4.columns.values):
        worksheet.write(
            len(df_table1) + 3 + len(df_table2) + len(df_table3) + len(df_table5) + 12,
            col_num,
            value,
            header_format,
        )
        col_letter = chr(65 + col_num)

        header_width = max(len(value) + 2, len(df_table4[value].astype(str).max()) + 2)
        worksheet.set_column(f"{col_letter}:{col_letter}", header_width)

    worksheet.set_row(len(df_table1) + len(df_table2) + 9, 30)

    writer.close()

    return response


@login_required
@permission_required("payroll.delete_payslip")
def payslip_bulk_delete(request):
    """
    This method is used to bulk delete for Payslip
    """
    ids = request.POST["ids"]
    ids = json.loads(ids)
    for id in ids:
        try:
            payslip = Payslip.objects.get(id=id)
            period = f"{payslip.start_date} to {payslip.end_date}"
            payslip.delete()
            messages.success(
                request,
                _("{employee} {period} payslip deleted.").format(
                    employee=payslip.employee_id, period=period
                ),
            )
        except Payslip.DoesNotExist:
            messages.error(request, _("Payslip not found."))
        except ProtectedError:
            messages.error(
                request,
                _("You cannot delete {payslip}").format(payslip=payslip),
            )
    return JsonResponse({"message": "Success"})


@login_required
@permission_required("payroll.change_payslip")
def slip_group_name_update(request):
    """
    This method is used to update the group of the payslip
    """
    new_name = request.POST["newName"]
    group_name = request.POST["previousName"]
    Payslip.objects.filter(group_name=group_name).update(group_name=new_name)
    return JsonResponse(
        {"type": "success", "message": "Batch name updated.", "new_name": new_name}
    )


@login_required
@permission_required("payroll.add_contract")
def contract_export(request):
    hx_request = request.META.get("HTTP_HX_REQUEST")
    if hx_request:
        export_filter = ContractFilter()
        export_column = ContractExportFieldForm()
        content = {
            "export_filter": export_filter,
            "export_column": export_column,
        }
        return render(
            request,
            "payroll/contract/contract_export_filter.html",
            context=content,
        )
    return export_data(
        request=request,
        model=Contract,
        filter_class=ContractFilter,
        form_class=ContractExportFieldForm,
        file_name="Contract_export",
    )


@login_required
@permission_required("payroll.delete_contract")
def contract_bulk_delete(request):
    """
    This method is used to bulk delete Contract
    """
    ids = request.POST["ids"]
    ids = json.loads(ids)
    for id in ids:
        try:
            contract = Contract.objects.get(id=id)
            name = contract.get_display_name()
            contract.delete()
            messages.success(
                request,
                _("{name} deleted.").format(name=name),
            )
        except Payslip.DoesNotExist:
            messages.error(request, _("Contract not found."))
        except ProtectedError:
            messages.error(
                request,
                _("You cannot delete {contract}").format(contract=contract),
            )
    return JsonResponse({"message": "Success"})


def equalize_lists_length(allowances, deductions):
    """
    Equalize the lengths of two lists by appending empty dictionaries to the shorter list.

    Args:
    deductions (list): List of dictionaries representing deductions.
    allowances (list): List of dictionaries representing allowances.

    Returns:
    tuple: Tuple containing two lists with equal lengths.
    """
    num_deductions = len(deductions)
    num_allowances = len(allowances)

    while num_deductions < num_allowances:
        deductions.append({"title": "", "amount": ""})
        num_deductions += 1

    while num_allowances < num_deductions:
        allowances.append({"title": "", "amount": ""})
        num_allowances += 1

    return deductions, allowances



logger = logging.getLogger(__name__)

_CHROMIUM_BROWSERS = (
    r"C:\Program Files\Google\Chrome\Application\chrome.exe",
    r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
    r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
    r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
)


def _chromium_pdf_to_bytes(target):
    """Render URL or file:// HTML to PDF using headless Chrome/Edge (same as browser print)."""
    fd, pdf_path = tempfile.mkstemp(suffix=".pdf")
    os.close(fd)
    try:
        for browser in _CHROMIUM_BROWSERS:
            if not os.path.isfile(browser):
                continue
            try:
                result = subprocess.run(
                    [
                        browser,
                        "--headless=new",
                        "--disable-gpu",
                        "--no-sandbox",
                        "--disable-dev-shm-usage",
                        f"--print-to-pdf={pdf_path}",
                        target,
                    ],
                    capture_output=True,
                    timeout=90,
                )
                if result.returncode != 0:
                    continue
                with open(pdf_path, "rb") as pdf_file:
                    data = pdf_file.read()
                if data.startswith(b"%PDF") and len(data) > 200:
                    return data
            except Exception as exc:
                logger.debug("Headless PDF via %s failed: %s", browser, exc)
        return None
    finally:
        try:
            os.remove(pdf_path)
        except OSError:
            pass


def _payslip_view_pdf_url(payslip_id, request):
    signer = TimestampSigner()
    token = signer.sign(str(payslip_id))
    return request.build_absolute_uri(
        reverse("view-payslip-pdf", kwargs={"payslip_id": payslip_id})
        + "?pdf_token="
        + token
    )

def _payslip_pdfkit_options():
    """wkhtmltopdf options tuned for emailed payslip PDFs."""
    return {
        "page-size": "A4",
        "margin-top": "8mm",
        "margin-bottom": "8mm",
        "margin-left": "10mm",
        "margin-right": "10mm",
        "encoding": "UTF-8",
        "enable-local-file-access": None,
        "print-media-type": None,
        "disable-smart-shrinking": None,
        "dpi": 300,
        "zoom": 1.0,
        "no-stop-slow-scripts": None,
        "javascript-delay": 500,
    }


def _payslip_pdf_via_same_url(payslip_id, request):
    """
    Same payslip page as on-screen (view-payslip-pdf), printed to PDF via headless Chrome/Edge.
    Falls back to wkhtmltopdf/pdfkit if needed.
    """
    try:
        url = _payslip_view_pdf_url(payslip_id, request)
        pdf_bytes = _chromium_pdf_to_bytes(url)
        if not pdf_bytes:
            pdf_bytes = pdfkit.from_url(url, False, options=_payslip_pdfkit_options())
        if pdf_bytes and pdf_bytes.startswith(b"%PDF"):
            response = HttpResponse(pdf_bytes, content_type="application/pdf")
            response["Content-Disposition"] = "inline; filename=payslip.pdf"
            return response
    except Exception as exc:
        logger.warning("Payslip PDF via URL failed: %s", exc)
    return None


def generate_payslip_pdf(template_path, context, html=False, payslip=None, request=None):
    """
    Generate a PDF file from an HTML template and context data.
    Tries pdfkit first; falls back to xhtml2pdf (pisa) if pdfkit fails
    (e.g. wkhtmltopdf not installed), so emailed payslips are always valid PDFs.

    Args:
        template_path (str): The path to the HTML template.
        context (dict): The context data to render the template.
        html (bool): If True, return raw HTML instead of a PDF.

    Returns:
        HttpResponse: A response with the generated PDF file or raw HTML.
    """
    try:
        # Render the HTML content from the template and context
        html_content = render_to_string(template_path, context)

        # Return raw HTML if requested
        if html:
            return HttpResponse(html_content, content_type="text/html")

        pdf_bytes = None

        # Headless Chrome/Edge: same engine as on-screen browser print
        html_fd, html_path = tempfile.mkstemp(suffix=".html")
        try:
            with os.fdopen(html_fd, "w", encoding="utf-8") as html_file:
                html_file.write(html_content)
            html_uri = "file:///" + os.path.normpath(html_path).replace("\\", "/")
            pdf_bytes = _chromium_pdf_to_bytes(html_uri)
        except Exception:
            pdf_bytes = None
        finally:
            try:
                os.remove(html_path)
            except OSError:
                pass

        # Try pdfkit (wkhtmltopdf) if installed
        if not pdf_bytes or not pdf_bytes.startswith(b"%PDF"):
            try:
                pdf_bytes = pdfkit.from_string(html_content, False, options=_payslip_pdfkit_options())
            except Exception:
                pdf_bytes = None

        # Last resort: xhtml2pdf (pisa) - layout may differ from on-screen
        if not pdf_bytes or not pdf_bytes.startswith(b"%PDF"):
            logger.warning("Payslip PDF: using pisa fallback (install Chrome or wkhtmltopdf for on-screen match)")
            result = BytesIO()
            pdf_status = pisa.CreatePDF(
                src=html_content, dest=result, encoding="utf-8"
            )
            if pdf_status.err:
                return HttpResponse(
                    _("Error generating PDF (pdfkit and pisa failed)"),
                    status=500,
                )
            result.seek(0)
            pdf_bytes = result.getvalue()

        if not pdf_bytes or not pdf_bytes.startswith(b"%PDF"):
            return HttpResponse(
                _("Error generating PDF: no valid output"),
                status=500,
            )

        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = "inline; filename=payslip.pdf"
        return response
    except Exception as e:
        return HttpResponse(f"Error generating PDF: {str(e)}", status=500)


def get_view_payslip_pdf_context(
    payslip, request=None, for_pdf=False, pdf_kit_render=False
):
    """
    Build the exact same context used by view_payslip_pdf (view-payslip-pdf/<id>/).
    Use for_pdf=False for on-screen and email (pdfkit) so layout matches.
    Use for_pdf=True only for pisa fallback (table layout).
    """
    company = Company.objects.filter(hq=True).first()
    viewer_employee = None
    if request and getattr(request.user, "employee_get", None):
        viewer_employee = request.user.employee_get
    if viewer_employee is None:
        viewer_employee = payslip.employee_id

    date_format = "MMM. D, YYYY"
    info = EmployeeWorkInformation.objects.filter(
        employee_id=viewer_employee
    ).select_related("company_id").first()
    if info and getattr(info, "company_id", None) and getattr(info.company_id, "date_format", None):
        date_format = info.company_id.date_format

    data = payslip.pay_head_data.copy()
    start_date_str = data["start_date"]
    end_date_str = data["end_date"]
    start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
    end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
    month_start_name = start_date.strftime("%B %d, %Y")
    month_end_name = end_date.strftime("%B %d, %Y")
    formatted_start_date = start_date.strftime("%b. %d, %Y")
    formatted_end_date = end_date.strftime("%b. %d, %Y")
    for format_name, format_string in HORILLA_DATE_FORMATS.items():
        if format_name == date_format:
            formatted_start_date = start_date.strftime(format_string)
            formatted_end_date = end_date.strftime(format_string)
            break

    data["month_start_name"] = month_start_name
    data["month_end_name"] = month_end_name
    data["formatted_start_date"] = formatted_start_date
    data["formatted_end_date"] = formatted_end_date
    data["employee"] = payslip.employee_id
    data["payslip"] = payslip
    data["json_data"] = data.copy()
    data["json_data"]["employee"] = payslip.employee_id.id
    data["json_data"]["payslip"] = payslip.id
    data["instance"] = payslip
    data["currency"] = (
        PayrollSettings.objects.first().currency_symbol
        if PayrollSettings.objects.exists()
        else ""
    )
    data["all_deductions"] = []
    for deduction_list in [
        data["basic_pay_deductions"],
        data["gross_pay_deductions"],
        data["pretax_deductions"],
        data["post_tax_deductions"],
        data["tax_deductions"],
        data["net_deductions"],
    ]:
        data["all_deductions"].extend(deduction_list)
    data["all_allowances"] = data["allowances"].copy()
    equalize_lists_length(data["allowances"], data["all_deductions"])
    data["zipped_data"] = zip(data["allowances"], data["all_deductions"])

    host = request.get_host() if request else ""
    protocol = "https" if request and request.is_secure() else "http"
    data["host"] = host
    data["protocol"] = protocol
    data["company"] = company

    paid_days, lop_days, working_days = _get_paid_days_lop_days_from_salary_data(payslip)
    total_working_days = 30
    total_leaves = 0
    if payslip.pay_head_data and "total_leaves" in payslip.pay_head_data:
        total_leaves = payslip.pay_head_data.get("total_leaves", 0)
    if paid_days is None:
        lop_days = 0
        paid_days = 30
        working_days = None
        if payslip.pay_head_data:
            pay_data = payslip.pay_head_data
            if "unpaid_days" in pay_data:
                lop_days = pay_data.get("unpaid_days", 0)
            elif "loss_of_pay" in pay_data:
                lop_amount = pay_data.get("loss_of_pay", 0)
                actual_basic_pay = pay_data.get(
                    "actual_basic_pay", payslip.basic_pay + lop_amount
                )
                if actual_basic_pay > 0:
                    daily_rate = actual_basic_pay / 30
                    lop_days = (
                        round(lop_amount / daily_rate, 1) if daily_rate > 0 else 0
                    )
            if "total_working_days" in pay_data:
                total_working_days = pay_data.get("total_working_days", 30)
            paid_days = (
                pay_data.get("paid_days", total_working_days - lop_days)
                if "paid_days" in pay_data
                else total_working_days - lop_days
            )
            total_leaves = pay_data.get("total_leaves", 0) if "total_leaves" in pay_data else 0

    data["total_leaves"] = total_leaves
    _sync_payslip_lop_from_salary_data(data, payslip)
    if "paid_days" not in data:
        data["paid_days"] = paid_days
    if "total_working_days" not in data:
        data["total_working_days"] = (
            int(working_days) if working_days is not None else total_working_days
        )
    data["lop_days"] = data.get("unpaid_days", data.get("lop_days", 0))
    data["lop_amount"] = data.get("loss_of_pay", 0)

    # Always add net_pay_in_words so view and emailed PDF show the same (no "Loading...")
    net_pay_val = data.get("net_pay", 0)
    try:
        data["net_pay_in_words"] = _number_to_words_indian(net_pay_val)
    except (TypeError, ValueError):
        data["net_pay_in_words"] = ""

    # Logo / watermark for on-screen, pdfkit email, and pisa fallback
    data["watermark_src"] = _payslip_watermark_src(host, protocol)
    data["logo_src"] = _payslip_logo_src(host, protocol)
    if for_pdf:
        data["for_pdf"] = True
        data["currency_symbol"] = "Rs. "
    elif pdf_kit_render:
        data["pdf_kit_render"] = True

    data["payslip_line_rows"] = _build_payslip_line_rows(data)
    data.update(_payslip_header_detail_fields(payslip, data))

    return data


def _payslip_header_detail_fields(payslip, data):
    """
    Fields for Employee Details / Payroll Details on the salary slip PDF.
    PAN / UAN / ESI are read from employee.additional_info when present.
    """
    employee = payslip.employee_id
    account = ""
    try:
        bank = employee.employee_bank_details
        if bank and bank.account_number:
            account = str(bank.account_number).strip()
    except Exception:
        account = ""
    if len(account) >= 4:
        last4 = ("*" * 12) + account[-4:]
    elif account:
        last4 = ("*" * 12) + account
    else:
        last4 = "—"

    extra = employee.additional_info if isinstance(employee.additional_info, dict) else {}

    def _extra(*keys):
        for key in keys:
            val = extra.get(key)
            if val not in (None, ""):
                return val
        return "—"

    if payslip.status == "paid":
        date_of_payment = payslip.end_date.strftime("%d-%m-%Y")
    else:
        date_of_payment = "—"

    return {
        "salary_month": payslip.start_date.strftime("%B %Y"),
        "pay_period_display": f"{data.get('formatted_start_date', payslip.start_date)} – {data.get('formatted_end_date', payslip.end_date)}",
        "date_of_payment": date_of_payment,
        "bank_account_last4": last4,
        "employee_pan": _extra("pan", "PAN", "pan_number", "PAN Number"),
        "employee_uan_pf": _extra(
            "uan", "UAN", "pf", "PF", "pf_number", "uan_number", "UAN / PF"
        ),
        "employee_esi": _extra("esi", "ESI", "esi_number", "ESI Number"),
    }


def _build_payslip_line_rows(data):
    """Pair earnings and deductions into table rows for the salary slip layout."""
    earnings = [{"title": "Basic", "amount": data.get("basic_pay", 0)}]
    for item in data.get("all_allowances") or []:
        title = item.get("title") if isinstance(item, dict) else getattr(item, "title", None)
        amount = item.get("amount") if isinstance(item, dict) else getattr(item, "amount", None)
        if title and amount:
            earnings.append({"title": title, "amount": amount})

    deductions = []
    for item in data.get("all_deductions") or []:
        title = item.get("title") if isinstance(item, dict) else getattr(item, "title", None)
        amount = item.get("amount") if isinstance(item, dict) else getattr(item, "amount", None)
        if title and amount:
            compact = (title or "").lower().replace(" ", "").replace("(", "").replace(")", "")
            if compact in ("incometax", "federaltax") or "incometax" in compact:
                continue
            deductions.append({"title": title, "amount": amount})
    if data.get("loss_of_pay") and not any(
        d["title"].lower().replace(" ", "") in ("lossofpay", "lop")
        for d in deductions
    ):
        deductions.append({"title": "Loss of Pay", "amount": data["loss_of_pay"]})

    def _is_professional_tax(title):
        compact = (title or "").lower().replace(" ", "").replace("(", "").replace(")", "")
        return compact in ("professionaltax", "pt", "ptprofessionaltax") or "professionaltax" in compact

    deductions = [d for d in deductions if _is_professional_tax(d["title"])] + [
        d for d in deductions if not _is_professional_tax(d["title"])
    ]

    row_count = max(len(earnings), len(deductions), 1)
    rows = []
    for i in range(row_count):
        earn = earnings[i] if i < len(earnings) else None
        ded = deductions[i] if i < len(deductions) else None
        rows.append(
            {
                "earning_title": earn["title"] if earn else "",
                "earning_amount": earn["amount"] if earn else None,
                "deduction_title": ded["title"] if ded else "",
                "deduction_amount": ded["amount"] if ded else None,
            }
        )
    return rows


def _number_to_words_indian(n):
    """Convert integer amount to Indian style words, e.g. 23550 -> 'TWENTY THREE THOUSAND FIVE HUNDRED FIFTY RUPEES ONLY'."""
    ones = ["", "ONE", "TWO", "THREE", "FOUR", "FIVE", "SIX", "SEVEN", "EIGHT", "NINE"]
    teens = ["TEN", "ELEVEN", "TWELVE", "THIRTEEN", "FOURTEEN", "FIFTEEN", "SIXTEEN", "SEVENTEEN", "EIGHTEEN", "NINETEEN"]
    tens = ["", "", "TWENTY", "THIRTY", "FORTY", "FIFTY", "SIXTY", "SEVENTY", "EIGHTY", "NINETY"]
    n = int(round(float(n)))

    def hundreds(val):
        if val == 0:
            return ""
        out = ""
        if val >= 100:
            out += ones[val // 100] + " HUNDRED "
            val %= 100
        if val >= 20:
            out += tens[val // 10] + " "
            val %= 10
        elif val >= 10:
            out += teens[val - 10] + " "
            return out
        if val > 0:
            out += ones[val] + " "
        return out

    if n == 0:
        return "ZERO RUPEES ONLY"
    out = ""
    if n >= 100000:
        out += hundreds(n // 100000) + "LAKH "
        n %= 100000
    if n >= 1000:
        out += hundreds(n // 1000) + "THOUSAND "
        n %= 1000
    out += hundreds(n)
    return (out.strip() or "ZERO") + " RUPEES ONLY"


def _payslip_logo_src(host, protocol):
    """Return logo URL/path for PDF: file:// when no host (email), else full URL so pisa/pdfkit can load it."""
    static_path = "images/ui/geekonomy-logo-mail.png"
    if host:
        return f"{protocol}://{host}{static(static_path)}"
    path = find(static_path)
    if path:
        return "file:///" + os.path.normpath(path).replace("\\", "/")
    return ""


def _payslip_watermark_src(host, protocol):
    """Grey Geekonomy logo used as payslip page watermark."""
    static_path = "payroll/images/geekonomy-grey-logo-watermark.png"
    if host:
        return f"{protocol}://{host}{static(static_path)}"
    path = find(static_path)
    if path:
        return "file:///" + os.path.normpath(path).replace("\\", "/")
    return ""


def _build_payslip_pdf_context(payslip, host="", protocol="http"):
    """
    Build context dict for payslip PDF template. Matches view_created_payslip logic
    so the emailed PDF matches what the user sees in HRMS (view-payslip/<id>/).
    Uses the payslip employee's company for address when available.
    """
    employee = payslip.employee_id
    # Use employee's company for address so emailed PDF matches HRMS view; fallback to HQ
    info = EmployeeWorkInformation.objects.filter(employee_id=employee).select_related("company_id").first()
    company = None
    if info and getattr(info, "company_id", None):
        company = info.company_id
    if company is None:
        company = Company.objects.filter(hq=True).first()
    date_format = "MMM. D, YYYY"
    if info and info.company_id:
        if getattr(info.company_id, "date_format", None):
            date_format = info.company_id.date_format

    # Same data build as view_created_payslip so numbers match HRMS
    data = payslip.pay_head_data.copy()
    _sync_payslip_lop_from_salary_data(data, payslip)

    data["employee"] = employee
    data["payslip"] = payslip
    data["instance"] = payslip
    data["json_data"] = data.copy()
    data["json_data"]["employee"] = payslip.employee_id.id
    data["json_data"]["payslip"] = payslip.id

    start_date_str = data["start_date"]
    end_date_str = data["end_date"]
    start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
    end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
    formatted_start_date = start_date.strftime("%b. %d, %Y")
    formatted_end_date = end_date.strftime("%b. %d, %Y")
    for format_name, format_string in HORILLA_DATE_FORMATS.items():
        if format_name == date_format:
            formatted_start_date = start_date.strftime(format_string)
            formatted_end_date = end_date.strftime(format_string)
            break

    data["all_deductions"] = []
    for deduction_list in [
        data["basic_pay_deductions"],
        data["gross_pay_deductions"],
        data["pretax_deductions"],
        data["post_tax_deductions"],
        data["tax_deductions"],
        data["net_deductions"],
    ]:
        data["all_deductions"].extend(deduction_list)
    data["all_allowances"] = data["allowances"].copy()
    equalize_lists_length(data["allowances"], data["all_deductions"])
    data["zipped_data"] = zip(data["allowances"], data["all_deductions"])

    net_pay_val = data.get("net_pay", 0)
    try:
        net_pay_in_words = _number_to_words_indian(net_pay_val)
    except (TypeError, ValueError):
        net_pay_in_words = ""

    data.update(
        {
            "month_start_name": start_date.strftime("%B %d, %Y"),
            "month_end_name": end_date.strftime("%B %d, %Y"),
            "formatted_start_date": formatted_start_date,
            "formatted_end_date": formatted_end_date,
            "currency": PayrollSettings.objects.first().currency_symbol if PayrollSettings.objects.exists() else "",
            "host": host,
            "protocol": protocol,
            "company": company,
            "for_pdf": True,
            "logo_src": _payslip_logo_src(host, protocol),
            "net_pay_in_words": net_pay_in_words,
            "currency_display": "Rs. ",  # Avoid ₹ in PDF when font lacks the glyph (pisa shows black square)
        }
    )
    # Ensure template has paid_days/lop_days/total_working_days when not set by salary data (fallback for display)
    if "paid_days" not in data:
        pay_data = payslip.pay_head_data or {}
        data["total_working_days"] = int(pay_data.get("total_working_days", 30))
        data["paid_days"] = pay_data.get("paid_days", data["total_working_days"])
        data["unpaid_days"] = pay_data.get("unpaid_days", 0)
        data["lop_days"] = data["unpaid_days"]
    else:
        data["lop_days"] = data["unpaid_days"]
        data["total_working_days"] = data.get("total_working_days", 30)
    return data


def payslip_pdf_content(payslip_id):
    """
    Generate payslip PDF content without request. For automatic email (scheduler).
    Uses same context as view-payslip-pdf so the mailed PDF is identical to the HRMS view.
    Returns HttpResponse with PDF, or None if payslip not found.
    """
    payslip = Payslip.objects.filter(id=payslip_id).first()
    if not payslip:
        return None
    context = get_view_payslip_pdf_context(
        payslip, request=None, for_pdf=False, pdf_kit_render=True
    )
    return generate_payslip_pdf(
        "payroll/payslip/payslip_pdf.html",
        context=context,
        html=False,
        payslip=payslip,
        request=None,
    )


def payslip_pdf(request, id):
    """
    Return the payslip as PDF. Prefer fetching the SAME view-payslip-pdf URL and
    converting to PDF so the mailed PDF is identical to what you see in HRMS.
    """
    from .component_views import filter_payslip

    if not Payslip.objects.filter(id=id).exists():
        return render(request, "405.html")
    payslip = Payslip.objects.get(id=id)
    if not (
        request.user.has_perm("payroll.view_payslip")
        or payslip.employee_id.employee_user_id == request.user
    ):
        return redirect(filter_payslip)

    # Same PDF as HRMS: fetch the exact view-payslip-pdf page and convert to PDF (no separate generation)
    same_pdf = _payslip_pdf_via_same_url(id, request)
    if same_pdf is not None and getattr(same_pdf, "content", b"").startswith(b"%PDF"):
        return same_pdf

    # Fallback if URL fetch fails (e.g. wkhtmltopdf not installed or server not reachable)
    data = get_view_payslip_pdf_context(
        payslip, request=request, for_pdf=False, pdf_kit_render=True
    )
    return generate_payslip_pdf(
        "payroll/payslip/payslip_pdf.html",
        context=data,
        html=False,
        payslip=payslip,
        request=request,
    )


@login_required
@permission_required("payroll.view_contract")
def contract_select(request):
    page_number = request.GET.get("page")

    if page_number == "all":
        employees = Contract.objects.all()

    contract_ids = [str(emp.id) for emp in employees]
    total_count = employees.count()

    context = {"contract_ids": contract_ids, "total_count": total_count}

    return JsonResponse(context, safe=False)


@login_required
def contract_select_filter(request):
    page_number = request.GET.get("page")
    filtered = request.GET.get("filter")
    filters = json.loads(filtered) if filtered else {}

    if page_number == "all":
        contract_filter = ContractFilter(filters, queryset=Contract.objects.all())

        # Get the filtered queryset
        filtered_employees = contract_filter.qs

        contract_ids = [str(emp.id) for emp in filtered_employees]
        total_count = filtered_employees.count()

        context = {"contract_ids": contract_ids, "total_count": total_count}

        return JsonResponse(context)


@login_required
def payslip_select(request):
    page_number = request.GET.get("page")

    if page_number == "all":
        if request.user.has_perm("payroll.view_payslip"):
            employees = Payslip.objects.all()
        else:
            employees = Payslip.objects.filter(
                employee_id__employee_user_id=request.user
            )

    payslip_ids = [str(emp.id) for emp in employees]
    total_count = employees.count()

    context = {"payslip_ids": payslip_ids, "total_count": total_count}

    return JsonResponse(context, safe=False)


@login_required
def payslip_select_filter(request):
    page_number = request.GET.get("page")
    filtered = request.GET.get("filter")
    filters = json.loads(filtered) if filtered else {}

    if page_number == "all":
        payslip_filter = PayslipFilter(filters, queryset=Payslip.objects.all())

        # Get the filtered queryset
        filtered_employees = payslip_filter.qs

        payslip_ids = [str(emp.id) for emp in filtered_employees]
        total_count = filtered_employees.count()

        context = {"payslip_ids": payslip_ids, "total_count": total_count}

        return JsonResponse(context)


@login_required
def create_payrollrequest_comment(request, payroll_id):
    """
    This method renders form and template to create Reimbursement request comments
    """
    from payroll.forms.forms import ReimbursementRequestCommentForm

    payroll = Reimbursement.objects.filter(id=payroll_id).first()
    emp = request.user.employee_get
    form = ReimbursementRequestCommentForm(
        initial={"employee_id": emp.id, "request_id": payroll_id}
    )

    if request.method == "POST":
        form = ReimbursementRequestCommentForm(request.POST)
        if form.is_valid():
            form.instance.employee_id = emp
            form.instance.request_id = payroll
            form.save()
            comments = ReimbursementrequestComment.objects.filter(
                request_id=payroll_id
            ).order_by("-created_at")
            no_comments = False
            if not comments.exists():
                no_comments = True
            form = ReimbursementRequestCommentForm(
                initial={"employee_id": emp.id, "request_id": payroll_id}
            )
            messages.success(request, _("Comment added successfully!"))

            if payroll.employee_id.employee_work_info.reporting_manager_id is not None:

                if request.user.employee_get.id == payroll.employee_id.id:
                    rec = (
                        payroll.employee_id.employee_work_info.reporting_manager_id.employee_user_id
                    )
                    notify.send(
                        request.user.employee_get,
                        recipient=rec,
                        verb=f"{payroll.employee_id}'s reimbursement request has received a comment.",
                        verb_ar=f"تلقى طلب استرداد نفقات {payroll.employee_id} تعليقًا.",
                        verb_de=f"{payroll.employee_id}s Rückerstattungsantrag hat einen Kommentar erhalten.",
                        verb_es=f"La solicitud de reembolso de gastos de {payroll.employee_id} ha recibido un comentario.",
                        verb_fr=f"La demande de remboursement de frais de {payroll.employee_id} a reçu un commentaire.",
                        redirect=reverse("view-reimbursement"),
                        icon="chatbox-ellipses",
                    )
                elif (
                    request.user.employee_get.id
                    == payroll.employee_id.employee_work_info.reporting_manager_id.id
                ):
                    rec = payroll.employee_id.employee_user_id
                    notify.send(
                        request.user.employee_get,
                        recipient=rec,
                        verb="Your reimbursement request has received a comment.",
                        verb_ar="تلقى طلب استرداد نفقاتك تعليقًا.",
                        verb_de="Ihr Rückerstattungsantrag hat einen Kommentar erhalten.",
                        verb_es="Tu solicitud de reembolso ha recibido un comentario.",
                        verb_fr="Votre demande de remboursement a reçu un commentaire.",
                        redirect=reverse("view-reimbursement"),
                        icon="chatbox-ellipses",
                    )
                else:
                    rec = [
                        payroll.employee_id.employee_user_id,
                        payroll.employee_id.employee_work_info.reporting_manager_id.employee_user_id,
                    ]
                    notify.send(
                        request.user.employee_get,
                        recipient=rec,
                        verb=f"{payroll.employee_id}'s reimbursement request has received a comment.",
                        verb_ar=f"تلقى طلب استرداد نفقات {payroll.employee_id} تعليقًا.",
                        verb_de=f"{payroll.employee_id}s Rückerstattungsantrag hat einen Kommentar erhalten.",
                        verb_es=f"La solicitud de reembolso de gastos de {payroll.employee_id} ha recibido un comentario.",
                        verb_fr=f"La demande de remboursement de frais de {payroll.employee_id} a reçu un commentaire.",
                        redirect=reverse("view-reimbursement"),
                        icon="chatbox-ellipses",
                    )
            else:
                rec = payroll.employee_id.employee_user_id
                notify.send(
                    request.user.employee_get,
                    recipient=rec,
                    verb="Your reimbursement request has received a comment.",
                    verb_ar="تلقى طلب استرداد نفقاتك تعليقًا.",
                    verb_de="Ihr Rückerstattungsantrag hat einen Kommentar erhalten.",
                    verb_es="Tu solicitud de reembolso ha recibido un comentario.",
                    verb_fr="Votre demande de remboursement a reçu un commentaire.",
                    redirect=reverse("view-reimbursement"),
                    icon="chatbox-ellipses",
                )

            return render(
                request,
                "payroll/reimbursement/reimbursement_comment.html",
                {
                    "comments": comments,
                    "no_comments": no_comments,
                    "request_id": payroll_id,
                },
            )
    return render(
        request,
        "payroll/reimbursement/reimbursement_comment.html",
        {"form": form, "request_id": payroll_id},
    )


@login_required
@hx_request_required
def view_payrollrequest_comment(request, payroll_id):
    """
    This method is used to show Reimbursement request comments
    """
    comments = ReimbursementrequestComment.objects.filter(
        request_id=payroll_id
    ).order_by("-created_at")

    req = Reimbursement.objects.get(id=payroll_id)
    no_comments = False
    if not comments.exists():
        no_comments = True

    if request.FILES:
        files = request.FILES.getlist("files")
        comment_id = request.GET["comment_id"]
        comment = ReimbursementrequestComment.objects.get(id=comment_id)
        attachments = []
        for file in files:
            file_instance = ReimbursementFile()
            file_instance.file = file
            file_instance.save()
            attachments.append(file_instance)
        comment.files.add(*attachments)
    return render(
        request,
        "payroll/reimbursement/reimbursement_comment.html",
        {
            "comments": comments,
            "no_comments": no_comments,
            "request_id": payroll_id,
            "req": req,
        },
    )


@login_required
def delete_payrollrequest_comment(request, comment_id):
    """
    This method is used to delete Reimbursement request comments
    """
    script = ""
    comment = ReimbursementrequestComment.objects.filter(id=comment_id)
    comment.delete()
    messages.success(request, _("Comment deleted successfully!"))
    return HttpResponse(script)


@login_required
def delete_reimbursement_comment_file(request):
    """
    Used to delete attachment
    """
    script = ""
    ids = request.GET.getlist("ids")
    records = ReimbursementFile.objects.filter(id__in=ids)
    if not request.user.has_perm("payroll.delete_reimbursmentfile"):
        records = records.filter(employee_id__employee_user_id=request.user)
    records.delete()
    messages.success(request, _("File deleted successfully"))
    return HttpResponse(script)


@login_required
@permission_required("payroll.add_payrollgeneralsetting")
def initial_notice_period(request):
    """
    This method is used to set initial value notice period
    """
    notice_period = eval_validate(request.GET["notice_period"])
    settings = PayrollGeneralSetting.objects.first()
    settings = settings if settings else PayrollGeneralSetting()
    settings.notice_period = max(notice_period, 0)
    settings.save()
    messages.success(
        request, _("The initial notice period has been successfully updated.")
    )
    if request.META.get("HTTP_HX_REQUEST"):
        return HttpResponse()
    return HttpResponseRedirect(request.META.get("HTTP_REFERER", "/"))


# ===========================Auto payslip generate================================


@login_required
@permission_required("payroll.view_PayslipAutoGenerate")
def auto_payslip_settings_view(request):
    payslip_auto_generate = PayslipAutoGenerate.objects.all()

    context = {"payslip_auto_generate": payslip_auto_generate}
    return render(request, "payroll/settings/auto_payslip_settings.html", context)


@login_required
@hx_request_required
@permission_required("payroll.change_PayslipAutoGenerate")
def create_or_update_auto_payslip(request, auto_id=None):
    auto_payslip = None
    if auto_id:
        auto_payslip = PayslipAutoGenerate.objects.get(id=auto_id)
    form = PayslipAutoGenerateForm(instance=auto_payslip)
    if request.method == "POST":
        form = PayslipAutoGenerateForm(request.POST, instance=auto_payslip)
        if form.is_valid():
            auto_payslip = form.save()
            company = (
                auto_payslip.company_id if auto_payslip.company_id else "All company"
            )
            messages.success(
                request, _(f"Payslip Auto generate for {company} created successfully ")
            )
            return HttpResponse("<script>window.location.reload()</script>")
    return render(
        request, "payroll/settings/auto_payslip_create_or_update.html", {"form": form}
    )


@login_required
@permission_required("payroll.change_PayslipAutoGenerate")
def activate_auto_payslip_generate(request):
    """
    ajax function to update is active field in PayslipAutoGenerate.
    Args:
    - isChecked: Boolean value representing the state of PayslipAutoGenerate,
    - autoId: Id of PayslipAutoGenerate object
    """
    isChecked = request.POST.get("isChecked")
    autoId = request.POST.get("autoId")
    payslip_auto = PayslipAutoGenerate.objects.get(id=autoId)
    if isChecked == "true":
        payslip_auto.auto_generate = True
        response = {
            "type": "success",
            "message": _("Auto paslip generate activated successfully."),
        }
    else:
        payslip_auto.auto_generate = False
        response = {
            "type": "success",
            "message": _("Auto paslip generate deactivated successfully."),
        }
    payslip_auto.save()
    return JsonResponse(response)


@login_required
@hx_request_required
@permission_required("payroll.delete_PayslipAutoGenerate")
def delete_auto_payslip(request, auto_id):
    """
    Delete a PayslipAutoGenerate object.

    Args:
        auto_id: The ID of PayslipAutoGenerate object to delete.

    Returns:
        Redirects to the contract view after successfully deleting the contract.

    """
    try:
        auto_payslip = PayslipAutoGenerate.objects.get(id=auto_id)
        if not auto_payslip.auto_generate:
            company = (
                auto_payslip.company_id if auto_payslip.company_id else "All company"
            )
            auto_payslip.delete()
            messages.success(
                request, _(f"Payslip auto generate for {company} deleted successfully.")
            )
        else:
            messages.info(request, _(f"Active 'Payslip auto generate' cannot delete."))
        return HttpResponse("<script>window.location.reload();</script>")
    except PayslipAutoGenerate.DoesNotExist:
        messages.error(request, _("Payslip auto generate not found."))
    return HttpResponseRedirect(request.META.get("HTTP_REFERER", "/"))


@login_required
@permission_required("payroll.view_contract")
def employee_payroll_summary(request):
    """
    Employee Payroll Summary view combining payroll, leave, and deductions data
    """
    from django.db.models import Q, Sum, F
    from datetime import date, datetime, timedelta
    from base.methods import paginator_qry
    from django.apps import apps
    
    # Get filter parameters
    search = request.GET.get('search')
    employee_id = request.GET.get('employee_id')
    department_id = request.GET.get('department_id')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    month = request.GET.get('month')
    year = request.GET.get('year', str(datetime.now().year))
    period = request.GET.get('period')  # Format: "YYYY-MM"
    
    # Parse period parameter if provided
    if period:
        try:
            period_parts = period.split('-')
            if len(period_parts) == 2:
                year = int(period_parts[0])
                month = int(period_parts[1])
        except (ValueError, IndexError):
            pass
    
    # Convert to integers if they exist
    try:
        month = int(month) if month else None
        year = int(year) if year else None
    except (ValueError, TypeError):
        month = None
        year = None
    
    # Set default to current month if no filters provided
    if not month and not start_date and not end_date:
        month = datetime.now().month
        year = datetime.now().year
    
    # Build base queryset for employees with contracts, ordered by ID
    employees = Employee.objects.filter(
        contract_set__isnull=False,
        is_active=True
    ).distinct().order_by('id')
    
    # Apply filters
    if search:
        employees = employees.filter(
            Q(employee_first_name__icontains=search) |
            Q(employee_last_name__icontains=search) |
            Q(badge_id__icontains=search) |
            Q(id__icontains=search)
        )
    if employee_id:
        employees = employees.filter(id=employee_id)
    if department_id:
        employees = employees.filter(employee_work_info__department_id=department_id)
    
    # Prepare date filters for payslips and leaves
    date_filter = Q()
    if start_date and end_date:
        date_filter = Q(
            payslip__start_date__gte=start_date,
            payslip__end_date__lte=end_date
        )
    elif month and year:
        date_filter = Q(
            payslip__start_date__month=month,
            payslip__start_date__year=year
        )
    elif year:
        date_filter = Q(payslip__start_date__year=year)
    
    # Get employee summary data
    employee_summaries = []
    for employee in employees:
        # Get active contract
        active_contract = employee.contract_set.filter(contract_status='active').first()
        if not active_contract:
            continue
            
        # Get work information
        work_info = employee.employee_work_info
        
        # Get latest payslip for the period
        payslip_filter = Q(employee_id=employee)
        if start_date and end_date:
            payslip_filter &= Q(start_date__gte=start_date, end_date__lte=end_date)
        elif month and year:
            payslip_filter &= Q(start_date__month=month, start_date__year=year)
        elif year:
            payslip_filter &= Q(start_date__year=year)
            
        latest_payslip = Payslip.objects.filter(payslip_filter).order_by('-end_date').first()
        
        # Get leave information
        leave_summary = {}
        from django.apps import apps
        if apps.is_installed("leave"):
            from leave.models import AvailableLeave, LeaveRequest
            
            # Get available leave balances
            available_leaves = AvailableLeave.objects.filter(employee_id=employee)
            leave_balances = {}
            for avail_leave in available_leaves:
                leave_balances[avail_leave.leave_type_id.name] = {
                    'available': avail_leave.available_days,
                    'carryforward': avail_leave.carryforward_days,
                    'total': avail_leave.total_leave_days,
                    'is_paid': avail_leave.leave_type_id.payment == 'paid'
                }
            
            # Get leave requests for the period
            leave_filter = Q(employee_id=employee, status='approved')
            if start_date and end_date:
                leave_filter &= Q(start_date__lte=end_date, end_date__gte=start_date)
            elif month and year:
                leave_filter &= Q(
                    start_date__month=month, start_date__year=year
                ) | Q(
                    end_date__month=month, end_date__year=year
                )
            elif year:
                leave_filter &= Q(start_date__year=year) | Q(end_date__year=year)
                
            leave_requests = LeaveRequest.objects.filter(leave_filter).select_related('leave_type_id')
            
            # Calculate paid vs unpaid leaves for the month
            paid_leaves = 0
            unpaid_leaves = 0
            for req in leave_requests:
                if req.leave_type_id.payment == 'paid':
                    paid_leaves += req.requested_days or 0
                else:
                    unpaid_leaves += req.requested_days or 0
            
            leave_summary = {
                'balances': leave_balances,
                'requests': leave_requests,
                'total_days_taken': sum([req.requested_days or 0 for req in leave_requests]),
                'paid_leaves': paid_leaves,
                'unpaid_leaves': unpaid_leaves
            }
        
        # Get attendance summary
        attendance_summary = {}
        if apps.is_installed("attendance"):
            from attendance.models import Attendance, AttendanceOverTime
            
            attendance_filter = Q(employee_id=employee)
            if start_date and end_date:
                attendance_filter &= Q(attendance_date__gte=start_date, attendance_date__lte=end_date)
            elif month and year:
                attendance_filter &= Q(attendance_date__month=month, attendance_date__year=year)
            elif year:
                attendance_filter &= Q(attendance_date__year=year)
                
            attendances = Attendance.objects.filter(attendance_filter, attendance_validated=True)
            
            total_worked_hours = sum([att.at_work_second for att in attendances if att.at_work_second])
            total_overtime_hours = sum([att.overtime_second for att in attendances if att.overtime_second])
            
            # Calculate total working days for the month
            total_working_days = 0
            if month and year:
                from calendar import monthrange
                import calendar
                # Get first and last day of month
                first_day = date(year, month, 1)
                last_day = date(year, month, monthrange(year, month)[1])
                
                # Count working days (excluding weekends)
                current_date = first_day
                while current_date <= last_day:
                    # Check if it's a weekday (Monday=0, Sunday=6)
                    if current_date.weekday() < 5:  # Monday to Friday
                        total_working_days += 1
                    current_date += timedelta(days=1)
            else:
                # If no month specified, use days present as working days
                total_working_days = attendances.count()
            
            # Get monthly overtime summary
            monthly_ot = AttendanceOverTime.objects.filter(
                employee_id=employee,
                month=datetime(int(year), int(month), 1).strftime('%B').lower() if month else None,
                year=str(year)
            ).first() if month and year else None
            
            attendance_summary = {
                'total_days_present': attendances.count(),
                'total_working_days': total_working_days,
                'total_worked_hours': total_worked_hours,
                'total_overtime_hours': total_overtime_hours,
                'monthly_overtime': monthly_ot.overtime_second if monthly_ot else 0
            }
        
        # Get bank details
        try:
            bank_details = employee.employee_bank_details
        except:
            bank_details = None
        
        # Calculate payroll summary
        payroll_summary = {
            'basic_pay': latest_payslip.basic_pay if latest_payslip else active_contract.wage or 0,
            'gross_pay': latest_payslip.gross_pay if latest_payslip else 0,
            'total_deductions': latest_payslip.deduction if latest_payslip else 0,
            'net_pay': latest_payslip.net_pay if latest_payslip else 0,
            'status': latest_payslip.status if latest_payslip else 'No payslip generated',
            'pay_period': f"{latest_payslip.start_date} to {latest_payslip.end_date}" if latest_payslip else 'N/A'
        }
        
        # Get allowances and deductions breakdown
        allowances_breakdown = []
        deductions_breakdown = []
        if latest_payslip and latest_payslip.pay_head_data:
            pay_data = latest_payslip.pay_head_data
            allowances_breakdown = pay_data.get('allowances', [])
            deductions_breakdown = []
            for deduction_list in [
                pay_data.get('basic_pay_deductions', []),
                pay_data.get('gross_pay_deductions', []),
                pay_data.get('pretax_deductions', []),
                pay_data.get('post_tax_deductions', []),
                pay_data.get('tax_deductions', []),
                pay_data.get('net_deductions', [])
            ]:
                deductions_breakdown.extend(deduction_list)
        
        employee_summary = {
            'employee': employee,
            'work_info': work_info,
            'contract': active_contract,
            'payroll_summary': payroll_summary,
            'leave_summary': leave_summary,
            'attendance_summary': attendance_summary,
            'bank_details': bank_details,
            'allowances_breakdown': allowances_breakdown,
            'deductions_breakdown': deductions_breakdown,
        }
        
        employee_summaries.append(employee_summary)
    
    # Paginate results
    employee_summaries = paginator_qry(employee_summaries, request.GET.get("page"))
    
    # Prepare context for filters
    from base.models import Department
    departments = Department.objects.all()

    # Generate months list
    months = [
        (1, 'January'), (2, 'February'), (3, 'March'), (4, 'April'),
        (5, 'May'), (6, 'June'), (7, 'July'), (8, 'August'),
        (9, 'September'), (10, 'October'), (11, 'November'), (12, 'December')
    ]

    # Generate year range (current year ± 5 years)
    current_year = datetime.now().year
    year_range = range(current_year - 5, current_year + 1)
    
    # Generate period options for dropdown
    period_options = []
    for year_option in year_range:
        for month_num, month_name in months:
            period_options.append({
                'value': f"{year_option}-{month_num:02d}",
                'label': f"{month_name} {year_option}",
                'year': year_option,
                'month': month_num
            })
    
    # Sort by year and month (newest first)
    period_options.sort(key=lambda x: (x['year'], x['month']), reverse=True)
    
    # Get current period info
    current_period = f"{year}-{month:02d}" if month else None
    current_period_label = f"{dict(months).get(int(month), '')} {year}" if month and year else None
    
    # Get previous months for history
    history_periods = []
    if month and year:
        for i in range(1, 6):  # Last 5 months
            prev_date = datetime(year, month, 1) - timedelta(days=30*i)
            history_periods.append({
                'year': prev_date.year,
                'month': prev_date.month,
                'label': f"{dict(months).get(prev_date.month, '')} {prev_date.year}",
                'value': f"{prev_date.year}-{prev_date.month:02d}"
            })

    # Calculate active filter count
    filter_count = 0
    if department_id:
        filter_count += 1
    if start_date or end_date:
        filter_count += 1
    if current_period:
        filter_count += 1

    context = {
        'employee_summaries': employee_summaries,
        'departments': departments,
        'months': months,
        'year_range': year_range,
        'period_options': period_options,
        'current_period': current_period,
        'current_period_label': current_period_label,
        'history_periods': history_periods,
        'filter_count': filter_count,
        'current_filters': {
            'search': search,
            'employee_id': employee_id,
            'department_id': department_id,
            'start_date': start_date,
            'end_date': end_date,
            'month': month,
            'year': year,
        }
    }

    # Handle export requests
    export_format = request.GET.get('export')
    if export_format:
        if export_format == 'pdf':
            return export_payroll_summary_pdf(employee_summaries, current_period_label, context)
        elif export_format == 'excel':
            return export_payroll_summary_excel(employee_summaries, current_period_label, context)
    
    return render(request, "payroll/payroll_summary/employee_payroll_summary.html", context)


@login_required
@permission_required("payroll.view_contract")
def employee_month_details(request):
    """
    AJAX endpoint to get employee attendance and payslip details for a specific month
    Returns JSON data
    """
    from django.http import JsonResponse
    from django.db.models import Q
    from datetime import date, datetime
    from calendar import monthrange
    import json
    
    employee_id = request.GET.get('employee_id')
    period = request.GET.get('period')  # Format: "YYYY-MM"
    
    if not employee_id or not period:
        return JsonResponse({'error': 'Missing required parameters'}, status=400)
    
    try:
        # Parse period
        period_parts = period.split('-')
        if len(period_parts) != 2:
            return JsonResponse({'error': 'Invalid period format'}, status=400)
        
        year = int(period_parts[0])
        month = int(period_parts[1])
        
        # Get employee
        employee = Employee.objects.get(id=employee_id)
        
        # Get first and last day of month
        first_day = date(year, month, 1)
        last_day = date(year, month, monthrange(year, month)[1])
        
        # Calculate total working days (excluding weekends)
        total_working_days = 0
        current_date = first_day
        while current_date <= last_day:
            if current_date.weekday() < 5:  # Monday to Friday
                total_working_days += 1
            current_date += timedelta(days=1)
        
        # Get attendance data
        attendance_data = {
            'total_working_days': total_working_days,
            'days_present': 0
        }
        
        from django.apps import apps
        if apps.is_installed("attendance"):
            from attendance.models import Attendance
            attendances = Attendance.objects.filter(
                employee_id=employee,
                attendance_date__year=year,
                attendance_date__month=month,
                attendance_validated=True
            )
            attendance_data['days_present'] = attendances.count()
        
        # Get leave data
        leave_data = {
            'paid_leaves': 0.0,
            'unpaid_leaves': 0.0
        }
        
        if apps.is_installed("leave"):
            from leave.models import LeaveRequest
            
            # Get approved leave requests for the month
            leave_requests = LeaveRequest.objects.filter(
                employee_id=employee,
                status='approved'
            ).filter(
                Q(start_date__month=month, start_date__year=year) |
                Q(end_date__month=month, end_date__year=year) |
                Q(start_date__lte=first_day, end_date__gte=last_day)
            ).select_related('leave_type_id')
            
            for req in leave_requests:
                # Calculate days that fall within this month
                leave_start = max(req.start_date, first_day)
                leave_end = min(req.end_date or req.start_date, last_day)
                
                if leave_start <= leave_end:
                    # Count working days in the leave period within this month
                    leave_days = 0
                    check_date = leave_start
                    while check_date <= leave_end:
                        if check_date.weekday() < 5:  # Weekdays only
                            leave_days += 1
                        check_date += timedelta(days=1)
                    
                    if req.leave_type_id.payment == 'paid':
                        leave_data['paid_leaves'] += leave_days
                    else:
                        leave_data['unpaid_leaves'] += leave_days
        
        # Get payslip data - read directly from payslip model and pay_head_data
        # No calculations - just read what's stored
        payslip_data = {
            'paid_days': 0,
            'working_days': 0,
            'basic_pay': 0.0,
            'net_pay': 0.0,
            'status': 'No Payslip Generated'
        }
        
        # Find payslip for this month
        payslip = Payslip.objects.filter(
            employee_id=employee,
            start_date__year=year,
            start_date__month=month
        ).order_by('-end_date').first()
        
        # Calculate paid_days from actual attendance (not from payslip)
        # Paid Days = Days Present + Paid Leaves
        payslip_data['paid_days'] = attendance_data['days_present'] + leave_data['paid_leaves']
        
        if payslip:
            # Read directly from payslip model fields
            payslip_data['status'] = payslip.status or 'draft'
            payslip_data['basic_pay'] = float(payslip.basic_pay or 0)
            payslip_data['net_pay'] = float(payslip.net_pay or 0)
            
            # Read working_days from pay_head_data (JSON field)
            if payslip.pay_head_data:
                pay_data = payslip.pay_head_data
                payslip_data['working_days'] = float(pay_data.get('total_working_days', total_working_days))
            else:
                payslip_data['working_days'] = total_working_days
        else:
            # No payslip - get basic pay from contract and use calculated working days
            payslip_data['working_days'] = total_working_days
            active_contract = employee.contract_set.filter(contract_status='active').first()
            if active_contract:
                payslip_data['basic_pay'] = float(active_contract.wage or 0)
        
        return JsonResponse({
            'attendance': attendance_data,
            'leaves': leave_data,
            'payslip': payslip_data
        })
        
    except Employee.DoesNotExist:
        return JsonResponse({'error': 'Employee not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def export_payroll_summary_pdf(employee_summaries, period_label, context):
    """
    Export payroll summary to PDF
    """
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    import io
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="payroll_summary_{period_label.replace(" ", "_")}.pdf"'
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=12,
        spaceAfter=12,
        textColor=colors.darkblue
    )
    
    # Build PDF content
    story = []
    
    # Title
    title = Paragraph(f"Employee Payroll Summary - {period_label}", title_style)
    story.append(title)
    story.append(Spacer(1, 20))
    
    # Employee summaries
    for summary in employee_summaries:
        # Employee header
        emp_header = f"{summary['employee'].get_full_name()} (ID: {summary['employee'].badge_id or 'N/A'})"
        story.append(Paragraph(emp_header, heading_style))
        
        # Safely get department and position
        department = 'N/A'
        position = 'N/A'
        if summary.get('work_info'):
            try:
                department = summary['work_info'].department_id.department if summary['work_info'].department_id else 'N/A'
            except AttributeError:
                department = 'N/A'
            try:
                position = summary['work_info'].job_position_id.job_position if summary['work_info'].job_position_id else 'N/A'
            except AttributeError:
                position = 'N/A'
        
        # Get phone number
        phone = summary['employee'].phone or 'N/A'
        
        # Create table data
        table_data = [
            ['Field', 'Value'],
            ['Department', department],
            ['Phone Number', phone],
            ['Position', position],
            ['Gross Salary', f"₹{summary['payroll_summary']['gross_pay']:.2f}"],
            ['Total Deductions', f"₹{summary['payroll_summary']['total_deductions']:.2f}"],
            ['Net Salary', f"₹{summary['payroll_summary']['net_pay']:.2f}"],
            ['Status', summary['payroll_summary']['status'].title()],
        ]
        
        # Create table
        table = Table(table_data, colWidths=[2*inch, 3*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        story.append(Spacer(1, 20))
    
    # Build PDF
    doc.build(story)
    pdf_content = buffer.getvalue()
    buffer.close()
    
    response.write(pdf_content)
    return response


def export_payroll_summary_excel(employee_summaries, period_label, context):
    """
    Export payroll summary to Excel
    """
    from django.http import HttpResponse
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="payroll_summary_{period_label.replace(" ", "_")}.xlsx"'
    
    # Create workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = f"Payroll Summary {period_label}"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'Employee ID', 'Employee Name', 'Department', 'Phone Number', 'Position',
        'Gross Salary', 'Total Deductions', 'Net Salary', 'Status',
        'Pay Period', 'Bank Name', 'Account Number'
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Write data
    for row, summary in enumerate(employee_summaries, 2):
        # Safely get department and position
        department = 'N/A'
        position = 'N/A'
        if summary.get('work_info'):
            try:
                department = summary['work_info'].department_id.department if summary['work_info'].department_id else 'N/A'
            except AttributeError:
                department = 'N/A'
            try:
                position = summary['work_info'].job_position_id.job_position if summary['work_info'].job_position_id else 'N/A'
            except AttributeError:
                position = 'N/A'
        
        # Safely get bank details
        bank_name = 'N/A'
        account_number = 'N/A'
        if summary.get('bank_details'):
            try:
                bank_name = summary['bank_details'].bank_name or 'N/A'
                account_number = summary['bank_details'].account_number or 'N/A'
            except AttributeError:
                bank_name = 'N/A'
                account_number = 'N/A'
        
        # Get phone number
        phone = summary['employee'].phone or 'N/A'
        
        data = [
            summary['employee'].id,
            summary['employee'].get_full_name(),
            department,
            phone,
            position,
            summary['payroll_summary']['gross_pay'],
            summary['payroll_summary']['total_deductions'],
            summary['payroll_summary']['net_pay'],
            summary['payroll_summary']['status'].title(),
            summary['payroll_summary']['pay_period'],
            bank_name,
            account_number
        ]
        
        for col, value in enumerate(data, 1):
            cell = worksheet.cell(row=row, column=col, value=value)
            cell.border = border
            if col in [5, 6, 7]:  # Currency columns
                cell.number_format = '₹#,##0.00'
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save to response
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response.write(buffer.getvalue())
    buffer.close()
    
    return response


@login_required
@permission_required("payroll.view_contract")
def employee_overview(request, employee_id):
    """
    Employee Overview page with employee info and salary summary
    """
    from datetime import datetime, timedelta
    from django.db.models import Q, Sum, F
    
    try:
        employee = Employee.objects.get(id=employee_id)
    except Employee.DoesNotExist:
        return render(request, "payroll/payroll_summary/employee_not_found.html")
    
    # Get work information
    work_info = employee.employee_work_info
    
    # Get active contract
    active_contract = employee.contract_set.filter(contract_status='active').first()
    
    # Get current period payslip
    current_month = datetime.now().month
    current_year = datetime.now().year
    
    latest_payslip = Payslip.objects.filter(
        employee_id=employee,
        start_date__month=current_month,
        start_date__year=current_year
    ).order_by('-end_date').first()
    
    # Get bank details
    try:
        bank_details = employee.employee_bank_details
    except:
        bank_details = None
    
    # Calculate payroll summary
    payroll_summary = {
        'basic_pay': latest_payslip.basic_pay if latest_payslip else active_contract.wage or 0,
        'gross_pay': latest_payslip.gross_pay if latest_payslip else 0,
        'total_deductions': latest_payslip.deduction if latest_payslip else 0,
        'net_pay': latest_payslip.net_pay if latest_payslip else 0,
        'status': latest_payslip.status if latest_payslip else 'No payslip generated',
        'pay_period': f"{latest_payslip.start_date} to {latest_payslip.end_date}" if latest_payslip else 'N/A'
    }
    
    context = {
        'employee': employee,
        'work_info': work_info,
        'contract': active_contract,
        'payroll_summary': payroll_summary,
        'bank_details': bank_details,
        'current_month': current_month,
        'current_year': current_year,
    }
    
    return render(request, "payroll/payroll_summary/employee_overview.html", context)


@login_required
@permission_required("payroll.view_contract")
def employee_leave_attendance(request, employee_id):
    """
    Employee Leave & Attendance page with complete monthly data
    """
    from datetime import datetime, timedelta
    from django.db.models import Q, Sum, F
    from django.apps import apps
    
    try:
        employee = Employee.objects.get(id=employee_id)
    except Employee.DoesNotExist:
        return render(request, "payroll/payroll_summary/employee_not_found.html")
    
    # Get filter parameters
    year = request.GET.get('year', datetime.now().year)
    month = request.GET.get('month', datetime.now().month)
    
    # Generate months list
    months = [
        (1, 'January'), (2, 'February'), (3, 'March'), (4, 'April'),
        (5, 'May'), (6, 'June'), (7, 'July'), (8, 'August'),
        (9, 'September'), (10, 'October'), (11, 'November'), (12, 'December')
    ]
    
    # Generate year range
    current_year = datetime.now().year
    year_range = range(current_year - 5, current_year + 1)
    
    # Get leave information for the selected month
    leave_summary = {}
    if apps.is_installed("leave"):
        from leave.models import AvailableLeave, LeaveRequest
        
        # Get available leave balances
        available_leaves = AvailableLeave.objects.filter(employee_id=employee)
        leave_balances = {}
        for avail_leave in available_leaves:
            leave_balances[avail_leave.leave_type_id.name] = {
                'available': avail_leave.available_days,
                'carryforward': avail_leave.carryforward_days,
                'total': avail_leave.total_leave_days,
                'is_paid': avail_leave.leave_type_id.payment == 'paid'
            }
        
        # Get leave requests for the selected month
        leave_filter = Q(employee_id=employee, status='approved')
        leave_filter &= Q(
            start_date__month=month, start_date__year=year
        ) | Q(
            end_date__month=month, end_date__year=year
        )
        
        leave_requests = LeaveRequest.objects.filter(leave_filter)
        
        # Get all leave requests for history
        all_leave_requests = LeaveRequest.objects.filter(
            employee_id=employee, status='approved'
        ).order_by('-start_date')[:12]  # Last 12 months
        
        leave_summary = {
            'balances': leave_balances,
            'requests': leave_requests,
            'all_requests': all_leave_requests,
            'total_days_taken': sum([req.requested_days for req in leave_requests])
        }
    
    # Get attendance summary for the selected month
    attendance_summary = {}
    monthly_attendance_data = []
    
    if apps.is_installed("attendance"):
        from attendance.models import Attendance, AttendanceOverTime
        
        # Get payslip data first to get working days and working dates
        payslip = None
        total_working_days = 30  
        working_dates = []
        
        try:
            payslip = Payslip.objects.filter(
                employee_id=employee,
                start_date__month=month,
                start_date__year=year
            ).first()
            
            if payslip and payslip.pay_head_data:
                pay_data = payslip.pay_head_data
                paid_days = pay_data.get('paid_days', 0)
                lop_days = pay_data.get('unpaid_days', 0)
                # Total company working days = paid days + LOP days
                total_working_days = paid_days + lop_days
                
                # Get working dates if available
                if 'working_dates' in pay_data:
                    working_dates = pay_data.get('working_dates', [])
        except:
            pass
        
        # Get attendance for selected month - only on working days if available
        if working_dates:
            # Use exact working dates from payslip
            attendances = Attendance.objects.filter(
                employee_id=employee,
                attendance_date__month=month,
                attendance_date__year=year,
                attendance_validated=True,
                attendance_date__in=working_dates
            )
        else:
            # Fallback: get all attendance if no working dates available
            attendances = Attendance.objects.filter(
                employee_id=employee,
                attendance_date__month=month,
                attendance_date__year=year,
                attendance_validated=True
            )
        
        total_worked_hours = sum([att.at_work_second for att in attendances if att.at_work_second])
        total_overtime_hours = sum([att.overtime_second for att in attendances if att.overtime_second])
        
        # Get monthly attendance data for last 12 months
        for i in range(12):
            target_date = datetime(int(year), int(month), 1) - timedelta(days=30*i)
            # Get payslip data for this month
            month_payslip = None
            month_total_working_days = 30  # Default fallback (company uses 30 days formula)
            month_working_dates = []
            
            try:
                month_payslip = Payslip.objects.filter(
                    employee_id=employee,
                    start_date__month=target_date.month,
                    start_date__year=target_date.year
                ).first()
                
                if month_payslip and month_payslip.pay_head_data:
                    month_pay_data = month_payslip.pay_head_data
                    month_paid_days = month_pay_data.get('paid_days', 0)
                    month_lop_days = month_pay_data.get('unpaid_days', 0)
                    # Total company working days = paid days + LOP days
                    month_total_working_days = month_paid_days + month_lop_days
                    
                    # Get working dates if available
                    if 'working_dates' in month_pay_data:
                        month_working_dates = month_pay_data.get('working_dates', [])
            except:
                pass
            
            # Get attendance for this month - only on working days if available
            if month_working_dates:
                # Use exact working dates from payslip
                month_attendances = Attendance.objects.filter(
                    employee_id=employee,
                    attendance_date__month=target_date.month,
                    attendance_date__year=target_date.year,
                    attendance_validated=True,
                    attendance_date__in=month_working_dates
                )
            else:
                # Fallback: get all attendance if no working dates available
                month_attendances = Attendance.objects.filter(
                    employee_id=employee,
                    attendance_date__month=target_date.month,
                    attendance_date__year=target_date.year,
                    attendance_validated=True
                )
            
            monthly_attendance_data.append({
                'month': target_date.month,
                'year': target_date.year,
                'month_name': dict(months)[target_date.month],
                'days_present': month_attendances.count(),
                'total_working_days': month_total_working_days,
                'total_hours': sum([att.at_work_second for att in month_attendances if att.at_work_second]),
                'overtime_hours': sum([att.overtime_second for att in month_attendances if att.overtime_second])
            })
        
        attendance_summary = {
            'total_days_present': attendances.count(),
            'total_working_days': total_working_days,
            'total_worked_hours': total_worked_hours,
            'total_overtime_hours': total_overtime_hours,
            'monthly_data': monthly_attendance_data
        }
    
    context = {
        'employee': employee,
        'leave_summary': leave_summary,
        'attendance_summary': attendance_summary,
        'months': months,
        'year_range': year_range,
        'current_month': int(month),
        'current_year': int(year),
        'current_month_name': dict(months)[int(month)],
    }
    
    return render(request, "payroll/payroll_summary/employee_leave_attendance.html", context)


@login_required
@permission_required("payroll.view_contract")
def employee_payslip_breakdown(request, employee_id):
    """
    Employee Payslip Breakdown page with complete monthly payslip data
    """
    from datetime import datetime, timedelta
    from django.db.models import Q, Sum, F
    
    try:
        employee = Employee.objects.get(id=employee_id)
    except Employee.DoesNotExist:
        return render(request, "payroll/payroll_summary/employee_not_found.html")
    
    # Get filter parameters
    year = request.GET.get('year', datetime.now().year)
    month = request.GET.get('month', datetime.now().month)
    
    # Generate months list
    months = [
        (1, 'January'), (2, 'February'), (3, 'March'), (4, 'April'),
        (5, 'May'), (6, 'June'), (7, 'July'), (8, 'August'),
        (9, 'September'), (10, 'October'), (11, 'November'), (12, 'December')
    ]
    
    # Generate year range
    current_year = datetime.now().year
    year_range = range(current_year - 5, current_year + 1)
    
    # Get active contract
    active_contract = employee.contract_set.filter(contract_status='active').first()
    
    # Get payslip for selected month
    payslip = Payslip.objects.filter(
        employee_id=employee,
        start_date__month=month,
        start_date__year=year
    ).order_by('-end_date').first()
    
    # Get all payslips for history (last 12 months)
    monthly_payslips = []
    for i in range(12):
        target_date = datetime(int(year), int(month), 1) - timedelta(days=30*i)
        month_payslip = Payslip.objects.filter(
            employee_id=employee,
            start_date__month=target_date.month,
            start_date__year=target_date.year
        ).order_by('-end_date').first()
        
        if month_payslip:
            # Calculate LOP for this month
            month_lop_days = 0
            month_lop_amount = 0
            
            if month_payslip.pay_head_data:
                month_pay_data = month_payslip.pay_head_data
                
                # First, check if unpaid_days is directly stored in pay_head_data
                if 'unpaid_days' in month_pay_data:
                    month_lop_days = month_pay_data.get('unpaid_days', 0)
                    month_lop_amount = month_pay_data.get('loss_of_pay', 0)
                # If not found, check if loss_of_pay is directly stored in pay_head_data
                elif 'loss_of_pay' in month_pay_data:
                    month_lop_amount = month_pay_data.get('loss_of_pay', 0)
                    # Calculate LOP days based on actual basic pay per day (not the reduced basic pay)
                    # Get the original basic pay before LOP deduction
                    actual_basic_pay = month_pay_data.get('actual_basic_pay', month_payslip.basic_pay + month_lop_amount)
                    if actual_basic_pay > 0:
                        daily_rate = actual_basic_pay / 30  # Assuming 30 days per month
                        month_lop_days = round(month_lop_amount / daily_rate, 1) if daily_rate > 0 else 0
                
                # If not found directly, look for LOP in all deduction categories
                if month_lop_amount == 0:
                    all_deduction_lists = [
                        month_pay_data.get('basic_pay_deductions', []),
                        month_pay_data.get('gross_pay_deductions', []),
                        month_pay_data.get('pretax_deductions', []),
                        month_pay_data.get('post_tax_deductions', []),
                        month_pay_data.get('tax_deductions', []),
                        month_pay_data.get('net_deductions', [])
                    ]
                    
                    # Also check all other possible keys in pay_head_data
                    for key, value in month_pay_data.items():
                        if isinstance(value, list) and 'deduction' in key.lower():
                            all_deduction_lists.append(value)
                    
                    # Search through all deduction lists
                    for deduction_list in all_deduction_lists:
                        for deduction in deduction_list:
                            deduction_title = deduction.get('title', '').lower()
                            # More comprehensive LOP detection
                            lop_keywords = ['lop', 'loss of pay', 'loss of pay days', 'lop days', 'loss', 'absent']
                            if any(keyword in deduction_title for keyword in lop_keywords):
                                month_lop_amount = deduction.get('amount', 0)
                                # Calculate LOP days based on actual basic pay per day (not the reduced basic pay)
                                # Get the original basic pay before LOP deduction
                                actual_basic_pay = month_pay_data.get('actual_basic_pay', month_payslip.basic_pay + month_lop_amount)
                                if actual_basic_pay > 0:
                                    daily_rate = actual_basic_pay / 30  # Assuming 30 days per month
                                    month_lop_days = round(month_lop_amount / daily_rate, 1) if daily_rate > 0 else 0
                                break
                        if month_lop_amount > 0:
                            break
            
            monthly_payslips.append({
                'month': target_date.month,
                'year': target_date.year,
                'month_name': dict(months)[target_date.month],
                'payslip': month_payslip,
                'basic_pay': month_payslip.basic_pay,
                'gross_pay': month_payslip.gross_pay,
                'total_deductions': month_payslip.deduction,
                'net_pay': month_payslip.net_pay,
                'status': month_payslip.status,
                'allowances': month_payslip.pay_head_data.get('allowances', []) if month_payslip.pay_head_data else [],
                'lop_days': month_lop_days,
                'lop_amount': month_lop_amount,
                'deductions': []
            })
            
            # Get deductions breakdown
            if month_payslip.pay_head_data:
                pay_data = month_payslip.pay_head_data
                deductions = []
                for deduction_list in [
                    pay_data.get('basic_pay_deductions', []),
                    pay_data.get('gross_pay_deductions', []),
                    pay_data.get('pretax_deductions', []),
                    pay_data.get('post_tax_deductions', []),
                    pay_data.get('tax_deductions', []),
                    pay_data.get('net_deductions', [])
                ]:
                    deductions.extend(deduction_list)
                monthly_payslips[-1]['deductions'] = deductions
    
    # Calculate LOP days and amount
    lop_days = 0
    lop_amount = 0
    
    if payslip and payslip.pay_head_data:
        pay_data = payslip.pay_head_data
        
        # First, check if unpaid_days is directly stored in pay_head_data
        if 'unpaid_days' in pay_data:
            lop_days = pay_data.get('unpaid_days', 0)
            lop_amount = pay_data.get('loss_of_pay', 0)
        # If not found, check if loss_of_pay is directly stored in pay_head_data
        elif 'loss_of_pay' in pay_data:
            lop_amount = pay_data.get('loss_of_pay', 0)
            # Calculate LOP days based on actual basic pay per day (not the reduced basic pay)
            # Get the original basic pay before LOP deduction
            actual_basic_pay = pay_data.get('actual_basic_pay', payslip.basic_pay + lop_amount)
            if actual_basic_pay > 0:
                daily_rate = actual_basic_pay / 30  # Assuming 30 days per month
                lop_days = round(lop_amount / daily_rate, 1) if daily_rate > 0 else 0
        
        # If not found directly, look for LOP in all deduction categories
        if lop_amount == 0:
            all_deduction_lists = [
                pay_data.get('basic_pay_deductions', []),
                pay_data.get('gross_pay_deductions', []),
                pay_data.get('pretax_deductions', []),
                pay_data.get('post_tax_deductions', []),
                pay_data.get('tax_deductions', []),
                pay_data.get('net_deductions', [])
            ]
            
            # Also check all other possible keys in pay_head_data
            for key, value in pay_data.items():
                if isinstance(value, list) and 'deduction' in key.lower():
                    all_deduction_lists.append(value)
            
            # Search through all deduction lists
            for deduction_list in all_deduction_lists:
                for deduction in deduction_list:
                    deduction_title = deduction.get('title', '').lower()
                    # More comprehensive LOP detection
                    lop_keywords = ['lop', 'loss of pay', 'loss of pay days', 'lop days', 'loss', 'absent']
                    if any(keyword in deduction_title for keyword in lop_keywords):
                        lop_amount = deduction.get('amount', 0)
                        # Calculate LOP days based on actual basic pay per day (not the reduced basic pay)
                        # Get the original basic pay before LOP deduction
                        actual_basic_pay = pay_data.get('actual_basic_pay', payslip.basic_pay + lop_amount)
                        if actual_basic_pay > 0:
                            daily_rate = actual_basic_pay / 30  # Assuming 30 days per month
                            lop_days = round(lop_amount / daily_rate, 1) if daily_rate > 0 else 0
                        break
                if lop_amount > 0:
                    break
    
    # Current payslip breakdown
    current_payslip_data = {
        'payslip': payslip,
        'basic_pay': payslip.basic_pay if payslip else active_contract.wage or 0,
        'gross_pay': payslip.gross_pay if payslip else 0,
        'total_deductions': payslip.deduction if payslip else 0,
        'net_pay': payslip.net_pay if payslip else 0,
        'status': payslip.status if payslip else 'No payslip generated',
        'allowances': payslip.pay_head_data.get('allowances', []) if payslip and payslip.pay_head_data else [],
        'lop_days': lop_days,
        'lop_amount': lop_amount,
        'deductions': []
    }
    
    # Get deductions breakdown for current payslip
    if payslip and payslip.pay_head_data:
        pay_data = payslip.pay_head_data
        deductions = []
        for deduction_list in [
            pay_data.get('basic_pay_deductions', []),
            pay_data.get('gross_pay_deductions', []),
            pay_data.get('pretax_deductions', []),
            pay_data.get('post_tax_deductions', []),
            pay_data.get('tax_deductions', []),
            pay_data.get('net_deductions', [])
        ]:
            deductions.extend(deduction_list)
        current_payslip_data['deductions'] = deductions
    
    # Debug: Get all deduction titles for troubleshooting
    debug_deduction_titles = []
    if payslip and payslip.pay_head_data:
        pay_data = payslip.pay_head_data
        for key, value in pay_data.items():
            if isinstance(value, list):
                for item in value:
                    if isinstance(item, dict) and 'title' in item:
                        debug_deduction_titles.append(f"{key}: {item.get('title', '')} - ₹{item.get('amount', 0)}")

    context = {
        'employee': employee,
        'current_payslip_data': current_payslip_data,
        'monthly_payslips': monthly_payslips,
        'months': months,
        'year_range': year_range,
        'current_month': int(month),
        'current_year': int(year),
        'current_month_name': dict(months)[int(month)],
        'debug_deduction_titles': debug_deduction_titles,  # For debugging LOP detection
    }
    
    return render(request, "payroll/payroll_summary/employee_payslip_breakdown.html", context)